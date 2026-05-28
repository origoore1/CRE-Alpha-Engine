"""
rent_roll_parser.py  v1.1
──────────────────────────
Institutional-Grade Rent Roll Parser for German Retail CRE

Changes vs v1.0
───────────────
• load_cost_sheet()     — reads itemised Cost Sheet; returns NOI breakdown dict
• compute_metrics()     — NOI now comes from Cost Sheet (itemised) not 80% proxy;
                          void costs (service charge + insurance) modelled separately;
                          noi_source flag tells caller which path was used
• compute_metrics() sig — accepts optional costs dict (from load_cost_sheet)
• analyse()             — passes costs through the full pipeline
"""

import pandas as pd
import numpy as np
from datetime import date
from typing import Optional


# =============================================================================
# 1. EXCEL READERS
# =============================================================================

def load_deal_overview(filepath) -> dict:
    """Read Deal Overview tab. Accepts file path or file-like object."""
    if hasattr(filepath, "seek"):
        filepath.seek(0)
    raw = pd.read_excel(filepath, sheet_name="Deal Overview",
                        header=None, dtype=str)

    def cell(row, col):
        try:
            v = raw.iat[row, col]
            return None if pd.isna(v) else v
        except Exception:
            return None

    def num(row, col):
        v = cell(row, col)
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").replace("%", ""))
        except ValueError:
            return None

    overview = {
        "property_name": cell(4,  1),
        "city":          cell(6,  1),
        "asset_type":    cell(7,  1),
        "gla_total":     num(8,   1),
        "asking_price":  num(15,  1),
        "equity":        num(16,  1),
        "grest_rate":    num(19,  1),
        "notar_rate":    num(20,  1),
        "broker_rate":   num(21,  1),
        "interest_rate": num(27,  1),
        "loan_term_yrs": num(28,  1),
        "amort_rate":    num(30,  1),
    }

    if overview["asking_price"] and overview["equity"]:
        overview["loan_amount"] = overview["asking_price"] - overview["equity"]
    else:
        overview["loan_amount"] = None

    if overview["asking_price"]:
        rates = [overview.get("grest_rate") or 0,
                 overview.get("notar_rate")  or 0,
                 overview.get("broker_rate") or 0]
        overview["total_acquisition_cost"] = overview["asking_price"] * (1 + sum(rates))
        # Store asset deal cost separately for share deal comparison
        overview["acq_cost_asset_deal"]  = overview["total_acquisition_cost"]
        # Share deal: GrESt is 0 (structure below 90% threshold)
        # Notar and broker fees still apply
        share_rates = [0,
                        overview.get("notar_rate") or 0,
                        overview.get("broker_rate") or 0]
        overview["acq_cost_share_deal"] = overview["asking_price"] * (1 + sum(share_rates))
        overview["grest_saving"]        = overview["asking_price"] * (overview.get("grest_rate") or 0)
    else:
        overview["total_acquisition_cost"] = None
        overview["acq_cost_asset_deal"]    = None
        overview["acq_cost_share_deal"]    = None
        overview["grest_saving"]           = None

    return overview


def load_rent_roll(filepath) -> pd.DataFrame:
    """Read rows 4-103 of the Rent Roll tab."""
    if hasattr(filepath, "seek"):
        filepath.seek(0)
    df = pd.read_excel(
        filepath,
        sheet_name="Rent Roll",
        header=2,
        nrows=100,
        dtype={
            "Unit\nRef":                    str,
            "Tenant\nName":                 str,
            "Trade\nSector":                str,
            "Anchor?\n(Y/N)":               str,
            "CPI\nClause (Y/N)":            str,
            "Service Charge\nRecov. (Y/N)": str,
            "Notes /\nObservations":        str,
        }
    )

    rename = {
        "Unit\nRef":                    "unit_ref",
        "Tenant\nName":                 "tenant",
        "Trade\nSector":                "sector",
        "Anchor?\n(Y/N)":               "is_anchor",
        "Leased\nArea (m\xb2)":         "area_sqm",
        "Passing Rent\n\u20ac/m\xb2/mo":"rent_per_sqm_mo",
        "Annual\nRent (\u20ac)":         "annual_rent",
        "Lease\nStart":                 "lease_start",
        "Lease\nExpiry":                "lease_expiry",
        "Break\nOption Date":           "break_option",
        "Unexpired\nTerm (Yrs)":        "unexpired_yrs",
        "CPI\nClause (Y/N)":            "cpi_clause",
        "CPI\nPass-Through %":          "cpi_passthrough",
        "CPI\nTrigger %":               "cpi_trigger",
        "Security\nDeposit (\u20ac)":   "security_deposit",
        "Service Charge\nRecov. (Y/N)": "svc_chg_recoverable",
        "Rent-Free\nMonths":            "rent_free_months",
        "Notes /\nObservations":        "notes",
    }
    # Also handle ASCII versions of the column names
    rename_ascii = {
        "Unit\nRef":                    "unit_ref",
        "Tenant\nName":                 "tenant",
        "Trade\nSector":                "sector",
        "Anchor?\n(Y/N)":               "is_anchor",
        "Leased\nArea (m2)":            "area_sqm",
        "Passing Rent\nEUR/m2/mo":      "rent_per_sqm_mo",
        "Annual\nRent (EUR)":           "annual_rent",
        "Lease\nStart":                 "lease_start",
        "Lease\nExpiry":                "lease_expiry",
        "Break\nOption Date":           "break_option",
        "Unexpired\nTerm (Yrs)":        "unexpired_yrs",
        "CPI\nClause (Y/N)":            "cpi_clause",
        "CPI\nPass-Through %":          "cpi_passthrough",
        "CPI\nTrigger %":               "cpi_trigger",
        "Security\nDeposit (EUR)":      "security_deposit",
        "Service Charge\nRecov. (Y/N)": "svc_chg_recoverable",
        "Rent-Free\nMonths":            "rent_free_months",
        "Notes /\nObservations":        "notes",
    }
    full_rename = {**rename_ascii, **rename}
    df.rename(columns={k: v for k, v in full_rename.items() if k in df.columns},
              inplace=True)

    # Map remaining columns by position if name matching failed
    col_map = {
        0: "unit_ref", 1: "tenant", 2: "sector", 3: "is_anchor",
        4: "area_sqm", 5: "rent_per_sqm_mo", 6: "annual_rent",
        7: "lease_start", 8: "lease_expiry", 9: "break_option",
        10: "unexpired_yrs", 11: "cpi_clause", 12: "cpi_passthrough",
        13: "cpi_trigger", 14: "security_deposit",
        15: "svc_chg_recoverable", 16: "rent_free_months", 17: "notes",
    }
    if "tenant" not in df.columns:
        df.columns = [col_map.get(i, f"col_{i}") for i in range(len(df.columns))]

    df = df[df["tenant"].notna() & (df["tenant"].astype(str).str.strip() != "")].copy()

    for col in ["area_sqm", "rent_per_sqm_mo", "annual_rent",
                "security_deposit", "rent_free_months"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # LAYER 1 FIX (cloud-safe): Column G annual_rent is an Excel formula cell.
    # When the file is created by openpyxl without LibreOffice recalc — e.g. on
    # cloud deployment or from deal_simulator — data_only=True returns None and
    # fillna(0) masks the problem. Recompute from raw inputs (always present).
    # Logic: only override rows where annual_rent is still 0 after fillna,
    # AND area_sqm > 0, so VACANT rows (area=0 or rent=0) are correctly left at 0.
    if all(c in df.columns for c in ["annual_rent", "area_sqm", "rent_per_sqm_mo"]):
        needs_calc = (df["annual_rent"] == 0) & (df["area_sqm"] > 0) & (df["rent_per_sqm_mo"] > 0)
        df.loc[needs_calc, "annual_rent"] = (
            df.loc[needs_calc, "area_sqm"] * df.loc[needs_calc, "rent_per_sqm_mo"] * 12
        )

    for col in ["cpi_passthrough", "cpi_trigger"]:
        if col in df.columns:
            series = pd.to_numeric(
                df[col].astype(str).str.replace("%", ""), errors="coerce"
            ).fillna(0)
            # If stored as percentage integer (e.g. 75), convert to decimal
            series = series.apply(lambda x: x / 100 if x > 1 else x)
            df[col] = series

    for col in ["lease_start", "lease_expiry", "break_option"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    today = pd.Timestamp(date.today())
    if "lease_expiry" in df.columns:
        df["unexpired_yrs"] = (df["lease_expiry"] - today).dt.days / 365.25
        df["unexpired_yrs"] = df["unexpired_yrs"].clip(lower=0)
    else:
        df["unexpired_yrs"] = 0.0

    df["is_vacant"] = df["tenant"].astype(str).str.upper().str.strip() == "VACANT"
    df.loc[df["is_vacant"], "annual_rent"]   = 0
    df.loc[df["is_vacant"], "unexpired_yrs"] = 0

    if "is_anchor" in df.columns:
        df["is_anchor"] = df["is_anchor"].astype(str).str.upper().str.strip() == "Y"
    else:
        df["is_anchor"] = False

    if "cpi_clause" not in df.columns:
        df["cpi_clause"] = "N"

    return df


def load_cost_sheet(filepath, rr_gross_rent: float = 0.0,
                    rr_gla_total: float = 0.0,
                    rr_vacant_area: float = 0.0) -> dict:
    """
    Read the Cost Sheet tab.

    Parameters
    ----------
    filepath        : path or file-like object
    rr_gross_rent   : gross passing rent computed from rent roll
                      (used as fallback when Cost Sheet B5 formula is None)
    rr_gla_total    : total GLA from rent roll (fallback for B6)
    rr_vacant_area  : total vacant area from rent roll (fallback for B8)
    These three parameters solve the cross-sheet formula link problem:
    B5/B6/B8 in the Cost Sheet link to Rent Roll summary cells but return
    None in cloud environments where LibreOffice recalc has not been run.
    Returns a dict: available, costs, total_costs, noi,
    noi_margin, void_costs, gross_rent_linked.
    """
    result = {
        "available":         False,
        "costs":             {},
        "total_costs":       0.0,
        "noi":               0.0,
        "noi_margin":        0.80,
        "void_costs":        0.0,
        "gross_rent_linked": 0.0,
    }

    try:
        from openpyxl import load_workbook as _lw
        import io as _io

        if hasattr(filepath, "read"):
            filepath.seek(0)   # seek BEFORE read — pointer may be advanced
            raw = filepath.read()
            wb = _lw(_io.BytesIO(raw), data_only=True)
        else:
            wb = _lw(filepath, data_only=True)

        if "Cost Sheet" not in wb.sheetnames:
            return result

        ws = wb["Cost Sheet"]

        def _v(row, col):
            try:
                v = ws.cell(row=row, column=col).value
                if v is None:
                    return None
                return float(v)
            except (TypeError, ValueError):
                return None

        result["gross_rent_linked"] = _v(5, 2) or 0.0

        # Cost Sheet col D (rows 14-23) contains Excel formulas:
        #   D14 = C14 * B5  (rate% * gross_rent)
        #   D16 = C16 * B6  (EUR/m2 * GLA)
        #   D18 = C18       (absolute EUR)
        # In cloud environments without LibreOffice, data_only=True returns None
        # for these formula cells. We recompute from col C (raw rate inputs)
        # which are always plain numeric values — never formulas.
        # This mirrors the exact formula logic hard-coded in the Excel template.
        # B5/B6/B8 are cross-sheet formula links to Rent Roll.
        # They return None in cloud environments (no LibreOffice recalc).
        # Fall back to the values passed in from the already-parsed rent roll.
        gla_total_cs   = _v(6, 2) or rr_gla_total
        vacant_area_cs = _v(8, 2) or rr_vacant_area
        gross_rent_cs  = _v(5, 2) or rr_gross_rent

        cost_keys = [
            "property_mgmt_fee", "asset_mgmt_fee", "building_insurance",
            "capex_reserve", "ground_rent", "grundsteuer",
            "void_service_charge", "void_insurance", "letting_costs", "other_costs",
        ]
        # Basis mirrors Excel template formulas exactly:
        #   rows 14,15,22 = rate% * gross_rent
        #   rows 16,17    = EUR/m2 * GLA
        #   rows 18,19,23 = absolute EUR (rate IS the value, basis=1)
        #   rows 20,21    = EUR/m2 * vacant_area
        cost_basis = [
            gross_rent_cs,   # 14 property_mgmt_fee
            gross_rent_cs,   # 15 asset_mgmt_fee
            gla_total_cs,    # 16 building_insurance
            gla_total_cs,    # 17 capex_reserve
            1.0,             # 18 ground_rent (absolute)
            1.0,             # 19 grundsteuer (absolute)
            vacant_area_cs,  # 20 void_service_charge
            vacant_area_cs,  # 21 void_insurance
            gross_rent_cs,   # 22 letting_costs
            1.0,             # 23 other_costs (absolute)
        ]
        costs = {}
        for i, (k, basis) in enumerate(zip(cost_keys, cost_basis)):
            col_d = _v(14 + i, 4)   # formula result (None if not recalculated)
            col_c = _v(14 + i, 3)   # raw rate input (always numeric)
            if col_d is not None and col_d != 0:
                costs[k] = col_d                  # TRUST Excel calculated value
            elif col_c is not None and col_c != 0:
                costs[k] = col_c * basis           # LAYER 2 FIX: recompute from rate
            else:
                costs[k] = 0.0

        result["costs"]       = costs
        result["total_costs"] = sum(costs.values())

        # LAYER 3 FIX: NOI total in B28 is also a formula cell.
        # Recompute from recalculated cost lines when it returns None.
        noi_from_sheet = _v(28, 2)
        if noi_from_sheet is not None and noi_from_sheet != 0:
            result["noi"] = noi_from_sheet
        else:
            result["noi"] = gross_rent_cs - result["total_costs"]

        gross = result["gross_rent_linked"]
        result["noi_margin"] = result["noi"] / gross if gross > 0 else 0.80
        result["void_costs"] = costs.get("void_service_charge", 0) + costs.get("void_insurance", 0)
        result["available"]  = result["gross_rent_linked"] > 0 or result["total_costs"] > 0

    except Exception as e:
        result["_error"] = str(e)

    return result


# =============================================================================
# 2. METRICS ENGINE
# =============================================================================

def compute_metrics(deal: dict, rr: pd.DataFrame,
                    costs: Optional[dict] = None,
                    noi_margin_override: Optional[float] = None) -> dict:
    """
    Full institutional underwriting metrics.

    NOI source priority:
      1. Itemised Cost Sheet  (if populated)
      2. noi_margin_override slider  (with void cost estimate)
      3. 80% fallback  (with void cost estimate)
    """
    active = rr[~rr["is_vacant"]]
    vacant = rr[rr["is_vacant"]]

    gross_passing_rent = float(active["annual_rent"].sum())
    total_leased_area  = float(active["area_sqm"].sum())
    total_vacant_area  = float(vacant["area_sqm"].sum())
    gla_total          = float(deal.get("gla_total") or (total_leased_area + total_vacant_area))
    vacancy_rate       = total_vacant_area / gla_total if gla_total else 0.0
    wtd_rent_per_sqm   = (gross_passing_rent / (total_leased_area * 12)
                          if total_leased_area else 0.0)

    wault = float(
        (active["annual_rent"] * active["unexpired_yrs"]).sum() / gross_passing_rent
    ) if gross_passing_rent > 0 else 0.0

    anchor_rent = float(active[active["is_anchor"]]["annual_rent"].sum())
    anchor_pct  = anchor_rent / gross_passing_rent if gross_passing_rent else 0.0

    n_active     = len(active)
    n_cpi        = int((active["cpi_clause"].astype(str).str.upper().str.strip() == "Y").sum())
    cpi_coverage = n_cpi / n_active if n_active else 0.0
    cpi_rows     = active[active["cpi_clause"].astype(str).str.upper().str.strip() == "Y"]
    avg_cpi_pt   = float(cpi_rows["cpi_passthrough"].mean()) if len(cpi_rows) > 0 else 0.0

    # NOI determination
    VOID_HOLD_RATE = 20.0   # EUR/m²/yr landlord bears on vacant units
    void_holding   = total_vacant_area * VOID_HOLD_RATE
    noi_source     = "fallback_80pct"
    void_costs     = void_holding
    cost_breakdown = {}

    if costs and costs.get("available") and costs.get("noi", 0) != 0:
        noi           = float(costs["noi"])
        noi_margin    = float(costs["noi_margin"])
        void_costs    = float(costs.get("void_costs", 0))
        cost_breakdown = costs.get("costs", {})
        noi_source    = "cost_sheet"
    elif noi_margin_override is not None and 0 < noi_margin_override <= 1:
        noi        = gross_passing_rent * noi_margin_override - void_holding
        noi_margin = noi / gross_passing_rent if gross_passing_rent else noi_margin_override
        noi_source = "slider_with_void_estimate"
    else:
        noi        = gross_passing_rent * 0.80 - void_holding
        noi_margin = noi / gross_passing_rent if gross_passing_rent else 0.80

    noi = max(noi, 0.0)

    asking_price  = float(deal.get("asking_price")  or 0)
    loan_amount   = float(deal.get("loan_amount")   or 0)
    equity        = float(deal.get("equity")        or 0)
    interest_rate = float(deal.get("interest_rate") or 0)
    amort_rate    = float(deal.get("amort_rate")    or 0)
    acq_cost      = float(deal.get("total_acquisition_cost") or asking_price)

    ltv             = loan_amount / asking_price if asking_price else 0.0
    annual_interest = loan_amount * interest_rate
    annual_amort    = loan_amount * amort_rate
    annual_debt_svc = annual_interest + annual_amort

    dscr         = noi / annual_debt_svc if annual_debt_svc else None
    icr          = noi / annual_interest if annual_interest else None
    giy          = gross_passing_rent / asking_price if asking_price else None
    niy          = noi / asking_price if asking_price else None
    cash_on_cash = (noi - annual_debt_svc) / equity if equity else None

    expiry_in_3y = float(active[active["unexpired_yrs"] <= 3]["annual_rent"].sum())
    expiry_risk  = expiry_in_3y / gross_passing_rent if gross_passing_rent else 0.0

    today = pd.Timestamp(date.today())
    if "break_option" in active.columns:
        imminent = active[
            active["break_option"].notna() &
            ((active["break_option"] - today).dt.days / 365.25 <= 1.5)
        ]
        break_risk_rent = float(imminent["annual_rent"].sum())
    else:
        break_risk_rent = 0.0

    top5     = (active.groupby("tenant")["annual_rent"]
                .sum().sort_values(ascending=False).head(5))
    top5_pct = (top5 / gross_passing_rent * 100).round(1)

    return {
        "gross_passing_rent":   gross_passing_rent,
        "noi_proxy":            noi,
        "noi_margin":           noi_margin,
        "noi_source":           noi_source,
        "void_costs":           void_costs,
        "cost_breakdown":       cost_breakdown,
        "total_leased_area":    total_leased_area,
        "total_vacant_area":    total_vacant_area,
        "gla_total":            gla_total,
        "vacancy_rate":         vacancy_rate,
        "wtd_rent_per_sqm_mo":  wtd_rent_per_sqm,
        "wault":                wault,
        "n_active_leases":      n_active,
        "n_vacant":             len(vacant),
        "expiry_risk_3yr_pct":  expiry_risk,
        "break_risk_rent":      break_risk_rent,
        "anchor_pct":           anchor_pct,
        "top5_tenants":         top5_pct.to_dict(),
        "cpi_coverage":         cpi_coverage,
        "avg_cpi_passthrough":  avg_cpi_pt,
        "asking_price":         asking_price,
        "loan_amount":          loan_amount,
        "equity":               equity,
        "total_acq_cost":       acq_cost,
        "ltv":                  ltv,
        "annual_debt_service":  annual_debt_svc,
        "dscr":                 dscr,
        "icr":                  icr,
        "giy":                  giy,
        "niy":                  niy,
        "cash_on_cash":         cash_on_cash,
        "debt_yield":           (noi / loan_amount) if loan_amount > 0 else None,
        # AfA: 2% linear depreciation on building value (purchase price × 75%)
        # §7 Abs.4 EStG — standard German commercial building depreciation
        # Not a cash cost — reduces taxable income. Land (25%) is excluded.
        "afa_annual":           asking_price * 0.75 * 0.02 if asking_price else None,
        "afa_building_value":   asking_price * 0.75 if asking_price else None,
    }


# =============================================================================
# 3. QA AUDITOR
# =============================================================================

class QAResult:
    def __init__(self, check, status, actual, benchmark, severity):
        self.check, self.status = check, status
        self.actual, self.benchmark, self.severity = actual, benchmark, severity


def run_qa_audit(metrics: dict, deal: dict) -> list:
    results = []

    def add(check, pass_c, warn_c, actual_s, bench, sev_fail):
        if pass_c:
            st, sev = "PASS", "INFO"
        elif warn_c:
            st, sev = "WARN", "YELLOW"
        else:
            st, sev = "FAIL", sev_fail
        results.append(QAResult(check, st, actual_s, bench, sev))

    m = metrics
    add("LTV <= 65%",
        m["ltv"] <= 0.65, m["ltv"] <= 0.70,
        f"{m['ltv']:.1%}", "<= 65%", "RED")

    dscr = m["dscr"]
    if dscr is not None:
        add("DSCR >= 1.30x",
            dscr >= 1.30, dscr >= 1.20,
            f"{dscr:.2f}x", ">= 1.30x", "RED")
    else:
        results.append(QAResult("DSCR >= 1.30x", "N/A", "No debt data", ">= 1.30x", "YELLOW"))

    icr = m["icr"]
    if icr is not None:
        add("ICR >= 1.75x",
            icr >= 1.75, icr >= 1.50,
            f"{icr:.2f}x", ">= 1.75x", "RED")

    add("WAULT >= 4 years",
        m["wault"] >= 4.0, m["wault"] >= 2.5,
        f"{m['wault']:.1f} yrs", ">= 4 yrs", "RED")

    add("Vacancy <= 10%",
        m["vacancy_rate"] <= 0.10, m["vacancy_rate"] <= 0.15,
        f"{m['vacancy_rate']:.1%}", "<= 10%", "YELLOW")

    add("Anchor Concentration <= 55%",
        m["anchor_pct"] <= 0.55, m["anchor_pct"] <= 0.65,
        f"{m['anchor_pct']:.1%}", "<= 55%", "YELLOW")

    add("3-Year Expiry Risk <= 30%",
        m["expiry_risk_3yr_pct"] <= 0.30, m["expiry_risk_3yr_pct"] <= 0.45,
        f"{m['expiry_risk_3yr_pct']:.1%}", "<= 30%", "YELLOW")

    giy = m["giy"]
    if giy is not None:
        add("GIY >= 5.5% (retail DE 2026)",
            giy >= 0.055, giy >= 0.045,
            f"{giy:.2%}", ">= 5.5%", "YELLOW")

    add("NOI > 0",
        m["noi_proxy"] > 0, False,
        f"EUR {m['noi_proxy']:,.0f}", "> 0", "RED")

    add("Equity > 0",
        m["equity"] > 0, False,
        f"EUR {m['equity']:,.0f}", "> 0", "RED")

    dy = m.get("debt_yield")
    if dy is not None:
        add("Debt Yield >= 7.5% (NOI / Loan)",
            dy >= 0.075, dy >= 0.060,
            f"{dy:.2%}", ">= 7.5%", "YELLOW")

    if m.get("noi_source") in ("fallback_80pct", "slider_with_void_estimate"):
        results.append(QAResult(
            "NOI - Cost Sheet populated",
            "WARN", m["noi_source"].replace("_", " "),
            "Itemised Cost Sheet", "YELLOW",
        ))

    return results


# =============================================================================
# 4. VERDICT ENGINE
# =============================================================================

def generate_verdict(metrics: dict, qa: list) -> dict:
    red_fails    = [r for r in qa if r.status == "FAIL" and r.severity == "RED"]
    yellow_warns = [r for r in qa if r.status in ("FAIL", "WARN") and r.severity == "YELLOW"]

    reasons, red_flags = [], []
    for r in red_fails:
        red_flags.append(f"CRITICAL  {r.check}: {r.actual} (required {r.benchmark})")

    m = metrics
    if m["wault"]          >= 5:     reasons.append(f"Strong WAULT of {m['wault']:.1f} years provides income visibility")
    if m.get("giy") and m["giy"] >= 0.065: reasons.append(f"Attractive yield of {m['giy']:.1%} vs. market benchmark")
    if m["vacancy_rate"]   <= 0.05:  reasons.append(f"Low vacancy ({m['vacancy_rate']:.1%}) signals strong market demand")
    if m.get("dscr") and m["dscr"] >= 1.50: reasons.append(f"Robust DSCR of {m['dscr']:.2f}x comfortable debt headroom")
    if m["cpi_coverage"]   >= 0.80:  reasons.append(f"{m['cpi_coverage']:.0%} of leases CPI-indexed -- inflation protection")
    for r in yellow_warns[:3]:
        reasons.append(f"CAUTION  {r.check}: {r.actual} (benchmark {r.benchmark})")

    if len(red_fails) >= 2:
        rec, rat = "DECLINE",     "Multiple critical covenant breaches. Not suitable for IC submission."
    elif len(red_fails) == 1:
        rec, rat = "INVESTIGATE", f"One critical issue: {red_fails[0].check}. Must resolve before IC."
    elif len(yellow_warns) >= 3:
        rec, rat = "INVESTIGATE", "Several caution flags. Further due diligence required."
    else:
        rec, rat = "PASS",        "Deal meets institutional benchmarks. Proceed to full due diligence."

    return {"recommendation": rec, "rationale": rat,
            "top_reasons": reasons[:5], "red_flags": red_flags}


# =============================================================================
# 5. CONSOLE IC MEMO
# =============================================================================

def print_ic_memo(deal, metrics, qa, verdict, costs=None):
    D, S = "=" * 72, "-" * 72
    src_label = {"cost_sheet": "Itemised Cost Sheet",
                 "slider_with_void_estimate": "Sidebar slider + void estimate",
                 "fallback_80pct": "80% proxy fallback"
                 }.get(metrics.get("noi_source", ""), "unknown")

    print(f"\n{D}")
    print("  INSTITUTIONAL RETAIL UNDERWRITER  |  IC MEMO")
    print(D)
    print(f"  Property : {deal.get('property_name','--')}")
    print(f"  Location : {deal.get('city','--')}  |  {deal.get('asset_type','--')}")
    print(f"  Date     : {date.today().strftime('%d %B %Y')}")
    print(S)
    print(f"\n  RENT ROLL")
    print(f"    Active Leases     : {metrics['n_active_leases']}")
    print(f"    Vacant Units      : {metrics['n_vacant']}")
    print(f"    GLA / Vacancy     : {metrics['gla_total']:,.0f} m2  |  {metrics['vacancy_rate']:.1%}")
    print(f"    Gross Rent        : EUR {metrics['gross_passing_rent']:,.0f} / yr")
    print(f"\n  NOI  [{src_label}]")
    if costs and costs.get("costs"):
        for k, v in costs["costs"].items():
            if v: print(f"    {k:<28}: -EUR {v:,.0f}")
    print(f"    Void costs        : -EUR {metrics['void_costs']:,.0f}")
    print(f"    NET OPERATING INC :  EUR {metrics['noi_proxy']:,.0f} / yr  ({metrics['noi_margin']:.1%})")
    print(f"\n  LEASE QUALITY")
    print(f"    WAULT             : {metrics['wault']:.2f} yrs")
    print(f"    Anchor conc.      : {metrics['anchor_pct']:.1%}")
    print(f"    3yr expiry risk   : {metrics['expiry_risk_3yr_pct']:.1%}")
    print(f"\n  FINANCIALS")
    print(f"    LTV               : {metrics['ltv']:.1%}")
    giy_s  = f"{metrics['giy']:.2%}"  if metrics.get("giy")  else "--"
    niy_s  = f"{metrics['niy']:.2%}"  if metrics.get("niy")  else "--"
    dscr_s = f"{metrics['dscr']:.2f}x" if metrics.get("dscr") else "--"
    icr_s  = f"{metrics['icr']:.2f}x"  if metrics.get("icr")  else "--"
    print(f"    GIY / NIY         : {giy_s} / {niy_s}")
    print(f"    DSCR / ICR        : {dscr_s} / {icr_s}")
    print(f"\n  QA AUDIT")
    for r in qa:
        icon = "OK" if r.status == "PASS" else ("!!" if r.status == "WARN" else "XX")
        print(f"    [{icon}]  {r.check:<42} {r.actual:<14} {r.benchmark}")
    print(f"\n{D}")
    print(f"  VERDICT: {verdict['recommendation']}  --  {verdict['rationale']}")
    if verdict["red_flags"]:
        print("\n  RED FLAGS:")
        for f in verdict["red_flags"]: print(f"    {f}")
    if verdict["top_reasons"]:
        print("\n  KEY FACTORS:")
        for r in verdict["top_reasons"]: print(f"    {r}")
    print(f"{D}\n")


# =============================================================================
# 6. MAIN
# =============================================================================

def analyse(filepath):
    """Full pipeline. Returns (deal, rr, costs, metrics, qa, verdict)."""
    print(f"\nLoading: {filepath}")
    deal    = load_deal_overview(filepath)
    rr      = load_rent_roll(filepath)
    # Pass rent roll aggregates so Cost Sheet can resolve cross-sheet
    # formula links without LibreOffice recalc (cloud-safe)
    _rr_active  = rr[~rr["is_vacant"]]
    _rr_vacant  = rr[rr["is_vacant"]]
    costs   = load_cost_sheet(
        filepath,
        rr_gross_rent  = float(_rr_active["annual_rent"].sum()),
        rr_gla_total   = float(rr["area_sqm"].sum()),
        rr_vacant_area = float(_rr_vacant["area_sqm"].sum()),
    )
    metrics = compute_metrics(deal, rr, costs)
    qa      = run_qa_audit(metrics, deal)
    verdict = generate_verdict(metrics, qa)
    print_ic_memo(deal, metrics, qa, verdict, costs)
    return deal, rr, costs, metrics, qa, verdict


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "DE_Retail_Underwriter_Template.xlsx"
    analyse(path)
