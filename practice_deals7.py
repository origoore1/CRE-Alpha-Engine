"""
German Retail Underwriter — Practice Deal Generator
====================================================
Generates realistic simulated deals for learning and tool testing.
Each deal is based on real German retail CRE market patterns.
Includes teaching notes explaining what to look for and why.

Usage:
    python practice_deals.py              → generate all 20 deals
    python practice_deals.py --deal 5     → generate deal #5 only
    python practice_deals.py --difficulty hard → filter by difficulty
"""
import io, shutil, sys, argparse
import datetime as _dtm  # stable alias: immune to test clock-patching of `date`
from datetime import datetime, date, timedelta
from pathlib import Path

# ── design-date anchoring ─────────────────────────────────────────────────────
# Lease dates in the deal configs are literals, written when each deal was
# designed. Without correction, every deal's WAULT decays as real time passes
# and verdicts silently flip (deal 12 drifted INVESTIGATE → DECLINE this way).
# generate_practice_deal() shifts all lease dates by (today − design_date) so
# each deal keeps its designed WAULT and verdict forever. A deal config may
# carry its own "design_date"; otherwise this default applies.
DESIGN_DATE_DEFAULT = date(2026, 5, 28)

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent))
from deal_simulator import MASTER_TEMPLATE, DEAL_CELLS, RR_COLS, COST_ROWS, RR_FIRST_DATA_ROW, _clear_rent_roll
from openpyxl import load_workbook

# ─── PRACTICE DEAL LIBRARY ────────────────────────────────────────────────────

PRACTICE_DEALS = [

# ════════════════════════════════════════════════════════════════════════════════
# BEGINNER LEVEL — Learn the basics
# ════════════════════════════════════════════════════════════════════════════════

{
    "id": 1, "difficulty": "BEGINNER", "verdict_expected": "PASS",
    "title": "The Textbook Deal — Regensburg Fachmarktzentrum",
    "learning_objective": "Understand what a clean, IC-ready retail deal looks like. All metrics green.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ✓ DSCR comfortably above 1.30x — the bank's primary covenant
  ✓ WAULT above 7 years — income visibility well beyond loan term
  ✓ GIY above 6.5% — acceptable yield for secondary German retail
  ✓ LTV below 62% — comfortable equity cushion
  ✓ Edeka anchor with 9yr remaining — strongest food retail covenant in Germany

KEY LEARNING:
  This is your benchmark. Every deal you analyse should be compared to this one.
  Notice how clean the NOI Bridge looks — no void costs because vacancy is zero.
  The DSCR at 1.59x gives you 29bps of headroom before covenant breach.

WHAT WOULD MAKE IT BETTER:
  - Anchor concentration at 62% is above the 55% benchmark — single risk point
  - A second strong tenant (dm, Rossmann) would diversify income
  - IRR at 10yr is only 5.8% — this is a long-hold income asset, not a value-add play
""",
    "deal": {
        "property_name": "Scharnitzpark Regensburg",
        "street_address": "Bajuwarenstrasse 2",
        "city": "Regensburg / Bayern",
        "asset_type": "Fachmarktzentrum (Retail Park)",
        "gla_sqm": 4800, "year_built": 2014, "parking": 210,
        "asking_price": 9_400_000, "equity": 3_700_000,
        "grest_rate": 0.035, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Edeka Minden-Hannover","Food & Grocery","Y",3200,10.50,
         date(2019,4,1),date(2033,3,31),None,"Y",0.75,0.10,25200,"Y",0,"Anchor — 7yr+ remaining"),
        ("U-02","dm-drogerie markt","Health & Beauty","N",620,17.00,
         date(2021,1,1),date(2031,12,31),None,"Y",1.00,0.10,10710,"Y",0,"Strong dm covenant"),
        ("U-03","KiK Textilien","Discount Fashion","N",480,8.50,
         date(2020,7,1),date(2029,6,30),None,"N",0.00,0.00,3672,"Y",0,"Stable discount fashion"),
        ("U-04","Lotto/Tabak/Backshop","Personal Services","N",500,11.00,
         date(2022,1,1),date(2030,12,31),None,"N",0.00,0.00,4950,"Y",0,"Service unit"),
    ],
    "costs": {
        "property_mgmt_pct":0.015,"asset_mgmt_pct":0.005,
        "insurance_sqm":2.20,"capex_sqm":4.50,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.030,"other_abs":2000,
    },
},

{
    "id": 2, "difficulty": "BEGINNER", "verdict_expected": "DECLINE",
    "title": "The Warning Sign — Duisburg Vacant Centre",
    "learning_objective": "Recognise a clearly broken deal. Multiple simultaneous failures.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ✗ 60% vacancy — barely any income to service the debt
  ✗ DSCR 0.13x — the property generates EUR 0.13 for every EUR 1.00 of debt service
  ✗ GIY 2.1% — seller is pricing the asset on hope, not reality
  ✗ WAULT 2.1yr on the few active tenants — even the income you have is short-lived

KEY LEARNING:
  When you see multiple RED flags simultaneously, don't look for fixes — walk away.
  The Deal Repair Engine will show a 52% price reduction needed. That's not a
  negotiation — that's telling you the seller's pricing is fundamentally wrong.
  A EUR 4.2M ask on an asset generating EUR 113k NOI implies a 2.7% yield.
  At 4.1% debt cost, every euro borrowed loses money before amortisation.

WHAT THE SELLER WILL SAY:
  "The asset has huge repositioning potential." This is true but irrelevant —
  repositioning risk isn't priced in. The right question is: what do I pay for
  the cash flow today, and what does the empty space cost me while I re-let it?
""",
    "deal": {
        "property_name": "Stadtgalerie Duisburg-Meiderich",
        "street_address": "Emscherstrasse 71",
        "city": "Duisburg / Nordrhein-Westfalen",
        "asset_type": "Nahversorgungszentrum (distressed)",
        "gla_sqm": 5200, "year_built": 2001, "parking": 95,
        "asking_price": 4_200_000, "equity": 1_200_000,
        "grest_rate": 0.065, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Netto Marken-Discount","Food & Grocery","Y",900,9.50,
         date(2020,1,1),date(2026,12,31),None,"N",0.00,0.00,9405,"Y",0,"Short WAULT — renewal at risk"),
        ("U-02","Steuerberater Kanzlei","Office","N",350,8.00,
         date(2021,4,1),date(2027,3,31),None,"N",0.00,0.00,2520,"N",0,"Office — not retail"),
        ("U-03","VACANT","—","N",800,0,None,None,None,"N",0,0,0,"N",0,"Dark 18 months"),
        ("U-04","VACANT","—","N",900,0,None,None,None,"N",0,0,0,"N",0,"Former fashion anchor"),
        ("U-05","VACANT","—","N",1200,0,None,None,None,"N",0,0,0,"N",0,"Former anchor — structural void"),
        ("U-06","VACANT","—","N",1050,0,None,None,None,"N",0,0,0,"N",0,"Services corridor — unlettable"),
    ],
    "costs": {
        "property_mgmt_pct":0.020,"asset_mgmt_pct":0.008,
        "insurance_sqm":2.80,"capex_sqm":8.00,"ground_rent_abs":0,
        "grundsteuer_abs":6000,"void_svc_sqm":22.0,"void_ins_sqm":2.00,
        "letting_pct":0.100,"other_abs":8000,
    },
},

# ════════════════════════════════════════════════════════════════════════════════
# INTERMEDIATE — Learn to spot hidden risks
# ════════════════════════════════════════════════════════════════════════════════

{
    "id": 3, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "The Erbpacht Trap — Frankfurt Suburb",
    "learning_objective": "Understand how ground rent (Erbpacht) destroys NOI even on a full asset.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ⚠️ Ground rent EUR 85,000/yr in Cost Sheet — this is the trap
  ⚠️ DSCR drops from apparent 1.70x (IO) to 1.14x (amortising)
  ⚠️ IRR negative at 3yr, 5yr, 7yr — ground rent never goes away

KEY LEARNING:
  This is the #1 trap for non-German buyers. An Erbpacht (ground lease) means
  you own the building but NOT the land. You pay ground rent forever.
  At EUR 85,000/yr, the Erbpacht is equivalent to having EUR 1.3M of additional
  debt at 6.5% — debt that never gets repaid and grows with CPI.

  Notice how the GIY looks fine (6.04%) but the Deal Repair Engine shows a
  27% price cut needed for 7% IRR. The ground rent is why.

HOW TO NEGOTIATE:
  Ask the seller: can the Erbpacht be bought out? German law allows this.
  Buyout typically costs 15-20× annual ground rent = EUR 1.3–1.7M.
  This is your negotiating chip: offer EUR 6.1M with EUR 1.5M Erbpacht buyout,
  rather than EUR 7.8M with ongoing ground rent.
""",
    "deal": {
        "property_name": "Rewe Center Dreieich-Sprendlingen",
        "street_address": "Hauptstrasse 180",
        "city": "Dreieich / Hessen",
        "asset_type": "Nahversorgungszentrum",
        "gla_sqm": 3200, "year_built": 2003, "parking": 145,
        "asking_price": 7_800_000, "equity": 3_100_000,
        "grest_rate": 0.060, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Rewe Markt GmbH","Food & Grocery","Y",2000,11.50,
         date(2018,10,1),date(2033,9,30),None,"Y",0.75,0.10,25875,"Y",0,"Anchor 7.5yr remaining"),
        ("U-02","dm-drogerie markt","Health & Beauty","N",560,17.50,
         date(2020,4,1),date(2033,3,31),None,"Y",1.00,0.10,11025,"Y",0,"7yr remaining"),
        ("U-03","Lotto/Tabak","Personal Services","N",180,14.00,
         date(2022,10,1),date(2027,9,30),None,"N",0.00,0.00,1764,"Y",0,"1.5yr remaining"),
        ("U-04","Blumenladen","Personal Services","N",460,8.50,
         date(2021,4,1),date(2029,3,31),None,"N",0.00,0.00,2907,"Y",0,"Below market rent"),
    ],
    "costs": {
        "property_mgmt_pct":0.015,"asset_mgmt_pct":0.005,
        "insurance_sqm":2.50,"capex_sqm":6.00,"ground_rent_abs":85000,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.040,"other_abs":2500,
    },
},

{
    "id": 4, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "The Trophy Trap — Aldi Ingolstadt",
    "learning_objective": "Learn that covenant quality alone doesn't justify price. GIY must work.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ⚠️ GIY 4.32% — Aldi covenant is being priced like a Berlin trophy asset
  ⚠️ DSCR 1.06x — barely above water even with interest-only
  ⚠️ IRR negative at 3yr and 5yr — this is a very long-hold play

KEY LEARNING:
  Aldi is one of the strongest covenants in German retail (parent: Aldi Süd,
  revenues EUR 30B+). But covenant quality doesn't override math.
  At EUR 5.8M asking price and EUR 250,560 gross rent, you're paying a 4.32% yield.
  With 4.1% debt cost, the spread is only 22bps — barely enough to cover costs.

  Brokers market these as "inflation-protected bond-like assets."
  That's true — but bonds don't require EUR 200,000 in acquisition costs
  and EUR 50,000/yr in asset management.

WHAT WOULD MAKE IT WORK:
  The Deal Repair Engine should show: max price EUR 4.56M for DSCR clearance.
  At EUR 4.56M, GIY = 5.50% — still tight but workable.
  The question to ask yourself: is Aldi Ingolstadt worth 21% more than that math?
""",
    "deal": {
        "property_name": "Aldi Süd Fachmarkt Ingolstadt-West",
        "street_address": "Manchinger Strasse 85",
        "city": "Ingolstadt / Bayern",
        "asset_type": "Fachmarkt Single-Tenant Food",
        "gla_sqm": 1800, "year_built": 2011, "parking": 85,
        "asking_price": 5_800_000, "equity": 2_200_000,
        "grest_rate": 0.035, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Aldi Süd Dienstleistungs GmbH","Food & Grocery","Y",1800,11.60,
         date(2018,4,1),date(2033,3,31),None,"Y",0.75,0.10,12528,"Y",0,
         "Single tenant. Aldi Süd (Albrecht-Gruppe) — top-tier German covenant."),
    ],
    "costs": {
        "property_mgmt_pct":0.010,"asset_mgmt_pct":0.003,
        "insurance_sqm":2.00,"capex_sqm":3.50,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.020,"other_abs":1500,
    },
},

{
    "id": 5, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "The LTV Squeeze — Kassel Retail Park",
    "learning_objective": "Understand how a small LTV breach combines with expiry risk and anchor concentration to flag a deal.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ⚠️ LTV 65.3% — just above the 65% bank maximum
  ⚠️ Anchor concentration 55.7% — Rewe just over the 55% single-tenant threshold
  ⚠️ 3-year expiry risk 44% — Rossmann (2.5yr), Takko (1yr) and Lotto (2yr) all expiring within 3yr
  ✓ DSCR ~1.55x — passes comfortably
  ✓ WAULT 4.3yr — just above the 4yr minimum

KEY LEARNING:
  No single issue here is catastrophic, but three yellow warns together trigger INVESTIGATE.
  The LTV breach only requires EUR 22,500 more equity to fix.
  But the more important question is the cluster of short leases expiring within 3yr.
  44% of rent (Rossmann + Takko + Lotto) has fewer than 3 years remaining.
  Takko has been contracting nationally — renewal is uncertain.

  Anchor concentration at 55.7% (Rewe) is also a warn: if Rewe doesn't renew,
  the asset loses more than half its income overnight.

HOW TO APPROACH NEGOTIATION:
  Option A: EUR 23k more equity drops LTV to exactly 65% (easy fix)
  Option B: Negotiate 1% price reduction → EUR 73,500 → LTV drops to 64.3%
  Option C: Require Rossmann and Takko to sign renewal heads-of-terms before closing
""",
    "deal": {
        "property_name": "Nordpark Kassel-Rothenditmold",
        "street_address": "Leipziger Strasse 200",
        "city": "Kassel / Hessen",
        "asset_type": "Fachmarktzentrum (Retail Park)",
        "gla_sqm": 3800, "year_built": 2009, "parking": 165,
        "asking_price": 7_350_000, "equity": 2_550_000,
        "grest_rate": 0.060, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Rewe Markt GmbH","Food & Grocery","Y",2100,11.50,
         date(2019,4,1),date(2032,3,31),None,"Y",0.75,0.10,23100,"Y",0,"Anchor 6yr — 55.7% of rent"),
        ("U-02","Rossmann GmbH","Health & Beauty","N",480,15.50,
         date(2020,10,1),date(2028,9,30),None,"Y",1.00,0.10,8928,"Y",0,"2.5yr remaining — expiry risk"),
        ("U-03","Takko Fashion","Discount Fashion","N",780,8.00,
         date(2020,4,1),date(2027,3,31),None,"N",0.00,0.00,5616,"Y",0,
         "1yr remaining — Takko contracting nationally"),
        ("U-04","Lotto/Post","Personal Services","N",440,12.50,
         date(2021,1,1),date(2028,12,31),None,"N",0.00,0.00,4950,"Y",0,"2yr remaining — expiry risk"),
    ],
    "costs": {
        "property_mgmt_pct":0.015,"asset_mgmt_pct":0.005,
        "insurance_sqm":2.30,"capex_sqm":5.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.050,"other_abs":2000,
    },
},

# ════════════════════════════════════════════════════════════════════════════════
# ADVANCED — Complex real-world scenarios
# ════════════════════════════════════════════════════════════════════════════════

{
    "id": 6, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "The Best Deal in 2026 — Hamburg Convenience",
    "learning_objective": "Understand why Hamburg neighbourhood retail outperforms in 2026.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ✓ DSCR 1.62x — strongest in any PASS scenario
  ✓ GIY 6.60% — excellent for Hamburg prime suburban
  ✓ WAULT 7.8yr — income well secured
  ✓ Zero vacancy — fully let, zero void costs

KEY LEARNING:
  This is the best risk-adjusted play in 2026. Why?
  Three factors converge: (1) Hamburg is Germany's strongest retail catchment,
  (2) convenience/neighbourhood format is structurally resilient to e-commerce,
  (3) Bayern GrESt equivalent (Hamburg 5.5%) is below NRW — lower entry cost.

  Notice the IRR profile: even at 10yr hold it's only ~6.5%. This is a stable
  income asset with low upside. The value is the certainty, not the growth.

ADVANCED QUESTION: What happens if Lidl exercises the 2029 break option?
  They occupy 1,800m² at EUR 9.20/m². Losing them means EUR 198,720/yr gone.
  DSCR would drop from 1.62x to ~0.95x — below water.
  This is why break options must always be stress-tested, even on PASS deals.
""",
    "deal": {
        "property_name": "Hamburger Strasse Nahversorgungszentrum",
        "street_address": "Hamburger Strasse 45",
        "city": "Hamburg / Hamburg",
        "asset_type": "Nahversorgungszentrum",
        "gla_sqm": 3600, "year_built": 2017, "parking": 140,
        "asking_price": 8_200_000, "equity": 3_200_000,
        "grest_rate": 0.055, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Lidl Dienstleistung GmbH","Food & Grocery","Y",1800,9.20,
         date(2020,4,1),date(2035,3,31),date(2029,3,31),"Y",0.75,0.10,16560,"Y",0,
         "Anchor — break option 2029 is key risk"),
        ("U-02","Rossmann GmbH","Health & Beauty","N",560,16.50,
         date(2021,1,1),date(2033,12,31),None,"Y",1.00,0.10,11088,"Y",0,"Strong 7yr"),
        ("U-03","Ernstings Family","Discount Fashion","N",480,10.50,
         date(2019,10,1),date(2029,9,30),None,"N",0.00,0.00,4032,"Y",0,"3yr remaining"),
        ("U-04","Fressnapf GmbH","Pet Supplies","N",420,13.50,
         date(2021,4,1),date(2033,3,31),None,"Y",1.00,0.10,6804,"Y",0,"Long lease"),
        ("U-05","Subway/Services","Personal Services","N",340,12.00,
         date(2022,1,1),date(2030,12,31),None,"N",0.00,0.00,3468,"Y",0,"Food court unit"),
    ],
    "costs": {
        "property_mgmt_pct":0.013,"asset_mgmt_pct":0.004,
        "insurance_sqm":1.90,"capex_sqm":4.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.025,"other_abs":2000,
    },
},

{
    "id": 7, "difficulty": "ADVANCED", "verdict_expected": "INVESTIGATE",
    "title": "The Mixed-Use Trap — Cologne Office-Retail",
    "learning_objective": "Recognise when a 'retail' deal is actually a mixed-use risk.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ⚠️ Two office tenants expiring in 9-12 months — 34% of income at risk
  ⚠️ WAULT 3.5yr — office tenants dragging the weighted average down
  ⚠️ GIY 5.04% — seller pricing as retail despite heavy office exposure
  ⚠️ DSCR 1.21x — only just above the 1.20x WARN floor; very thin headroom
  ⚠️ 29.6% vacancy — two large units already dark

KEY LEARNING:
  The seller is marketing this as 'retail' but 34% of passing rent is office.
  Office market in suburban Cologne (Porz) is structurally weak post-COVID.
  When those leases expire, effective vacancy jumps from 30% to over 60%.

  The DSCR of 1.21x is the tipping point: barely above the WARN threshold.
  Even a 2% rent reduction on renewal pushes DSCR into RED territory.

ADVANCED QUESTION: What does a stress scenario look like?
  If both office tenants leave: gross rent drops from EUR 428k to EUR 279k.
  Void costs increase as 1,360m² more becomes vacant.
  Estimated stress NOI: ~EUR 130k. Stress DSCR: ~0.53x — catastrophic.
  This deal requires office renewal heads-of-terms as a condition precedent to closing.
""",
    "deal": {
        "property_name": "Porz Galerie Köln-Porz Mixed Use",
        "street_address": "Friedrich-Ebert-Ufer 50",
        "city": "Koeln / Nordrhein-Westfalen",
        "asset_type": "Mixed Use Retail + Office",
        "gla_sqm": 4800, "year_built": 1998, "parking": 165,
        "asking_price": 8_500_000, "equity": 4_500_000,
        "grest_rate": 0.065, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Rewe Markt GmbH","Food & Grocery","Y",1600,10.50,
         date(2020,4,1),date(2031,3,31),None,"Y",0.75,0.10,18900,"Y",0,"Ground floor anchor"),
        ("U-02","Rossmann GmbH","Health & Beauty","N",420,15.50,
         date(2021,10,1),date(2030,9,30),None,"Y",1.00,0.10,7245,"Y",0,"4.5yr remaining"),
        ("U-03","Büro Mieter GmbH","Office","N",800,9.50,
         date(2020,1,1),date(2026,12,31),None,"N",0.00,0.00,5700,"N",0,
         "OFFICE — expires Dec 2026. Weak Porz office market."),
        ("U-04","Steuerberater Kanzlei","Office","N",560,8.50,
         date(2019,4,1),date(2027,3,31),None,"N",0.00,0.00,2800,"N",0,
         "OFFICE — 1yr remaining. Likely to downsize."),
        ("U-05","VACANT","—","N",420,0,None,None,None,"N",0,0,0,"N",0,
         "Former office — dark 18 months"),
        ("U-06","VACANT","—","N",1000,0,None,None,None,"N",0,0,0,"N",0,
         "Large retail void — 24 months dark"),
    ],
    "costs": {
        "property_mgmt_pct":0.018,"asset_mgmt_pct":0.006,
        "insurance_sqm":3.00,"capex_sqm":7.00,"ground_rent_abs":0,
        "grundsteuer_abs":4500,"void_svc_sqm":20.0,"void_ins_sqm":2.00,
        "letting_pct":0.080,"other_abs":5000,
    },
},

{
    "id": 8, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "The Pricing Question — Munich Premium Core",
    "learning_objective": "Understand when a perfect PASS still raises an investment question.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ✓ Zero fails, zero warnings — technically perfect
  ✓ WAULT 8.1yr — exceptional income visibility
  ✓ Bayern GrESt 3.5% — lowest in Germany
  ✓ Edeka + dm + Rossmann + Fressnapf — institutional-grade tenant mix

  BUT:
  ⚠️ 10yr IRR only ~5.8% — below most fund hurdle rates (7-8%)
  ⚠️ GIY 5.77% — priced for a 2021 world at 1.5% interest rates
  ⚠️ True day-1 equity EUR 6.9M on a EUR 14.5M asset

KEY LEARNING:
  This is the 2021 pricing problem. A Bayern pension fund is selling at a price
  that made sense when the 10yr Bund was at -0.5%. With 10yr Bund at 2.8% in 2026,
  the risk premium for owning retail real estate has compressed.

  A 5.77% GIY vs 4.1% debt cost = 167bps spread. Institutional minimum is usually
  200-250bps. At EUR 14.5M asking, you're accepting spread compression.

ADVANCED QUESTION: At what price does this become a buy?
  For 7% levered IRR: ~EUR 12.2M (Deal Repair will calculate this)
  That's a 16% discount to asking — the real negotiation question.
""",
    "deal": {
        "property_name": "Germering Fachmarktzentrum München-West",
        "street_address": "Untere Bahnhofstrasse 38",
        "city": "Germering / Bayern",
        "asset_type": "Fachmarktzentrum Core",
        "gla_sqm": 5200, "year_built": 2016, "parking": 230,
        "asking_price": 14_500_000, "equity": 5_600_000,
        "grest_rate": 0.035, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.041, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.020, "io_period": 2,
    },
    "rent_roll": [
        ("U-01","Edeka München","Food & Grocery","Y",2600,12.50,
         date(2019,4,1),date(2035,3,31),None,"Y",0.75,0.10,36375,"Y",0,"9yr remaining"),
        ("U-02","dm-drogerie markt","Health & Beauty","N",620,18.50,
         date(2020,10,1),date(2035,9,30),None,"Y",1.00,0.10,12838,"Y",0,"9.5yr remaining"),
        ("U-03","Rossmann GmbH","Health & Beauty","N",420,17.00,
         date(2021,4,1),date(2033,3,31),None,"Y",1.00,0.10,7956,"Y",0,"7yr remaining"),
        ("U-04","Fressnapf GmbH","Pet Supplies","N",480,13.00,
         date(2019,10,1),date(2034,9,30),None,"Y",1.00,0.10,7020,"Y",0,"8.5yr remaining"),
        ("U-05","Ernstings Family","Discount Fashion","N",380,10.50,
         date(2022,4,1),date(2031,3,31),None,"N",0.00,0.00,2999,"Y",0,"5yr remaining"),
        ("U-06","Subway/Services","Personal Services","N",700,12.00,
         date(2021,10,1),date(2030,9,30),None,"N",0.00,0.00,4200,"Y",0,"4.5yr remaining"),
    ],
    "costs": {
        "property_mgmt_pct":0.012,"asset_mgmt_pct":0.004,
        "insurance_sqm":1.80,"capex_sqm":4.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.025,"other_abs":2000,
    },
},


# ════════════════════════════════════════════════════════════════════════════════
# ADVANCED — Real Lahav L.R. Germany Acquisitions
# ════════════════════════════════════════════════════════════════════════════════

{
    "id": 9, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "The Single-Tenant Bet — JAWOLL Neighborhood Center",
    "learning_objective": "Understand why a mechanical PASS is NOT the end of the analysis — 100% single-tenant concentration demands professional judgment beyond the score.",
    "teaching_notes": """
ENGINE VERDICT: PASS. The advanced lesson: PASS does not mean buy.

WHAT THE ENGINE SEES:
  PASS: DSCR ~2.4x, GIY ~7.8%, WAULT ~11yr, LTV 64%, vacancy 0%
  YELLOW: Anchor Concentration 100% (single flag, not enough for INVESTIGATE)

WHY THE ENGINE SAYS PASS:
  The QA engine needs 3 YELLOW fails or 1 RED fail for INVESTIGATE. Anchor
  concentration is a YELLOW fail. With only 1 YELLOW, the engine returns PASS.
  All other metrics are genuinely excellent. Mechanically, this is a clean deal.

WHY A PROFESSIONAL SAYS INVESTIGATE:
  When 100% of income is from one tenant, every metric becomes binary:
    Either JAWOLL pays → DSCR 2.4x, all covenants met
    Or JAWOLL exits → NOI collapses → instant covenant breach
  The 2037 expiry protects you — but JAWOLL has ~90 stores, not 500.
  A covenant failure or operational restructure cannot be ruled out over 11 years.

  Run the Scenario Engine NOW (Section 10):
    Severity 2 (anchor exits): NOI falls sharply → DECLINE verdict
    Severity 1 (renews at 85%): viable — this is the base case you are betting on
  The SWOT analysis will show 100% as the dominant threat.

LAHAV'S EDGE:
  Lahav holds 3 JAWOLL assets and knows renewal behaviour from direct experience.
  A private investor without this network faces pure binary risk.
  The 7,000 m² expansion plot is the real upside: JAWOLL renewal + expansion
  materially increases site value beyond the income-based valuation.

WHAT TO BRING TO IC:
  - Renewal Intention Letter from JAWOLL
  - Covenant analysis: is JAWOLL growing, stable, or consolidating?
  - Re-letting plan for 7,290 m² if JAWOLL does not renew in 2037
""",
    "deal": {
        "property_name": "JAWOLL Neighborhood Center Germany",
        "street_address": "Regensburg Retail Park",
        "city": "Regensburg / Bayern",
        "asset_type": "Nahversorgungszentrum Single-Tenant",
        "gla_sqm": 7290, "year_built": 2008, "parking": 180,
        "asking_price": 5_000_000, "equity": 1_800_000,
        "grest_rate": 0.035, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.0385, "loan_term_yrs": 20,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.005, "io_period": 0,
    },
    "rent_roll": [
        ("U-01","JAWOLL GmbH","Food & Non-Food Discount","Y",7290,4.46,
         date(2020,6,1),date(2037,5,31),None,"Y",0.75,0.10,39000,"Y",0,
         "Single tenant. JAWOLL ~90 stores Germany. Discount food+non-food. No break option to 2037."),
    ],
    "costs": {
        "property_mgmt_pct":0.010,"asset_mgmt_pct":0.003,
        "insurance_sqm":2.00,"capex_sqm":4.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.020,"other_abs":1500,
    },
},

{
    "id": 10, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "The High-Yield Question — Dormagen Trinkgut",
    "learning_objective": "Understand why a 10%+ yield produces a PASS score — and what risks the score cannot see.",
    "teaching_notes": """
ENGINE VERDICT: PASS. The advanced lesson: high yield is a signal, not an answer.

WHAT THE ENGINE SEES:
  PASS: GIY ~10.6%, DSCR ~3.6x, WAULT ~12.7yr, LTV 60.2%, vacancy 0%
  YELLOW: Anchor Concentration 100% (single flag, not enough for INVESTIGATE)

WHY THE ENGINE SAYS PASS:
  With one YELLOW fail (single-tenant concentration), the engine returns PASS.
  By conventional metrics, this is correct: the income is very strong, the lease
  is long, and the DSCR is exceptional. Mechanically, a clean deal.

WHY A PROFESSIONAL PAUSES:
  The German market is efficient. A ~10.5% GIY on a fully-let 12yr lease is
  the market pricing in risks that the income metrics cannot see:

  1. SPECIALIST FORMAT RISK: Trinkgut is a beverage specialist, not a grocery anchor.
     The format is concentrated (beverages/drinks) — re-letting 2,800 m² to another
     beverage retailer at lease expiry (2038) will be very difficult.
     A REWE or Edeka would need 400 m² of refurbishment to take this space.

  2. EXIT RISK: At 2038, your buyer is another private investor — not an institution.
     Institutional funds exclude single-tenant specialist retail. The exit yield
     at sale will likely be higher (10%+) unless the format has changed.

  3. ACQUISITION COST: NRW GrESt 6.5% = EUR 245,050. This is 16.3% of equity.
     Explore share deal (Anteilskauf) — potential saving ~EUR 180k net.

  4. COVENANT SCORE: Trinkgut is not in ANCHOR_NAMES — it will not receive the
     strong covenant scores that REWE/Edeka/Aldi get. The SWOT analysis will show
     this clearly in the Weaknesses quadrant.

WHAT WOULD MAKE IT A BUY:
  - Trinkgut signs a 2042+ lease extension before closing
  - Structural survey confirms conversion viability to grocery format
  - Price at EUR 3.2M (GIY ~12.4%) pricing in re-letting risk at 2038 expiry
""",
    "deal": {
        "property_name": "Dormagen Retail Center",
        "street_address": "Dormagen Stadtmitte",
        "city": "Dormagen / Nordrhein-Westfalen",
        "asset_type": "Nahversorgungszentrum",
        "gla_sqm": 2800, "year_built": 1999, "parking": 95,
        "asking_price": 3_770_000, "equity": 1_500_000,
        "grest_rate": 0.065, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.0385, "loan_term_yrs": 20,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.005, "io_period": 0,
    },
    "rent_roll": [
        ("U-01","Trinkgut GmbH","Food & Grocery","Y",2800,11.90,
         date(2016,1,1),date(2038,12,31),None,"Y",0.75,0.10,35000,"Y",0,
         "Trinkgut — beverage specialist discount retail. Lease to 2038. High yield reflects secondary location and specialist format."),
    ],
    "costs": {
        "property_mgmt_pct":0.012,"asset_mgmt_pct":0.004,
        "insurance_sqm":2.50,"capex_sqm":6.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":0,"void_ins_sqm":0,
        "letting_pct":0.030,"other_abs":2000,
    },
},

{
    "id": 11, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "The Discount Portfolio — Borken & Altenkunstadt",
    "learning_objective": "Learn how a diversified discount tenant mix changes the risk profile vs single-tenant deals.",
    "teaching_notes": """
WHAT TO LOOK FOR:
  ✓ No single tenant above 35% of income — textbook diversification
  ✓ Aldi Süd + JAWOLL = two strong anchors with complementary formats
  ✓ WAULT ~11-12yr — well above the 4yr bank minimum
  ✓ Fressnapf, JYSK, Tedox — European chains with stable expansion models
  ⚠️ 500 m² vacant unit (U-06) — True Yield Gap shows re-letting upside
  ⚠️ NRW GrESt 6.5% — share deal saves ~EUR 552,500 net

KEY LEARNING:
  Compare this to Deals 9 and 10. Same discount retail universe, very different
  risk profile. JAWOLL alone (U-01) would be an INVESTIGATE. Add Aldi Süd (U-02)
  and the portfolio becomes a PASS — because no single exit breaks the DSCR.

  True Yield Gap Module: the 500 m² vacant unit (U-06) represents ~EUR 55,000/yr
  of ERV upside. At a 9.5% cap rate, that's ~EUR 580,000 of additional value
  if let — equivalent to a 5.9% discount to asking price in value terms.

NRW GREST CALCULATION:
  GrESt = EUR 9,850,000 × 6.5% = EUR 640,250
  Share deal (Anteilskauf, 94.9% acquisition): saves 94.9% × EUR 640,250 ≈ EUR 607,597 gross
  Less legal/notarial restructuring costs (~EUR 55,000) → net saving ~EUR 552,500
  This is 15.8% of equity — material. Worth structuring if hold > 10 years.

SCENARIO ENGINE:
  Stress (sev 2): JAWOLL exits (anchor_rent = EUR 352,800). Remaining NOI still
  covers debt service from Aldi + three minor tenants → INVESTIGATE, not DECLINE.
  This is the diversification premium in action.
""",
    "deal": {
        "property_name": "Borken Altenkunstadt Retail Portfolio",
        "street_address": "Borken Retail Park",
        "city": "Borken / Nordrhein-Westfalen",
        "asset_type": "Fachmarktzentrum (Retail Park)",
        "gla_sqm": 8200, "year_built": 2005, "parking": 280,
        "asking_price": 9_850_000, "equity": 3_500_000,
        "grest_rate": 0.065, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.0385, "loan_term_yrs": 20,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.005, "io_period": 0,
    },
    "rent_roll": [
        ("U-01","Jawoll GmbH","Food & Non-Food Discount","Y",2800,10.50,
         date(2019,1,1),date(2037,12,31),None,"Y",0.75,0.10,35280,"Y",0,
         "Anchor — Jawoll to 2037"),
        ("U-02","Aldi Süd","Food & Grocery","Y",1200,12.00,
         date(2020,4,1),date(2035,3,31),None,"Y",0.75,0.10,17280,"Y",0,
         "Strong Aldi Süd covenant to 2035"),
        ("U-03","Fressnapf GmbH","Pet Supplies","N",900,11.50,
         date(2021,1,1),date(2033,12,31),None,"Y",1.00,0.10,12420,"Y",0,
         "Long lease 7yr remaining"),
        ("U-04","JYSK GmbH","Home & Living","N",1800,9.00,
         date(2019,7,1),date(2033,6,30),None,"N",0.00,0.00,9720,"Y",0,
         "Scandinavian home retail"),
        ("U-05","Tedox KG","Discount Home","N",1000,8.50,
         date(2020,1,1),date(2033,12,31),None,"N",0.00,0.00,5100,"Y",0,
         "Discount home/textile"),
        ("U-06","VACANT","—","N",500,0,None,None,None,"N",0,0,0,"N",0,
         "Vacant unit — former services"),
    ],
    "costs": {
        "property_mgmt_pct":0.014,"asset_mgmt_pct":0.004,
        "insurance_sqm":2.20,"capex_sqm":5.00,"ground_rent_abs":0,
        "grundsteuer_abs":0,"void_svc_sqm":18.0,"void_ins_sqm":1.80,
        "letting_pct":0.040,"other_abs":3000,
    },
},

{
    "id": 12, "difficulty": "ADVANCED", "verdict_expected": "INVESTIGATE",
    "design_date": date(2024, 12, 1),  # anchors WAULT ≈ 2.2yr → mechanical INVESTIGATE as designed
    "title": "The Value-Add Trap — Kerken Woolworth Distressed",
    "learning_objective": "Understand why INVESTIGATE is the right mechanical score here — and why professional judgment upgrades it to DECLINE.",
    "teaching_notes": """
ENGINE VERDICT: INVESTIGATE. The advanced lesson: INVESTIGATE here means DECLINE.

WHAT THE ENGINE SEES:
  RED:    WAULT ~2.5yr (borderline below 2.5yr threshold — 1 RED fail)
  YELLOW: Vacancy 17.8% (above 15%)
  YELLOW: 3-Year Expiry Risk 100% (all rent expires within 3yr)
  1 RED fail → INVESTIGATE (engine needs 2 RED fails for DECLINE)

WHY THE ENGINE STOPS AT INVESTIGATE:
  The engine evaluates CURRENT income. Woolworth is currently an active tenant
  paying EUR 250,800/yr. Action is paying EUR 144,000/yr. Total NOI is positive
  (~EUR 276k). With only 1 RED fail, the engine correctly says: investigate further.

WHY YOU SHOULD MENTALLY UPGRADE THIS TO DECLINE:
  The engine cannot see what happens in 8 weeks. Woolworth's break option fires on
  June 30, 2026 — that is NOW. If Woolworth exercises the break:
    - Lose EUR 250,800/yr immediately (63.6% of gross rent)
    - Add void holding costs: 2,200 m² x EUR 22/m²/yr = EUR 48,400/yr additional costs
    - Remaining NOI from Action alone: EUR 144,000 - EUR 48,400 void - overheads
    - NOI goes NEGATIVE → RED fail on NOI > 0 → DECLINE verdict

  Run the Scenario Engine at Severity 2 to see this play out numerically.
  The Woolworth break option will also show in the Break Option Stress table.

  In addition: both active leases expire in 2028 — only 2 years from now.
  Even if the break is not exercised, you face 100% rollover risk in 2 years.

VALUE-ADD MYTH:
  EUR 556/m² for NRW retail sounds cheap. But you are not buying a building —
  you are buying a 2-year income stream with no guarantee of renewal.
  Re-letting 4,500 m² in a secondary NRW location typically takes 18-36 months.
  Void costs during that period: 4,500 m² x EUR 24.20/m²/yr = EUR 108,900/yr.

LAHAV'S EDGE (that you don't have):
  Lahav has 10-15yr lease extension options negotiated with both tenants and
  operator relationships to introduce replacement tenants before break fires.
  Without those relationships: maximum bid EUR 1.4-1.6M (stress-stabilised NOI).
""",
    "deal": {
        "property_name": "Kerken Retail Center Woolworth",
        "street_address": "Kerken Fachmarktzentrum",
        "city": "Kerken / Nordrhein-Westfalen",
        "asset_type": "Fachmarktzentrum (distressed)",
        "gla_sqm": 4500, "year_built": 1995, "parking": 120,
        "asking_price": 2_500_000, "equity": 1_200_000,
        "grest_rate": 0.065, "notar_rate": 0.015, "broker_rate": 0.010,
        "analysis_date": date.today(), "hold_period": 10,
        "interest_rate": 0.0385, "loan_term_yrs": 10,
        "amort_type": "Annuitaetendarlehen", "amort_rate": 0.005, "io_period": 0,
    },
    "rent_roll": [
        ("U-01","Woolworth GmbH","Discount Fashion","N",2200,9.50,
         date(2018,1,1),date(2028,12,31),date(2026,6,30),"N",0.00,0.00,6270,"Y",0,
         "Discount fashion. Break option June 2026 — imminent."),
        ("U-02","Action Germany","Discount Non-Food","N",1500,8.00,
         date(2021,4,1),date(2028,3,31),None,"N",0.00,0.00,3600,"Y",0,
         "Non-food discount 2yr remaining"),
        ("U-03","VACANT","—","N",800,0,None,None,None,"N",0,0,0,"N",0,
         "Former anchor — dark 12 months"),
    ],
    "costs": {
        "property_mgmt_pct":0.018,"asset_mgmt_pct":0.006,
        "insurance_sqm":3.00,"capex_sqm":8.00,"ground_rent_abs":0,
        "grundsteuer_abs":3500,"void_svc_sqm":22.0,"void_ins_sqm":2.20,
        "letting_pct":0.080,"other_abs":5000,
    },
},

]


# ─── EXCEL GENERATOR ──────────────────────────────────────────────────────────

def generate_practice_deal(deal_config: dict, output_dir: str = '.') -> str:
    """Generate Excel file for one practice deal."""
    d = deal_config['deal']
    rr = deal_config['rent_roll']
    costs = deal_config['costs']

    filename = f"Practice_{deal_config['id']:02d}_{deal_config['difficulty']}_{deal_config['verdict_expected']}_{deal_config['title'][:30].replace(' ','_').replace('—','').replace('/','')}.xlsx"
    filename = ''.join(c for c in filename if c.isalnum() or c in '._-')
    out_path = Path(output_dir) / filename

    shutil.copy2(MASTER_TEMPLATE, out_path)
    wb = load_workbook(out_path)

    ws_deal = wb['Deal Overview']
    for field, (row, col) in DEAL_CELLS.items():
        val = d.get(field)
        if val is not None:
            ws_deal.cell(row=row, column=col).value = val

    ws_rr = wb['Rent Roll']
    _clear_rent_roll(ws_rr)

    # Shift lease dates so the deal behaves as it did on its design date
    _design = deal_config.get("design_date", DESIGN_DATE_DEFAULT)
    _shift = timedelta(days=(date.today() - _design).days)

    def _sh(v):
        # _dtm.date (not the module-global `date`) so the check still matches
        # real date objects when tests replace `date` with a frozen subclass
        return v + _shift if isinstance(v, _dtm.date) else v

    for i, tenant in enumerate(rr):
        row = RR_FIRST_DATA_ROW + i
        (unit_ref, tenant_name, sector, anchor, area_sqm, rent_sqm_mo,
         lease_start, lease_expiry, break_option, cpi_clause, cpi_pt, cpi_trig,
         sec_dep, svc_rec, rf_mo, notes) = tenant
        lease_start, lease_expiry, break_option = _sh(lease_start), _sh(lease_expiry), _sh(break_option)
        ws_rr.cell(row=row, column=RR_COLS['unit_ref']).value          = unit_ref
        ws_rr.cell(row=row, column=RR_COLS['tenant']).value            = tenant_name
        ws_rr.cell(row=row, column=RR_COLS['sector']).value            = sector
        ws_rr.cell(row=row, column=RR_COLS['anchor']).value            = anchor
        ws_rr.cell(row=row, column=RR_COLS['area_sqm']).value          = area_sqm
        ws_rr.cell(row=row, column=RR_COLS['rent_sqm_mo']).value       = rent_sqm_mo
        ws_rr.cell(row=row, column=RR_COLS['lease_start']).value       = lease_start
        ws_rr.cell(row=row, column=RR_COLS['lease_expiry']).value      = lease_expiry
        ws_rr.cell(row=row, column=RR_COLS['break_option']).value      = break_option
        ws_rr.cell(row=row, column=RR_COLS['cpi_clause']).value        = cpi_clause
        ws_rr.cell(row=row, column=RR_COLS['cpi_passthrough']).value   = cpi_pt
        ws_rr.cell(row=row, column=RR_COLS['cpi_trigger']).value       = cpi_trig
        ws_rr.cell(row=row, column=RR_COLS['security_deposit']).value  = sec_dep
        ws_rr.cell(row=row, column=RR_COLS['svc_chg_recov']).value     = svc_rec
        ws_rr.cell(row=row, column=RR_COLS['rent_free_months']).value  = rf_mo
        ws_rr.cell(row=row, column=RR_COLS['notes']).value             = notes

    ws_cost = wb['Cost Sheet']
    from deal_simulator import COST_ROWS
    for cost_key, row_num in COST_ROWS.items():
        ws_cost.cell(row=row_num, column=3).value = costs.get(cost_key, 0)

    wb.save(out_path)
    return str(out_path)


def generate_teaching_notes(deal_config: dict) -> str:
    """Generate formatted teaching notes for a deal."""
    d = deal_config
    lines = [
        f"{'='*70}",
        f"PRACTICE DEAL #{d['id']} — {d['difficulty']}",
        f"{d['title']}",
        f"{'='*70}",
        f"Expected Verdict: {d['verdict_expected']}",
        f"Learning Objective: {d['learning_objective']}",
        f"",
        d['teaching_notes'].strip(),
        f"",
        f"Deal Parameters:",
        f"  Asking price: EUR {d['deal']['asking_price']:,}",
        f"  City: {d['deal']['city']}",
        f"  GLA: {d['deal']['gla_sqm']:,} m²",
        f"  Tenants: {len([t for t in d['rent_roll'] if t[1] != 'VACANT'])} active, "
        f"{len([t for t in d['rent_roll'] if t[1] == 'VACANT'])} vacant",
        f"{'='*70}",
    ]
    return '\n'.join(lines)


def run_all(output_dir='.', difficulty=None, verbose=True):
    """Generate all practice deals and verify them through the engine."""
    import io as _io
    import rent_roll_parser as rrp

    deals_to_run = PRACTICE_DEALS
    if difficulty:
        deals_to_run = [d for d in PRACTICE_DEALS if d['difficulty'] == difficulty.upper()]

    results = []
    for deal_config in deals_to_run:
        path = generate_practice_deal(deal_config, output_dir)

        # Verify through engine
        with open(path, 'rb') as f:
            data = _io.BytesIO(f.read())
        deal   = rrp.load_deal_overview(data)
        rr     = rrp.load_rent_roll(data)
        act    = rr[~rr['is_vacant']]; vac = rr[rr['is_vacant']]
        costs  = rrp.load_cost_sheet(data,
            rr_gross_rent  = float(act['annual_rent'].sum()),
            rr_gla_total   = float(rr['area_sqm'].sum()),
            rr_vacant_area = float(vac['area_sqm'].sum()))
        m  = rrp.compute_metrics(deal, rr, costs)
        qa = rrp.run_qa_audit(m, deal)
        v  = rrp.generate_verdict(m, qa)

        expected = deal_config['verdict_expected']
        actual   = v['recommendation']
        match    = actual == expected

        results.append({
            'id':       deal_config['id'],
            'title':    deal_config['title'][:40],
            'expected': expected,
            'actual':   actual,
            'match':    match,
            'path':     path,
            'giy':      m.get('giy',0),
            'dscr':     m.get('dscr',0),
            'wault':    m.get('wault',0),
        })

        if verbose:
            status = '✓' if match else '✗'
            print(f"  [{status}] Deal {deal_config['id']:02d} {actual:<12} "
                  f"GIY:{m.get('giy',0):.2%} DSCR:{m.get('dscr',0):.2f}x "
                  f"WAULT:{m.get('wault',0):.1f}yr — {deal_config['title'][:35]}")

    all_match = all(r['match'] for r in results)
    if verbose:
        print(f"\n{'All verdicts match expected' if all_match else 'SOME VERDICTS MISMATCH'}")

    return results


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate practice deals')
    parser.add_argument('--deal', type=int, help='Generate specific deal by ID')
    parser.add_argument('--difficulty', choices=['BEGINNER','INTERMEDIATE','ADVANCED'],
                        help='Filter by difficulty')
    parser.add_argument('--notes', action='store_true', help='Print teaching notes')
    args = parser.parse_args()

    print("German Retail Underwriter — Practice Deal Generator")
    print(f"{len(PRACTICE_DEALS)} deals available")
    print()

    if args.notes and args.deal:
        d = next((x for x in PRACTICE_DEALS if x['id'] == args.deal), None)
        if d: print(generate_teaching_notes(d))
    elif args.deal:
        d = next((x for x in PRACTICE_DEALS if x['id'] == args.deal), None)
        if d:
            path = generate_practice_deal(d)
            print(f"Generated: {Path(path).name}")
    else:
        print("Generating all practice deals...")
        run_all(difficulty=args.difficulty)
