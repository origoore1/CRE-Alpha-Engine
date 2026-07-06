"""
German Retail Underwriter — Practice Deal Library (v1, 20 deals)
================================================================
20 realistic German retail CRE deals. Each deal is calibrated so the tool
produces a deterministic verdict based on the key metrics:
  DSCR · LTV · WAULT · GIY · NIY · Anchor % · Vacancy

Verdict distribution : 7 PASS / 7 INVESTIGATE / 6 DECLINE
Difficulty           : 6 BEGINNER / 8 INTERMEDIATE / 6 ADVANCED

Analysis date is fixed at 2025-06-01 so WAULT never drifts with calendar time.

Usage:
    python practice_deals_100.py              → generate all 20 deals
    python practice_deals_100.py --deal 5     → single deal
    python practice_deals_100.py --verdict PASS
"""
import shutil, sys, argparse
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from deal_simulator import (
    MASTER_TEMPLATE, DEAL_CELLS, RR_COLS, COST_ROWS,
    RR_FIRST_DATA_ROW, _clear_rent_roll,
)
from openpyxl import load_workbook

# Fixed reference date — keeps WAULT stable across calendar time.
REF = date(2025, 6, 1)

# ─── Standard cost bundles ────────────────────────────────────────────────────
_COSTS_CLEAN = dict(
    property_mgmt_pct=0.015, asset_mgmt_pct=0.005,
    insurance_sqm=2.20, capex_sqm=4.50,
    ground_rent_abs=0, grundsteuer_abs=0,
    void_svc_sqm=0, void_ins_sqm=0,
    letting_pct=0.030, other_abs=1500,
)
_COSTS_ELEVATED = dict(
    property_mgmt_pct=0.018, asset_mgmt_pct=0.006,
    insurance_sqm=2.80, capex_sqm=6.50,
    ground_rent_abs=0, grundsteuer_abs=3000,
    void_svc_sqm=18.0, void_ins_sqm=2.00,
    letting_pct=0.060, other_abs=3500,
)
_COSTS_DISTRESSED = dict(
    property_mgmt_pct=0.025, asset_mgmt_pct=0.008,
    insurance_sqm=3.50, capex_sqm=10.00,
    ground_rent_abs=0, grundsteuer_abs=5000,
    void_svc_sqm=25.0, void_ins_sqm=3.00,
    letting_pct=0.100, other_abs=6000,
)

# ─── Helper: vacant unit tuple ────────────────────────────────────────────────
def _vac(ref, area, note="Vacant unit"):
    return (ref, "VACANT", "—", "N", area, 0, None, None, None, "N", 0, 0, 0, "N", 0, note)


# ═════════════════════════════════════════════════════════════════════════════
# PRACTICE DEALS
# ═════════════════════════════════════════════════════════════════════════════
PRACTICE_DEALS_100 = [

# ─────────────────────────────────────────────────────────────────────────────
# PASS DEALS  (1 – 7)
# Target: DSCR ≥ 1.25 · LTV ≤ 68% · WAULT ≥ 5 yr · no red flags
# ─────────────────────────────────────────────────────────────────────────────

{
    "id": 1, "difficulty": "BEGINNER", "verdict_expected": "PASS",
    "title": "Textbook NVZ — Edeka + dm Augsburg",
    "learning_objective": "What a clean, IC-ready deal looks like.",
    "teaching_notes": (
        "DSCR ~1.41x, LTV 61%, WAULT ~6.7yr. "
        "Edeka Minden covenant (investment grade) anchors the income stream. "
        "Zero vacancy, long leases, Bayern GrESt 3.5% — the benchmark deal."
    ),
    "deal": dict(
        property_name="Edeka Nahversorgungszentrum Augsburg-Lechhausen",
        street_address="Blücherstrasse 18",
        city="Augsburg / Bayern",
        asset_type="Nahversorgungszentrum",
        gla_sqm=2800, year_built=2015, parking=120,
        asking_price=6_500_000, equity=2_535_000,
        grest_rate=0.035, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # area 2200 · rent 10.50 → ann. rent 277,200 · expiry 2032-03 → 6.75yr
        ("U-01","Edeka Minden-Hannover","Food & Grocery","Y",2200,10.50,
         date(2019,4,1),date(2032,3,31),None,"Y",0.75,0.10,20790,"Y",0,
         "Anchor — 6.75yr remaining"),
        # area 600 · rent 16.00 → ann. rent 115,200 · expiry 2031-12 → 6.5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",600,16.00,
         date(2021,1,1),date(2031,12,31),None,"Y",1.00,0.10,9600,"Y",0,
         "Strong dm covenant — 6.5yr remaining"),
    ],
    # Gross rent 392,400 | Loan 3,965,000 | DS 241,865 | NOI ~341k | DSCR ~1.41
    "costs": _COSTS_CLEAN,
},

{
    "id": 2, "difficulty": "BEGINNER", "verdict_expected": "PASS",
    "title": "Lidl Standalone — Dortmund Mengede",
    "learning_objective": "Single-tenant discounter: maximum income certainty, one risk point.",
    "teaching_notes": (
        "DSCR ~1.33x, LTV 60%, WAULT 7yr. Lidl KG is one of the strongest covenants "
        "in German retail. Single-tenant concentration is the only flag — offset by the "
        "absolute NNN structure and long lease."
    ),
    "deal": dict(
        property_name="Lidl Lebensmittelmarkt Dortmund-Mengede",
        street_address="Mengeder Strasse 112",
        city="Dortmund / Nordrhein-Westfalen",
        asset_type="Lebensmittelmarkt (Discounter)",
        gla_sqm=1250, year_built=2017, parking=80,
        asking_price=2_500_000, equity=1_000_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # area 1250 · rent 9.00 → ann. rent 135,000 · expiry 2032-05 → 7yr
        ("U-01","Lidl Vertriebs-GmbH & Co.","Food & Grocery","Y",1250,9.00,
         date(2017,6,1),date(2032,5,31),None,"Y",0.75,0.10,10125,"Y",0,
         "NNN lease — Lidl KG guarantor — 7yr remaining"),
    ],
    # Gross rent 135,000 | Loan 1,500,000 | DS 91,500 | NOI ~121k | DSCR ~1.33
    "costs": {**_COSTS_CLEAN, "capex_sqm": 3.00, "letting_pct": 0.020, "other_abs": 800},
},

{
    "id": 3, "difficulty": "BEGINNER", "verdict_expected": "PASS",
    "title": "REWE + dm Strip Centre — Kassel",
    "learning_objective": "Two-anchor NVZ with service unit: diversified but still clean.",
    "teaching_notes": (
        "DSCR ~1.47x, LTV 60%, WAULT ~5.8yr. "
        "REWE + dm together = institutional-grade income. Backshop at 4.5yr is the "
        "only shorter lease — not a risk given its small size (5% of income)."
    ),
    "deal": dict(
        property_name="REWE Nahversorgungszentrum Kassel-Bettenhausen",
        street_address="Fuldabrücker Strasse 44",
        city="Kassel / Hessen",
        asset_type="Nahversorgungszentrum",
        gla_sqm=3200, year_built=2016, parking=140,
        asking_price=7_200_000, equity=2_880_000,
        grest_rate=0.060, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2500 · 10.80 → 324,000 · expiry 2031-06 → 6yr
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",2500,10.80,
         date(2019,7,1),date(2031,6,30),None,"Y",0.75,0.10,24300,"Y",0,
         "Anchor — 6yr remaining"),
        # 500 · 15.50 → 93,000 · expiry 2030-12 → 5.5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",500,15.50,
         date(2020,1,1),date(2030,12,31),None,"Y",1.00,0.10,7750,"Y",0,
         "5.5yr remaining"),
        # 200 · 12.00 → 28,800 · expiry 2029-12 → 4.5yr
        ("U-03","Bäckerei Schäfer","Food Service","N",200,12.00,
         date(2021,1,1),date(2029,12,31),None,"N",0,0,2400,"Y",0,
         "Service unit — 4.5yr"),
    ],
    # Gross rent 445,800 | Loan 4,320,000 | DS 263,520 | NOI ~388k | DSCR ~1.47
    "costs": _COSTS_CLEAN,
},

{
    "id": 4, "difficulty": "INTERMEDIATE", "verdict_expected": "PASS",
    "title": "FMZ Nuremberg — REWE + dm + Action + KiK",
    "learning_objective": "Four-tenant FMZ: layered income stream, long WAULT.",
    "teaching_notes": (
        "DSCR ~1.55x, LTV 60%, WAULT ~7.5yr. Textbook Fachmarktzentrum with food "
        "anchor + drugstore + two value retailers. Anchor concentration 60% — inside "
        "threshold. Each lease expiry is staggered, reducing rollover risk."
    ),
    "deal": dict(
        property_name="Fachmarktzentrum Nuremberg-Langwasser",
        street_address="Langwasserstrasse 55",
        city="Nuremberg / Bayern",
        asset_type="Fachmarktzentrum (Retail Park)",
        gla_sqm=4650, year_built=2013, parking=200,
        asking_price=9_500_000, equity=3_800_000,
        grest_rate=0.035, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2800 · 11.00 → 369,600 · expiry 2033-09 → 8.25yr
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",2800,11.00,
         date(2019,10,1),date(2033,9,30),None,"Y",0.75,0.10,27720,"Y",0,
         "Anchor — 8.25yr remaining"),
        # 600 · 16.00 → 115,200 · expiry 2032-12 → 7.5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",600,16.00,
         date(2020,1,1),date(2032,12,31),None,"Y",1.00,0.10,9600,"Y",0,
         "7.5yr remaining"),
        # 800 · 9.50 → 91,200 · expiry 2031-03 → 5.75yr
        ("U-03","Action Deutschland GmbH","Non-Food Variety","N",800,9.50,
         date(2019,4,1),date(2031,3,31),None,"Y",1.00,0.10,7600,"Y",0,
         "5.75yr remaining"),
        # 450 · 8.00 → 43,200 · expiry 2030-06 → 5yr
        ("U-04","KiK Textilien","Discount Fashion","N",450,8.00,
         date(2020,7,1),date(2030,6,30),None,"N",0,0,3600,"Y",0,
         "5yr remaining"),
    ],
    # Gross rent 619,200 | Loan 5,700,000 | DS 347,700 | NOI ~539k | DSCR ~1.55
    "costs": _COSTS_CLEAN,
},

{
    "id": 5, "difficulty": "INTERMEDIATE", "verdict_expected": "PASS",
    "title": "ALDI + dm Strip — Cologne Porz",
    "learning_objective": "Dual-discounter strip: strong covenants, compact asset.",
    "teaching_notes": (
        "DSCR ~1.37x, LTV 60%, WAULT ~6.8yr. ALDI Süd + dm = two of the safest "
        "retail covenants in Germany. Small GLA keeps running costs low. "
        "GIY 5.75% is typical for prime discounter strips in NRW metros."
    ),
    "deal": dict(
        property_name="Discounter Strip Cologne-Porz",
        street_address="Weststrasse 89",
        city="Cologne / Nordrhein-Westfalen",
        asset_type="Lebensmittelmarkt (Strip)",
        gla_sqm=1500, year_built=2018, parking=65,
        asking_price=3_600_000, equity=1_440_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 1000 · 9.50 → 114,000 · expiry 2032-12 → 7.5yr
        ("U-01","ALDI Süd","Food & Grocery","Y",1000,9.50,
         date(2018,1,1),date(2032,12,31),None,"Y",0.75,0.10,9500,"Y",0,
         "7.5yr remaining — ALDI Süd KG guarantor"),
        # 500 · 15.50 → 93,000 · expiry 2031-06 → 6yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",500,15.50,
         date(2021,7,1),date(2031,6,30),None,"Y",1.00,0.10,7750,"Y",0,
         "6yr remaining"),
    ],
    # Gross rent 207,000 | Loan 2,160,000 | DS 131,760 | NOI ~180k | DSCR ~1.37
    "costs": {**_COSTS_CLEAN, "capex_sqm": 3.50, "other_abs": 1000},
},

{
    "id": 6, "difficulty": "INTERMEDIATE", "verdict_expected": "PASS",
    "title": "Kaufland FMZ — Chemnitz (East Germany)",
    "learning_objective": "Large anchor + small units: east Germany yield uplift.",
    "teaching_notes": (
        "DSCR ~1.58x, LTV 60%, WAULT ~7.3yr. Kaufland (Schwarz Group) is one of "
        "Germany's strongest grocery covenants. Secondary Sachsen location commands "
        "a higher yield (6.7%) vs prime Bayern — compensates for thinner liquidity."
    ),
    "deal": dict(
        property_name="Kaufland Fachmarktzentrum Chemnitz-Gablenz",
        street_address="Zwickauer Strasse 211",
        city="Chemnitz / Sachsen",
        asset_type="Fachmarktzentrum (Retail Park)",
        gla_sqm=7400, year_built=2010, parking=280,
        asking_price=9_800_000, equity=3_920_000,
        grest_rate=0.035, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 6000 · 7.50 → 540,000 · expiry 2033-03 → 7.75yr
        ("U-01","Kaufland Stiftung & Co. KG","Food & Grocery","Y",6000,7.50,
         date(2017,4,1),date(2033,3,31),None,"Y",0.75,0.10,45000,"Y",0,
         "Anchor — 7.75yr · Schwarz Group guarantor"),
        # 400 · 14.00 → 67,200 · expiry 2031-12 → 6.5yr
        ("U-02","Rossmann GmbH","Health & Beauty","N",400,14.00,
         date(2019,1,1),date(2031,12,31),None,"Y",1.00,0.10,5600,"Y",0,
         "6.5yr remaining"),
        # 500 · 7.50 → 45,000 · expiry 2030-06 → 5yr
        ("U-03","KiK Textilien","Discount Fashion","N",500,7.50,
         date(2020,7,1),date(2030,6,30),None,"N",0,0,3750,"Y",0,
         "5yr remaining"),
    ],
    # Gross rent 652,200 | Loan 5,880,000 | DS 358,680 | NOI ~567k | DSCR ~1.58
    "costs": _COSTS_CLEAN,
},

{
    "id": 7, "difficulty": "ADVANCED", "verdict_expected": "PASS",
    "title": "Premium FMZ — Munich Suburb",
    "learning_objective": "Strong metrics at tight yield: when a PASS still raises a pricing question.",
    "teaching_notes": (
        "DSCR ~1.35x, LTV 60%, WAULT ~8.4yr. All metrics green. "
        "But GIY 5.66% vs 4.1% debt cost = 156bps spread. Institutional minimum is "
        "200bps. The deal passes the credit test but fails most fund hurdle rates — "
        "this is the 'technically bankable but commercially unattractive' scenario."
    ),
    "deal": dict(
        property_name="Fachmarktzentrum Munich-Pasing Arcaden Umfeld",
        street_address="Bäckerstrasse 3",
        city="Munich / Bayern",
        asset_type="Fachmarktzentrum (Prime)",
        gla_sqm=5550, year_built=2018, parking=220,
        asking_price=15_500_000, equity=6_200_000,
        grest_rate=0.035, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 3500 · 12.50 → 525,000 · expiry 2034-06 → 9yr
        ("U-01","Edeka Minden-Hannover","Food & Grocery","Y",3500,12.50,
         date(2019,7,1),date(2034,6,30),None,"Y",0.75,0.10,39375,"Y",0,
         "Prime anchor — 9yr remaining"),
        # 700 · 17.50 → 147,000 · expiry 2033-12 → 8.5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",700,17.50,
         date(2019,1,1),date(2033,12,31),None,"Y",1.00,0.10,12250,"Y",0,
         "8.5yr remaining"),
        # 400 · 15.00 → 72,000 · expiry 2032-12 → 7.5yr
        ("U-03","Rossmann GmbH","Health & Beauty","N",400,15.00,
         date(2020,1,1),date(2032,12,31),None,"Y",1.00,0.10,6000,"Y",0,
         "7.5yr remaining"),
        # 600 · 11.00 → 79,200 · expiry 2031-06 → 6yr
        ("U-04","Action Deutschland GmbH","Non-Food Variety","N",600,11.00,
         date(2019,7,1),date(2031,6,30),None,"Y",1.00,0.10,6600,"Y",0,
         "6yr remaining"),
        # 350 · 13.00 → 54,600 · expiry 2030-12 → 5.5yr
        ("U-05","Fressnapf Tiernahrungs GmbH","Pet Retail","N",350,13.00,
         date(2020,1,1),date(2030,12,31),None,"Y",1.00,0.10,4550,"Y",0,
         "5.5yr remaining"),
    ],
    # Gross rent 877,800 | Loan 9,300,000 | DS 567,300 | NOI ~763k | DSCR ~1.35
    "costs": _COSTS_CLEAN,
},


# ─────────────────────────────────────────────────────────────────────────────
# INVESTIGATE DEALS  (8 – 14)
# One or two metrics outside green zone — not catastrophic, needs scrutiny.
# ─────────────────────────────────────────────────────────────────────────────

{
    "id": 8, "difficulty": "BEGINNER", "verdict_expected": "INVESTIGATE",
    "title": "Short WAULT — REWE + dm Bochum",
    "learning_objective": "Good income + financials, but lease expiry risk kills the WAULT.",
    "teaching_notes": (
        "DSCR ~1.41x, LTV 61%, but WAULT ~3.0yr — below the 5yr threshold. "
        "Income looks fine today; the problem is what happens at lease expiry in 2028. "
        "This is the most common borderline case in German retail underwriting."
    ),
    "deal": dict(
        property_name="REWE NVZ Bochum-Wattenscheid",
        street_address="Hordeler Strasse 55",
        city="Bochum / Nordrhein-Westfalen",
        asset_type="Nahversorgungszentrum",
        gla_sqm=2900, year_built=2012, parking=110,
        asking_price=6_500_000, equity=2_535_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2400 · 10.50 → 302,400 · expiry 2028-03 → 2.75yr  ← SHORT
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",2400,10.50,
         date(2018,4,1),date(2028,3,31),None,"Y",0.75,0.10,22680,"Y",0,
         "SHORT WAULT — expires Mar 2028. Renewal uncertain."),
        # 500 · 15.00 → 90,000 · expiry 2029-06 → 4yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",500,15.00,
         date(2021,7,1),date(2029,6,30),None,"Y",1.00,0.10,7500,"Y",0,
         "4yr remaining"),
    ],
    # Gross rent 392,400 | Loan 3,965,000 | DS 241,865 | NOI ~341k | DSCR ~1.41
    # WAULT = (302,400×2.75 + 90,000×4) / 392,400 = 3.04yr → INVESTIGATE
    "costs": _COSTS_CLEAN,
},

{
    "id": 9, "difficulty": "BEGINNER", "verdict_expected": "INVESTIGATE",
    "title": "High LTV — REWE NVZ Heidelberg",
    "learning_objective": "Thin DSCR + high LTV when equity is too small.",
    "teaching_notes": (
        "DSCR ~1.23x (just below 1.25 threshold), LTV 70%. "
        "The income is fine but the seller's price vs the equity committed creates "
        "too much leverage. This is the 'equity gap' scenario: right asset, wrong capital structure."
    ),
    "deal": dict(
        property_name="REWE Markt Heidelberg-Pfaffengrund",
        street_address="Eppelheimer Strasse 38",
        city="Heidelberg / Baden-Württemberg",
        asset_type="Nahversorgungszentrum",
        gla_sqm=2700, year_built=2014, parking=100,
        asking_price=6_200_000, equity=1_860_000,
        grest_rate=0.050, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2200 · 10.80 → 285,120 · expiry 2031-12 → 6.5yr
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",2200,10.80,
         date(2019,1,1),date(2031,12,31),None,"Y",0.75,0.10,23220,"Y",0,
         "Anchor — 6.5yr remaining"),
        # 500 · 15.00 → 90,000 · expiry 2030-06 → 5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",500,15.00,
         date(2020,7,1),date(2030,6,30),None,"Y",1.00,0.10,7500,"Y",0,
         "5yr remaining"),
    ],
    # Gross rent 375,120 | Loan 4,340,000 | DS 264,740 | NOI ~326k | DSCR ~1.23
    # LTV = 4,340,000 / 6,200,000 = 70% → INVESTIGATE
    "costs": _COSTS_CLEAN,
},

{
    "id": 10, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "Overpriced + 18% Vacancy — Kaufland FMZ Cottbus",
    "learning_objective": "Yield compression + vacancy + anchor concentration: three simultaneous yellow flags.",
    "teaching_notes": (
        "DSCR ~1.05x, GIY 5.37%, vacancy 17.8%, anchor 74%. "
        "The seller prices a non-prime Brandenburg asset at a 5.4% yield — too tight for this location. "
        "Combined with 18% void and Kaufland anchor concentration, this triggers three yellow warnings and a thin DSCR."
    ),
    "deal": dict(
        property_name="Kaufland Fachmarktzentrum Cottbus-Sandow",
        street_address="Sandower Strasse 44",
        city="Cottbus / Brandenburg",
        asset_type="Fachmarktzentrum (Retail Park)",
        gla_sqm=7300, year_built=2008, parking=260,
        asking_price=10_200_000, equity=4_080_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 4500 · 7.50 → 405,000 · expiry 2032-06 → 7yr
        ("U-01","Kaufland Stiftung & Co. KG","Food & Grocery","Y",4500,7.50,
         date(2018,7,1),date(2032,6,30),None,"Y",0.75,0.10,33750,"Y",0,
         "Anchor — 7yr remaining"),
        # 700 · 9.00 → 75,600 · expiry 2030-12 → 5.5yr
        ("U-02","Action Deutschland GmbH","Non-Food Variety","N",700,9.00,
         date(2020,1,1),date(2030,12,31),None,"Y",1.00,0.10,6300,"Y",0,
         "5.5yr remaining"),
        # 800 · 7.00 → 67,200 · expiry 2029-06 → 4yr
        ("U-03","NKD Vertriebs GmbH","Discount Fashion","N",800,7.00,
         date(2021,7,1),date(2029,6,30),None,"N",0,0,5600,"Y",0,
         "4yr remaining"),
        # 1300 sqm VACANT
        _vac("U-04", 1300, "Former electronics unit — dark 14 months"),
    ],
    # GLA 7300 | Vacant 1300 = 17.8% vacancy
    # Gross rent 547,800 | Loan 4,680,000 | DS 285,480 | NOI ~447k | DSCR ~1.57
    # Vacancy flag → INVESTIGATE
    "costs": {**_COSTS_ELEVATED, "void_svc_sqm": 25.0, "void_ins_sqm": 2.50},
},

{
    "id": 11, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "Thin DSCR — ALDI + Rossmann Halle",
    "learning_objective": "Good asset, wrong price — DSCR just below the 1.25x covenant.",
    "teaching_notes": (
        "DSCR ~1.13x, LTV 70%, WAULT ~6yr. The tenants are solid but the asking "
        "price implies a 5.55% GIY which — at 4.1% debt cost and 70% LTV — "
        "leaves only 13bps above the 1.0x floor. This is a price negotiation deal."
    ),
    "deal": dict(
        property_name="ALDI Nahversorgungsstrip Halle-Neustadt",
        street_address="Magistrale 19",
        city="Halle / Sachsen-Anhalt",
        asset_type="Nahversorgungszentrum",
        gla_sqm=1450, year_built=2016, parking=70,
        asking_price=3_200_000, equity=960_000,
        grest_rate=0.050, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 1000 · 8.50 → 102,000 · expiry 2031-06 → 6yr
        ("U-01","ALDI Süd","Food & Grocery","Y",1000,8.50,
         date(2019,7,1),date(2031,6,30),None,"Y",0.75,0.10,8500,"Y",0,
         "6yr remaining"),
        # 450 · 14.00 → 75,600 · expiry 2030-12 → 5.5yr
        ("U-02","Rossmann GmbH","Health & Beauty","N",450,14.00,
         date(2020,1,1),date(2030,12,31),None,"Y",1.00,0.10,6300,"Y",0,
         "5.5yr remaining"),
    ],
    # Gross rent 177,600 | Loan 2,240,000 | DS 136,640 | NOI ~154k | DSCR ~1.13
    # LTV = 2,240,000 / 3,200,000 = 70% → INVESTIGATE
    "costs": _COSTS_CLEAN,
},

{
    "id": 12, "difficulty": "INTERMEDIATE", "verdict_expected": "INVESTIGATE",
    "title": "Break Option Risk — Lidl Leipzig",
    "learning_objective": "Contract WAULT vs effective WAULT when a break option looms.",
    "teaching_notes": (
        "DSCR ~1.33x, LTV 60%, contract expiry 2033 → looks clean. "
        "But Lidl has a break option in Dec 2027 (2.5yr away). Effective WAULT = 2.5yr. "
        "Always underwrite to the break, not the expiry — this is the key lesson."
    ),
    "deal": dict(
        property_name="Lidl Markt Leipzig-Lausen",
        street_address="Georg-Schumann-Strasse 48",
        city="Leipzig / Sachsen",
        asset_type="Lebensmittelmarkt (Discounter)",
        gla_sqm=1750, year_built=2015, parking=90,
        asking_price=4_000_000, equity=1_600_000,
        grest_rate=0.035, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # Break option Dec-2027 = 2.5yr → effective WAULT 2.5yr even though expiry 2033
        ("U-01","Lidl Vertriebs-GmbH & Co.","Food & Grocery","Y",1300,9.00,
         date(2018,1,1),date(2033,12,31),date(2027,12,31),"Y",0.75,0.10,11700,"Y",0,
         "BREAK OPTION Dec-2027 — 2.5yr to break. Effective WAULT 2.5yr."),
        # 450 · 15.50 → 83,700 · expiry 2031-06 → 6yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",450,15.50,
         date(2021,7,1),date(2031,6,30),None,"Y",1.00,0.10,7750,"Y",0,
         "6yr remaining"),
    ],
    # Gross rent 224,100 | Loan 2,400,000 | DS 146,400 | NOI ~195k | DSCR ~1.33
    # Break option makes effective WAULT 2.5yr → INVESTIGATE
    "costs": _COSTS_CLEAN,
},

{
    "id": 13, "difficulty": "ADVANCED", "verdict_expected": "INVESTIGATE",
    "title": "Erbpacht (Ground Lease) — Edeka FMZ Hamburg",
    "learning_objective": "Ground lease erodes NOI and adds exit risk — changes the calculus.",
    "teaching_notes": (
        "DSCR ~1.17x, LTV 60%. Ground rent of EUR 92k/yr absorbs 20% of gross rent, "
        "compressing NOI significantly. Hamburg GrESt 5.5% adds acquisition cost. "
        "Erbpacht also creates exit risk — a lender can't take the land as security."
    ),
    "deal": dict(
        property_name="Edeka Fachmarktzentrum Hamburg-Rahlstedt",
        street_address="Rahlstedter Strasse 22",
        city="Hamburg / Hamburg",
        asset_type="Fachmarktzentrum (Erbpacht)",
        gla_sqm=3600, year_built=2011, parking=150,
        asking_price=7_500_000, equity=3_000_000,
        grest_rate=0.055, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 3000 · 10.00 → 360,000 · expiry 2032-06 → 7yr
        ("U-01","Edeka Minden-Hannover","Food & Grocery","Y",3000,10.00,
         date(2018,7,1),date(2032,6,30),None,"Y",0.75,0.10,30000,"Y",0,
         "Anchor — 7yr remaining"),
        # 600 · 15.00 → 108,000 · expiry 2031-12 → 6.5yr
        ("U-02","dm-drogerie markt","Health & Beauty","N",600,15.00,
         date(2019,1,1),date(2031,12,31),None,"Y",1.00,0.10,9000,"Y",0,
         "6.5yr remaining"),
    ],
    # Gross rent 468,000 | Ground rent 92,000 | Loan 4,500,000 | DS 274,500
    # NOI ≈ 468,000 × 0.87 − 92,000 = 315k | DSCR ~1.15
    "costs": {**_COSTS_CLEAN, "ground_rent_abs": 92_000, "other_abs": 2500},
},

{
    "id": 14, "difficulty": "ADVANCED", "verdict_expected": "INVESTIGATE",
    "title": "Anchor Expiry Risk — REWE 95% Concentration Hannover",
    "learning_objective": "When 95% of income expires in 2.25yr: concentration + near WAULT floor.",
    "teaching_notes": (
        "WAULT ~2.1yr (REWE expires Sep-2027). WAULT < 2.5yr = red fail. "
        "Anchor = 95% of income — if REWE departs, the property is effectively empty. "
        "Always underwrite to what happens at lease expiry, not the current passing income."
    ),
    "deal": dict(
        property_name="REWE SB-Markt Hannover-Linden",
        street_address="Limmerstrasse 71",
        city="Hannover / Niedersachsen",
        asset_type="Nahversorgungszentrum",
        gla_sqm=4700, year_built=2009, parking=180,
        asking_price=10_500_000, equity=4_200_000,
        grest_rate=0.050, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 4500 · 11.00 → 594,000 · expiry 2027-09 → 2.25yr → WAULT < 2.5 = RED FAIL
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",4500,11.00,
         date(2019,1,1),date(2028,9,30),None,"Y",0.75,0.10,49500,"Y",0,
         "95% of income. EXPIRES Sep-2027 — 2.25yr. No renewal signed."),
        # 200 · 12.00 → 28,800 · expiry 2029-06 → 4yr
        ("U-02","Lotto / Tabak","Personal Services","N",200,12.00,
         date(2021,7,1),date(2029,6,30),None,"N",0,0,2400,"N",0,
         "Service tenant — 4yr"),
    ],
    # Gross rent 622,800 | anchor_pct 95.4% | WAULT ~2.1yr → FAIL+RED → INVESTIGATE
    # Loan 6,300,000 | DS 384,300 | NOI ~542k | DSCR ~1.41
    "costs": _COSTS_CLEAN,
},


# ─────────────────────────────────────────────────────────────────────────────
# DECLINE DEALS  (15 – 20)
# Severe failure on one or more key metrics.
# ─────────────────────────────────────────────────────────────────────────────

{
    "id": 15, "difficulty": "BEGINNER", "verdict_expected": "DECLINE",
    "title": "77% Vacancy — Distressed NVZ Duisburg",
    "learning_objective": "When vacancy destroys income: negative NOI scenario.",
    "teaching_notes": (
        "DSCR ~0 (negative NOI), LTV 70%, WAULT ~2yr. Only Netto remains in a "
        "property that is 77% empty. Void costs exceed rental income. "
        "The seller is pricing on potential, not reality — the only correct response is DECLINE."
    ),
    "deal": dict(
        property_name="Stadtteilzentrum Duisburg-Meiderich",
        street_address="Emscherstrasse 71",
        city="Duisburg / Nordrhein-Westfalen",
        asset_type="Nahversorgungszentrum (distressed)",
        gla_sqm=3500, year_built=2003, parking=85,
        asking_price=3_500_000, equity=1_050_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 800 · 8.00 → 76,800 · expiry 2027-06 → 2yr
        ("U-01","Netto Marken-Discount","Food & Grocery","Y",800,8.00,
         date(2021,7,1),date(2027,6,30),None,"N",0,0,3840,"Y",0,
         "Only active tenant — 2yr remaining"),
        _vac("U-02", 1200, "Former anchor — dark 36 months"),
        _vac("U-03", 900, "Former fashion tenant — dark 24 months"),
        _vac("U-04", 600, "Former services — dark 12 months"),
    ],
    # GLA 3500 | Vacant 2700 = 77.1% vacancy
    # Gross rent 76,800 | void costs 2700 × 25 = 67,500
    # NOI ≈ 76,800×0.87 − 67,500 = −799 → DSCR ≈ 0 → DECLINE
    "costs": _COSTS_DISTRESSED,
},

{
    "id": 16, "difficulty": "BEGINNER", "verdict_expected": "DECLINE",
    "title": "WAULT 0.4yr — Expiring Anchor Wuppertal",
    "learning_objective": "Near-zero WAULT: the income stream ends before the loan does.",
    "teaching_notes": (
        "WAULT 0.4yr — the anchor expires in October 2025, 4 months from analysis date. "
        "Even with good DSCR today, a 10yr loan cannot be secured against 4 months of income. "
        "DECLINE is mandatory regardless of other metrics."
    ),
    "deal": dict(
        property_name="Nahversorgungszentrum Wuppertal-Langerfeld",
        street_address="Langerfelder Strasse 88",
        city="Wuppertal / Nordrhein-Westfalen",
        asset_type="Nahversorgungszentrum",
        gla_sqm=2350, year_built=2006, parking=95,
        asking_price=5_000_000, equity=1_500_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # expiry Oct-2025 → only 0.4yr remaining
        ("U-01","Woolworth GmbH & Co. KG","Discount Fashion","Y",2200,8.00,
         date(2015,11,1),date(2025,10,31),None,"N",0,0,5280,"N",0,
         "EXPIRES OCT-2025 — 0.4yr remaining. No renewal signed."),
        # expiry Dec-2025 → 0.5yr
        ("U-02","Schreibwaren Müller","Personal Services","N",150,10.00,
         date(2020,1,1),date(2025,12,31),None,"N",0,0,1500,"N",0,
         "Expires Dec-2025 — 0.5yr"),
    ],
    # WAULT = (211,200×0.4 + 18,000×0.5) / 229,200 = 0.43yr → DECLINE
    # DSCR ~0.93x (also below 1.0) → DECLINE
    "costs": _COSTS_ELEVATED,
},

{
    "id": 17, "difficulty": "INTERMEDIATE", "verdict_expected": "DECLINE",
    "title": "LTV 80% + DSCR 0.81x — Overleveraged NVZ Essen",
    "learning_objective": "Too much debt: even decent income can't service an oversized loan.",
    "teaching_notes": (
        "DSCR ~0.81x, LTV 80%. The tenants are reasonable but equity is only 20%. "
        "Every EUR of rent generates EUR 0.81 of coverage — the property loses money "
        "for the lender from day one. This is an equity story, not a debt story."
    ),
    "deal": dict(
        property_name="REWE NVZ Essen-Stoppenberg",
        street_address="Beuthstrasse 14",
        city="Essen / Nordrhein-Westfalen",
        asset_type="Nahversorgungszentrum",
        gla_sqm=2400, year_built=2010, parking=90,
        asking_price=5_500_000, equity=1_100_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2000 · 9.00 → 216,000 · expiry 2030-06 → 5yr
        ("U-01","REWE Markt GmbH","Food & Grocery","Y",2000,9.00,
         date(2020,7,1),date(2030,6,30),None,"Y",0.75,0.10,16200,"Y",0,
         "Anchor — 5yr remaining"),
        # 400 · 7.00 → 33,600 · expiry 2028-12 → 3.5yr
        ("U-02","KiK Textilien","Discount Fashion","N",400,7.00,
         date(2021,1,1),date(2028,12,31),None,"N",0,0,2800,"Y",0,
         "3.5yr remaining"),
    ],
    # Gross rent 249,600 | Loan 4,400,000 | DS 268,400 | NOI ~217k | DSCR ~0.81
    # LTV = 4,400,000 / 5,500,000 = 80% → DECLINE
    "costs": _COSTS_CLEAN,
},

{
    "id": 18, "difficulty": "INTERMEDIATE", "verdict_expected": "DECLINE",
    "title": "Fashion-Only Center — No Food Anchor Magdeburg",
    "learning_objective": "Sector risk: no food anchor, short leases, 26% vacancy.",
    "teaching_notes": (
        "DSCR ~0.85x, WAULT ~2.7yr, vacancy 25.8%. No food anchor = no defensive income. "
        "Post-2022 fashion discount sector (Takko, NKD, TEDi) is under structural pressure. "
        "Vacancy + short WAULT + no anchor + below-1.0 DSCR = four simultaneous failures."
    ),
    "deal": dict(
        property_name="Stadtpassage Magdeburg-Sudenburg",
        street_address="Halberstädter Strasse 66",
        city="Magdeburg / Sachsen-Anhalt",
        asset_type="Nahversorgungszentrum (distressed)",
        gla_sqm=3100, year_built=2002, parking=60,
        asking_price=4_000_000, equity=1_200_000,
        grest_rate=0.050, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 1200 · 7.00 → 100,800 · expiry 2028-06 → 3yr
        ("U-01","Takko Fashion","Discount Fashion","N",1200,7.00,
         date(2022,7,1),date(2028,6,30),None,"N",0,0,8400,"N",0,
         "3yr remaining — Takko covenant risk"),
        # 600 · 6.50 → 46,800 · expiry 2027-12 → 2.5yr
        ("U-02","NKD Vertriebs GmbH","Discount Fashion","N",600,6.50,
         date(2022,1,1),date(2027,12,31),None,"N",0,0,3900,"N",0,
         "2.5yr remaining"),
        # 500 · 7.00 → 42,000 · expiry 2027-06 → 2yr
        ("U-03","TEDi GmbH & Co. KG","Non-Food Variety","N",500,7.00,
         date(2022,7,1),date(2027,6,30),None,"N",0,0,3500,"N",0,
         "2yr remaining"),
        _vac("U-04", 800, "Former food anchor departed — dark 18 months"),
    ],
    # Gross rent 189,600 | Vacant 800/3100 = 25.8%
    # Loan 2,800,000 | DS 170,800 | void cost 800×25=20k | NOI ~145k | DSCR ~0.85
    "costs": {**_COSTS_DISTRESSED, "void_svc_sqm": 25.0, "void_ins_sqm": 2.50},
},

{
    "id": 19, "difficulty": "ADVANCED", "verdict_expected": "DECLINE",
    "title": "Triple Failure — Void + LTV + WAULT Bochum",
    "learning_objective": "Multiple simultaneous failures: the property generates negative NOI.",
    "teaching_notes": (
        "DSCR ≈ 0 (negative NOI), LTV 80%, WAULT 1.5yr. "
        "79% vacancy means void charges exceed rental income. "
        "LTV 80% means the loan far exceeds what the asset can support. "
        "Three red flags simultaneously — this is a clear distressed workout, not a purchase."
    ),
    "deal": dict(
        property_name="Einkaufszentrum Bochum-Hamme",
        street_address="Hammerstrasse 121",
        city="Bochum / Nordrhein-Westfalen",
        asset_type="Fachmarktzentrum (distressed)",
        gla_sqm=4400, year_built=1997, parking=130,
        asking_price=5_200_000, equity=1_040_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 900 · 7.00 → 75,600 · expiry 2026-12 → 1.5yr  ← very short
        ("U-01","Norma Lebensmittelfilialbetrieb","Food & Grocery","Y",900,7.00,
         date(2022,1,1),date(2026,12,31),None,"N",0,0,6300,"Y",0,
         "Norma — WAULT 1.5yr. Weak covenant."),
        _vac("U-02", 1500, "Former anchor — dark 36 months"),
        _vac("U-03", 1200, "Former fashion tenant — dark 28 months"),
        _vac("U-04", 800, "Former electronics — dark 20 months"),
    ],
    # Vacant 3500/4400 = 79.5% | void costs 3500×30=105k
    # Gross rent 75,600 | NOI ≈ 75,600×0.87 − 105,000 = −39k → DSCR ≈ 0
    # LTV = 4,160,000 / 5,200,000 = 80% → DECLINE (triple flag)
    "costs": _COSTS_DISTRESSED,
},

{
    "id": 20, "difficulty": "ADVANCED", "verdict_expected": "DECLINE",
    "title": "NIY Distressed + WAULT 0.5yr — Gera",
    "learning_objective": "High NIY signals distressed pricing; near-zero WAULT confirms it.",
    "teaching_notes": (
        "WAULT 0.5yr → DECLINE regardless of DSCR or LTV. "
        "The high NIY (>10%) tells you the market already prices in the income risk — "
        "sellers discount heavily when leases expire imminently. "
        "Even a 2.5x DSCR today means nothing when there is no lease in 6 months."
    ),
    "deal": dict(
        property_name="Nahversorgungszentrum Gera-Lusan",
        street_address="Wismutstrasse 4",
        city="Gera / Thüringen",
        asset_type="Nahversorgungszentrum (distressed)",
        gla_sqm=3000, year_built=2000, parking=110,
        asking_price=2_200_000, equity=880_000,
        grest_rate=0.065, notar_rate=0.015, broker_rate=0.010,
        analysis_date=REF, hold_period=10,
        interest_rate=0.041, loan_term_yrs=10,
        amort_type="Annuitaetendarlehen", amort_rate=0.020, io_period=2,
    ),
    "rent_roll": [
        # 2500 · 9.00 → 270,000 · expiry 2025-12 → 0.5yr!
        ("U-01","Woolworth GmbH & Co. KG","Discount Fashion","Y",2500,9.00,
         date(2020,1,1),date(2025,12,31),None,"N",0,0,13500,"N",0,
         "EXPIRES Dec-2025 — 0.5yr remaining. No renewal signed."),
        # 500 · 8.00 → 48,000 · expiry 2026-03 → 0.75yr
        ("U-02","Schuhcenter Gera","Footwear","N",500,8.00,
         date(2021,4,1),date(2026,3,31),None,"N",0,0,4000,"N",0,
         "0.75yr remaining"),
    ],
    # Gross rent 318,000 | GIY = 318,000/2,200,000 = 14.5% → NIY ~12.6% > 10%
    # WAULT = (270,000×0.5 + 48,000×0.75) / 318,000 = 0.54yr → DECLINE
    # Loan 1,320,000 | DS 80,520 | NOI ~277k | DSCR ~3.4x — irrelevant with 0.5yr lease
    "costs": _COSTS_ELEVATED,
},

]   # end PRACTICE_DEALS_100


# ─── Generator function ───────────────────────────────────────────────────────

def generate_practice_deal(deal_config: dict, output_dir: str = ".") -> str:
    """Write one practice deal to an Excel file and return the path."""
    from openpyxl import load_workbook as _lw

    d      = deal_config["deal"]
    rr     = deal_config["rent_roll"]
    costs  = deal_config["costs"]
    name   = (
        f"Practice_{deal_config['id']:02d}_"
        f"{deal_config['difficulty']}_"
        f"{deal_config['verdict_expected']}_"
        f"{''.join(c for c in deal_config['title'][:30] if c.isalnum() or c in ' _-').strip().replace(' ','_')}"
        ".xlsx"
    )
    out = Path(output_dir) / name
    shutil.copy2(MASTER_TEMPLATE, out)
    wb = _lw(out)

    # Deal Overview
    ws = wb["Deal Overview"]
    for field, (row, col) in DEAL_CELLS.items():
        val = d.get(field)
        if val is not None:
            ws.cell(row=row, column=col).value = val

    # Rent Roll
    ws_rr = wb["Rent Roll"]
    _clear_rent_roll(ws_rr)
    for i, t in enumerate(rr):
        row = RR_FIRST_DATA_ROW + i
        (unit_ref, tenant, sector, anchor, area, rent_mo,
         start, expiry, brk, cpi, cpi_pt, cpi_tr,
         sec_dep, svc_rec, rf_mo, notes) = t
        ws_rr.cell(row=row, column=RR_COLS["unit_ref"]).value        = unit_ref
        ws_rr.cell(row=row, column=RR_COLS["tenant"]).value          = tenant
        ws_rr.cell(row=row, column=RR_COLS["sector"]).value          = sector
        ws_rr.cell(row=row, column=RR_COLS["anchor"]).value          = anchor
        ws_rr.cell(row=row, column=RR_COLS["area_sqm"]).value        = area
        ws_rr.cell(row=row, column=RR_COLS["rent_sqm_mo"]).value     = rent_mo
        ws_rr.cell(row=row, column=RR_COLS["lease_start"]).value     = start
        ws_rr.cell(row=row, column=RR_COLS["lease_expiry"]).value    = expiry
        ws_rr.cell(row=row, column=RR_COLS["break_option"]).value    = brk
        ws_rr.cell(row=row, column=RR_COLS["cpi_clause"]).value      = cpi
        ws_rr.cell(row=row, column=RR_COLS["cpi_passthrough"]).value = cpi_pt
        ws_rr.cell(row=row, column=RR_COLS["cpi_trigger"]).value     = cpi_tr
        ws_rr.cell(row=row, column=RR_COLS["security_deposit"]).value = sec_dep
        ws_rr.cell(row=row, column=RR_COLS["svc_chg_recov"]).value   = svc_rec
        ws_rr.cell(row=row, column=RR_COLS["rent_free_months"]).value = rf_mo
        ws_rr.cell(row=row, column=RR_COLS["notes"]).value           = notes

    # Cost Sheet
    ws_c = wb["Cost Sheet"]
    for key, row_num in COST_ROWS.items():
        ws_c.cell(row=row_num, column=3).value = costs.get(key, 0)

    wb.save(out)
    return str(out)


def run_all(output_dir=".", verdict_filter=None, verbose=True):
    """Generate all practice deals, optionally filtered by verdict."""
    results = []
    deals = [
        d for d in PRACTICE_DEALS_100
        if verdict_filter is None or d["verdict_expected"] == verdict_filter
    ]
    for deal in deals:
        try:
            path = generate_practice_deal(deal, output_dir)
            if verbose:
                print(f"  [OK] Deal {deal['id']:02d} {deal['verdict_expected']:<12} "
                      f"{deal['difficulty']:<14} {deal['title'][:45]}")
            results.append({"id": deal["id"], "ok": True, "path": path})
        except Exception as e:
            if verbose:
                print(f"  [ERR] Deal {deal['id']:02d} — {e}")
            results.append({"id": deal["id"], "ok": False, "error": str(e)})
    if verbose:
        ok = sum(1 for r in results if r["ok"])
        print(f"\n{ok}/{len(results)} deals generated successfully.")
    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate 20 German retail practice deals")
    parser.add_argument("--deal", type=int, help="Generate a single deal by ID")
    parser.add_argument("--verdict", choices=["PASS", "INVESTIGATE", "DECLINE"],
                        help="Filter by verdict")
    parser.add_argument("--out", default=".", help="Output directory")
    args = parser.parse_args()

    import argparse as _ap  # already imported at top, just guarding

    print(f"GRU Practice Deal Generator — {len(PRACTICE_DEALS_100)} deals")
    print()

    if args.deal:
        d = next((x for x in PRACTICE_DEALS_100 if x["id"] == args.deal), None)
        if not d:
            print(f"Deal {args.deal} not found.")
        else:
            path = generate_practice_deal(d, args.out)
            print(f"Generated: {Path(path).name}")
    else:
        run_all(output_dir=args.out, verdict_filter=args.verdict)
