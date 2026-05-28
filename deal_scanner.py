"""
deal_scanner.py  —  GRU Pipeline: German Retail Property Listing Scanner
=========================================================================
Searches ImmoScout24 Gewerbe and Immowelt Gewerbe for retail listings,
filters against GRU investment criteria, scores each deal 0-100, and
pre-populates the UW template for high-scoring candidates.

Usage
-----
    python deal_scanner.py                    # incremental (skip today if run already)
    python deal_scanner.py --force            # always run, overwrite today's file
    python deal_scanner.py --quiet            # suppress progress output
    python deal_scanner.py --show             # print latest report and exit

Important notes on scraping
----------------------------
• robots.txt is checked for every domain before any GET request.
• A 2-second delay is enforced between all requests.
• ImmoScout24 and Immowelt use client-side React rendering; a plain
  requests.get() typically returns an HTML shell only. Where possible
  this module attempts to extract embedded JSON (Next.js / __NEXT_DATA__
  patterns). If the site blocks or returns no listings the error is
  logged and the run continues.
• For full data, consider their official broker/API programmes or a
  headless browser (Playwright/Selenium).
"""

import argparse
import datetime
import json
import logging
import pathlib
import re
import shutil
import sys
import time
from typing import Optional
from urllib.parse import urljoin, urlencode, urlparse
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup

from constants import GREST_BY_STATE, GREST_DEFAULT

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_BASE     = pathlib.Path(__file__).parent
OUTPUT_DIR = _BASE / "deal_pipeline"
PRE_UW_DIR = OUTPUT_DIR / "pre_underwritten"
TEMPLATE   = _BASE / "DE_Retail_Underwriter_Template.xlsx"

# ---------------------------------------------------------------------------
# Filter / scoring configuration
# ---------------------------------------------------------------------------
FILTER = {
    "min_price":         1_000_000,
    "max_price":        15_000_000,
    "min_gla":               500,
    "max_gla":            15_000,
    "preferred_states": [
        "Bayern", "Hamburg", "Hessen",
        "Nordrhein-Westfalen", "Baden-Württemberg",
    ],
    "min_giy_if_stated": 0.055,
    "required_keywords": [
        "Lebensmittel", "Rewe", "Edeka", "Lidl", "Aldi", "Netto", "Penny",
        "Nahversorgung", "Fachmarkt", "Decathlon", "OBI", "Bauhaus",
        "MediaMarkt", "Saturn", "ATU", "Fressnapf", "JYSK", "Action",
        "Woolworth", "Jawoll", "Takko", "KiK", "dm", "Rossmann", "Müller",
    ],
    # Any single tenant with WAULT > 8yr AND GIY > 6% passes regardless of sector
    "anchor_quality_override":      True,
    "anchor_override_min_wault":    8.0,
    "anchor_override_min_giy":      0.06,
    "pre_populate_threshold":       70,   # score ≥ this → pre-fill Excel
}

SEARCH_TERMS = [
    "Einzelhandel", "Fachmarkt", "Nahversorgung", "Supermarkt", "Discounter",
]

REQUEST_DELAY = 2.0   # seconds between requests
BOT_UA = (
    "Mozilla/5.0 (compatible; GRU-DealScanner/1.0; "
    "+https://github.com/lahav-lr/gru; research use)"
)

# ---------------------------------------------------------------------------
# Site definitions
# ---------------------------------------------------------------------------
SITES = [
    {
        # ImmobilienScout24 — their /Suche/ API endpoints require auth (401).
        # /gewerbe.html is the public landing page; actual search requires login.
        # Kept here so the scanner attempts the page and logs the result.
        "name":       "ImmobilienScout24",
        "base_url":   "https://www.immobilienscout24.de",
        "robots_url": "https://www.immobilienscout24.de/robots.txt",
        "search_urls": [
            "https://www.immobilienscout24.de/gewerbe/einzelhandel.html",
        ],
        "parser": "immoscout24",
        "wait_selectors": [
            "ul.result-list",
            "div[class*='result-list']",
            "article[data-id]",
            "li[data-item-id]",
        ],
    },
    {
        # Immowelt — /gewerbeimmobilien confirmed 200; JS renders listing cards.
        # Note: serp-bff/estates returns 403 (CSRF-protected); XHR interception
        # captures analytics but not estate listings. Kept for DOM fallback.
        "name":       "Immowelt",
        "base_url":   "https://www.immowelt.de",
        "robots_url": "https://www.immowelt.de/robots.txt",
        "search_urls": [
            "https://www.immowelt.de/gewerbeimmobilien",
        ],
        "parser": "immowelt",
        "wait_selectors": [
            "article[class*='estate']",
            "article[class*='Estate']",
            "div[class*='EstateItem']",
            "div[class*='estateItem']",
            "div[class*='estate-item']",
            "article",
            "div[data-testid*='estate']",
            "div[data-testid*='listing']",
        ],
    },
    {
        # Kleinanzeigen (formerly eBay Kleinanzeigen) — retail Gewerbeimmobilien
        # category. Confirmed 200 with 27 article.aditem elements per page.
        # Mix of investment sales and rental units; price filter removes rentals.
        "name":       "Kleinanzeigen",
        "base_url":   "https://www.kleinanzeigen.de",
        "robots_url": "https://www.kleinanzeigen.de/robots.txt",
        "search_urls": [
            "https://www.kleinanzeigen.de/s-gewerbeimmobilien/kaufen/c277+gewerbeimmobilien.objektart_s:einzelhandel_kioske",
            "https://www.kleinanzeigen.de/s-gewerbeimmobilien/kaufen/c277+gewerbeimmobilien.objektart_s:gewerbeeinheit",
        ],
        "parser": "kleinanzeigen",
        "wait_selectors": [
            "article.aditem",
            "article[class*='aditem']",
            "#srchrslt-adtable",
        ],
    },
]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
log = logging.getLogger("deal_scanner")
_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(logging.Formatter("  %(message)s"))
# Safe emit: replaces unencodable chars so Windows cp1255 consoles don't crash
_orig_emit = _handler.emit
def _safe_emit(record):
    try:
        _orig_emit(record)
    except UnicodeEncodeError:
        try:
            msg = _handler.format(record).encode("ascii", errors="replace").decode("ascii")
            sys.stdout.write(msg + "\n")
            sys.stdout.flush()
        except Exception:
            pass
_handler.emit = _safe_emit
log.addHandler(_handler)
log.setLevel(logging.INFO)


# ===========================================================================
# robots.txt helper
# ===========================================================================
class _RobotsCache:
    """Lazy, per-domain robots.txt cache."""
    def __init__(self):
        self._cache: dict[str, RobotFileParser] = {}

    def allowed(self, url: str, session: requests.Session) -> bool:
        parsed   = urlparse(url)
        domain   = f"{parsed.scheme}://{parsed.netloc}"
        robots_url = urljoin(domain, "/robots.txt")
        if domain not in self._cache:
            rp = RobotFileParser(robots_url)
            try:
                resp = session.get(robots_url, timeout=8)
                rp.parse(resp.text.splitlines())
                log.info(f"robots.txt fetched: {robots_url}")
            except Exception as exc:
                log.warning(f"Could not fetch robots.txt ({robots_url}): {exc} — assuming allowed")
            self._cache[domain] = rp
            time.sleep(REQUEST_DELAY)
        return self._cache[domain].can_fetch(BOT_UA, url)


# ===========================================================================
# HTTP session
# ===========================================================================
def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent":      BOT_UA,
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate",
    })
    return s


# ===========================================================================
# Text extraction helpers
# ===========================================================================
_PRICE_RE = re.compile(
    r'(?:EUR|€)\s*([\d]{1,3}(?:[.,\s][\d]{3})*(?:[.,][\d]{1,2})?)'
    r'|'
    r'([\d]{1,3}(?:[.,\s][\d]{3})*(?:[.,][\d]{1,2})?)\s*(?:EUR|€)',
    re.IGNORECASE,
)
_GLA_RE = re.compile(
    r'([\d]{3,6}(?:[.,][\d]{1,2})?)\s*m[²2]',
    re.IGNORECASE,
)
_GIY_RE = re.compile(
    r'(?:Rendite|GIY|Mietrendite|Jahresrohrendite|Nettomietrendite)[^\d]*([\d]{1,2}[.,][\d]{0,2})\s*%',
    re.IGNORECASE,
)
_WAULT_RE = re.compile(
    r'(?:WALT|WAULT|Mietlaufzeit|Restlaufzeit)[^\d]*([\d]{1,2}(?:[.,][\d])?)\s*(?:Jahre?|J\.)',
    re.IGNORECASE,
)
_YEAR_RE = re.compile(r'\b(19[0-9]{2}|20[0-2][0-9])\b')

# City → Bundesland map (major German cities; extend as needed)
_CITY_STATE: dict[str, str] = {
    "münchen": "Bayern", "nürnberg": "Bayern", "augsburg": "Bayern",
    "ingolstadt": "Bayern", "regensburg": "Bayern", "würzburg": "Bayern",
    "erlangen": "Bayern", "fürth": "Bayern",
    "hamburg": "Hamburg",
    "frankfurt": "Hessen", "wiesbaden": "Hessen", "kassel": "Hessen",
    "darmstadt": "Hessen", "offenbach": "Hessen",
    "köln": "Nordrhein-Westfalen", "düsseldorf": "Nordrhein-Westfalen",
    "dortmund": "Nordrhein-Westfalen", "essen": "Nordrhein-Westfalen",
    "duisburg": "Nordrhein-Westfalen", "bochum": "Nordrhein-Westfalen",
    "wuppertal": "Nordrhein-Westfalen", "bonn": "Nordrhein-Westfalen",
    "münster": "Nordrhein-Westfalen", "bielefeld": "Nordrhein-Westfalen",
    "aachen": "Nordrhein-Westfalen", "gelsenkirchen": "Nordrhein-Westfalen",
    "dormagen": "Nordrhein-Westfalen", "kerken": "Nordrhein-Westfalen",
    "borken": "Nordrhein-Westfalen",
    "stuttgart": "Baden-Württemberg", "mannheim": "Baden-Württemberg",
    "karlsruhe": "Baden-Württemberg", "freiburg": "Baden-Württemberg",
    "heidelberg": "Baden-Württemberg", "heilbronn": "Baden-Württemberg",
    "ulm": "Baden-Württemberg",
    "berlin": "Berlin",
    "leipzig": "Sachsen", "dresden": "Sachsen", "chemnitz": "Sachsen",
    "hannover": "Niedersachsen", "braunschweig": "Niedersachsen",
    "osnabrück": "Niedersachsen", "oldenburg": "Niedersachsen",
    "bremen": "Bremen",
    "magdeburg": "Sachsen-Anhalt", "halle": "Sachsen-Anhalt",
    "erfurt": "Thüringen", "jena": "Thüringen",
    "rostock": "Mecklenburg-Vorpommern", "schwerin": "Mecklenburg-Vorpommern",
    "kiel": "Schleswig-Holstein", "lübeck": "Schleswig-Holstein",
    "saarbrücken": "Saarland",
    "potsdam": "Brandenburg", "cottbus": "Brandenburg",
}


def _guess_bundesland(text: str) -> str:
    """Best-effort Bundesland from any city-like text."""
    if not text:
        return ""
    lower = text.lower()
    for city, state in _CITY_STATE.items():
        if city in lower:
            return state
    return ""


def _parse_price(text: str) -> Optional[float]:
    for m in _PRICE_RE.finditer(text):
        raw = (m.group(1) or m.group(2) or "").strip()
        raw = re.sub(r'[\s]', '', raw).replace('.', '').replace(',', '.')
        try:
            v = float(raw)
            # Normalise: if < 10 000 it was probably stated in thousands (€)
            if v < 10_000:
                v *= 1_000
            if FILTER["min_price"] * 0.5 <= v <= FILTER["max_price"] * 3:
                return v
        except ValueError:
            pass
    return None


def _parse_gla(text: str) -> Optional[float]:
    for m in _GLA_RE.finditer(text):
        raw = m.group(1).replace('.', '').replace(',', '.')
        try:
            v = float(raw)
            if 50 < v < 100_000:
                return v
        except ValueError:
            pass
    return None


def _parse_giy(text: str) -> Optional[float]:
    m = _GIY_RE.search(text)
    if m:
        try:
            v = float(m.group(1).replace(',', '.'))
            if 1 < v < 25:
                return round(v / 100, 4)
        except ValueError:
            pass
    return None


def _parse_wault(text: str) -> Optional[float]:
    m = _WAULT_RE.search(text)
    if m:
        try:
            v = float(m.group(1).replace(',', '.'))
            if 0 < v < 30:
                return round(v, 1)
        except ValueError:
            pass
    return None


def _parse_year_built(text: str) -> Optional[int]:
    m = _YEAR_RE.search(text)
    if m:
        yr = int(m.group(1))
        if 1900 <= yr <= datetime.date.today().year:
            return yr
    return None


def _find_keywords(text: str) -> list:
    """Return FILTER keywords found in text. Uses word-boundary matching to
    avoid false positives (e.g. 'obi' inside 'Immobilien', 'dm' inside 'dm²').
    """
    results = []
    for kw in FILTER["required_keywords"]:
        pattern = r'(?<![A-Za-z\d])' + re.escape(kw) + r'(?![A-Za-z\d])'
        if re.search(pattern, text, re.IGNORECASE):
            results.append(kw)
    return results


def _extract_anchor(keywords: list) -> Optional[str]:
    for kw in keywords:
        if kw.lower() in {
            "rewe", "edeka", "lidl", "aldi", "netto", "penny",
            "kaufland", "obi", "bauhaus", "mediamarkt", "saturn",
            "dm", "rossmann", "müller",
        }:
            return kw
    return None


# ===========================================================================
# HTML parsers — one per site
# ===========================================================================
def _try_next_data(soup: BeautifulSoup) -> Optional[dict]:
    """Extract embedded Next.js __NEXT_DATA__ JSON if present."""
    tag = soup.find("script", {"id": "__NEXT_DATA__"})
    if tag and tag.string:
        try:
            return json.loads(tag.string)
        except json.JSONDecodeError:
            pass
    return None


def _try_window_state(html: str) -> Optional[dict]:
    """Try to extract window.__INITIAL_STATE__ or similar embedded JSON."""
    patterns = [
        r'window\.__INITIAL_STATE__\s*=\s*({.+?});',
        r'window\.__APP_STATE__\s*=\s*({.+?});',
        r'"@type"\s*:\s*"ItemList"(.{0,5000})',
    ]
    for pat in patterns:
        m = re.search(pat, html, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except json.JSONDecodeError:
                pass
    return None


def _parse_immoscout24(html: str, base_url: str) -> list:
    """
    Parse ImmoScout24 search result page. Returns list of raw listing dicts.

    ImmoScout24 is a React SPA. A plain GET typically returns an HTML shell
    with <div id="root"> and no listing data. We attempt:
    1. __NEXT_DATA__ embedded JSON
    2. ld+json schema.org markup
    3. CSS selectors for static listing cards (rarely present)
    """
    listings = []
    soup     = BeautifulSoup(html, "html.parser")

    # ── Attempt 1: Next.js SSR data ──────────────────────────────────────────
    next_data = _try_next_data(soup)
    if next_data:
        log.info("ImmoScout24: found __NEXT_DATA__ block")
        items = (
            next_data.get("props", {})
                     .get("pageProps", {})
                     .get("searchResult", {})
                     .get("resultlist", {})
                     .get("resultlistEntries", [{}])[0]
                     .get("resultlistEntry", [])
        )
        for item in items:
            expose = item.get("resultlistEntry", item)
            addr   = expose.get("address", {})
            re2    = expose.get("realEstate", {})
            text   = json.dumps(expose)
            listings.append({
                "title":       re2.get("title", ""),
                "address":     f"{addr.get('street','')} {addr.get('houseNumber','')}".strip(),
                "city":        addr.get("city", ""),
                "bundesland":  _guess_bundesland(addr.get("city", "")),
                "asking_price": _parse_price(str(re2.get("price", {}).get("value", ""))),
                "gla_sqm":     _parse_gla(str(re2.get("livingSpace") or re2.get("plotArea") or "")),
                "year_built":  re2.get("constructionYear"),
                "anchor_tenant": _extract_anchor(_find_keywords(text)),
                "giy_stated":  _parse_giy(text),
                "wault_stated": _parse_wault(text),
                "broker":      expose.get("realtor", {}).get("companyName", ""),
                "url":         urljoin(base_url, f"/expose/{expose.get('id', '')}"),
                "source":      "ImmoScout24",
                "raw_text":    text[:500],
                "keywords":    _find_keywords(text),
            })
        if listings:
            return listings

    # ── Attempt 2: schema.org JSON-LD ────────────────────────────────────────
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(tag.string or "")
            if isinstance(data, list):
                data = data[0]
            if data.get("@type") in ("Product", "RealEstateListing", "Offer"):
                text = json.dumps(data)
                listings.append({
                    "title":        data.get("name", ""),
                    "address":      str(data.get("address", "")),
                    "city":         data.get("address", {}).get("addressLocality", ""),
                    "bundesland":   _guess_bundesland(
                        data.get("address", {}).get("addressLocality", "")
                    ),
                    "asking_price": _parse_price(str(data.get("offers", {}).get("price", ""))),
                    "gla_sqm":      None,
                    "year_built":   None,
                    "anchor_tenant": _extract_anchor(_find_keywords(text)),
                    "giy_stated":   _parse_giy(text),
                    "wault_stated": _parse_wault(text),
                    "broker":       "",
                    "url":          data.get("url", ""),
                    "source":       "ImmoScout24",
                    "raw_text":     text[:500],
                    "keywords":     _find_keywords(text),
                })
        except (json.JSONDecodeError, AttributeError):
            pass
    if listings:
        return listings

    # ── Attempt 3: static HTML cards ─────────────────────────────────────────
    for card in soup.select(
        "article[data-id], li[data-item-id], div[class*='listing'], "
        "div[class*='result-list'], li[class*='result']"
    ):
        text  = card.get_text(" ", strip=True)
        if len(text) < 20:
            continue
        href  = card.find("a")
        url   = urljoin(base_url, href["href"]) if href and href.get("href") else ""
        listings.append({
            "title":        text[:80],
            "address":      "",
            "city":         "",
            "bundesland":   _guess_bundesland(text),
            "asking_price": _parse_price(text),
            "gla_sqm":      _parse_gla(text),
            "year_built":   _parse_year_built(text),
            "anchor_tenant": _extract_anchor(_find_keywords(text)),
            "giy_stated":   _parse_giy(text),
            "wault_stated": _parse_wault(text),
            "broker":       "",
            "url":          url,
            "source":       "ImmobilienScout24",
            "raw_text":     text[:500],
            "keywords":     _find_keywords(text),
        })

    # ── Diagnostic ────────────────────────────────────────────────────────────
    if not listings:
        script_tags = soup.find_all("script")
        _has_next   = any(t.get("id") == "__NEXT_DATA__" for t in script_tags)
        _has_ld     = any(t.get("type") == "application/ld+json" for t in script_tags)
        log.info(
            f"  IS24: 0 listings parsed. HTML={len(html):,} chars, "
            f"scripts={len(script_tags)}, __NEXT_DATA__={_has_next}, ld+json={_has_ld}. "
            f"Page likely requires JavaScript rendering."
        )
    return listings


def _parse_immowelt(html: str, base_url: str) -> list:
    """
    Parse Immowelt search result page. Same JS-SPA caveat as ImmoScout24.

    Attempts (in order):
    1. __NEXT_DATA__ embedded JSON (Next.js SSR)
    2. application/ld+json schema.org ItemList / ListItem data
    3. Inline JSON blobs containing 'estates' or 'estateList' keys
    4. Static HTML listing cards
    """
    listings = []
    soup     = BeautifulSoup(html, "html.parser")

    # ── Attempt 1: Next.js SSR ───────────────────────────────────────────────
    next_data = _try_next_data(soup)
    if next_data:
        log.info("Immowelt: found __NEXT_DATA__ block")
        # Try several common key paths
        for _path in [
            ["props", "pageProps", "estateList", "estates"],
            ["props", "pageProps", "searchResult", "estates"],
            ["props", "pageProps", "items"],
        ]:
            items = next_data
            for k in _path:
                items = items.get(k, {}) if isinstance(items, dict) else {}
            if isinstance(items, list) and items:
                break
        else:
            items = []

        for item in items:
            text  = json.dumps(item)
            addr  = item.get("address", {})
            listings.append({
                "title":        item.get("title", ""),
                "address":      f"{addr.get('street','')} {addr.get('houseNumber','')}".strip(),
                "city":         addr.get("city", ""),
                "bundesland":   _guess_bundesland(addr.get("city", "")),
                "asking_price": _parse_price(str(item.get("prices", {}).get("buy", {}).get("amount", ""))),
                "gla_sqm":      _parse_gla(str(item.get("areas", {}).get("living") or
                                               item.get("areas", {}).get("total") or "")),
                "year_built":   item.get("yearOfConstruction"),
                "anchor_tenant": _extract_anchor(_find_keywords(text)),
                "giy_stated":   _parse_giy(text),
                "wault_stated": _parse_wault(text),
                "broker":       item.get("broker", {}).get("companyName", ""),
                "url":          urljoin(base_url, f"/expose/{item.get('id', '')}"),
                "source":       "Immowelt",
                "raw_text":     text[:500],
                "keywords":     _find_keywords(text),
            })
        if listings:
            return listings

    # ── Attempt 2: schema.org JSON-LD (ItemList with individual ListItems) ───
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(tag.string or "")
            if not isinstance(data, (dict, list)):
                continue
            objs = data if isinstance(data, list) else [data]
            for obj in objs:
                if obj.get("@type") == "ItemList":
                    for elem in obj.get("itemListElement", []):
                        item = elem.get("item", elem)
                        text = json.dumps(item)
                        addr = item.get("address", {})
                        city = (addr.get("addressLocality") or
                                addr.get("addressRegion") or "")
                        listings.append({
                            "title":        item.get("name", ""),
                            "address":      str(addr.get("streetAddress", "")),
                            "city":         city,
                            "bundesland":   _guess_bundesland(city),
                            "asking_price": _parse_price(
                                str(item.get("offers", {}).get("price", ""))
                            ),
                            "gla_sqm":      None,
                            "year_built":   None,
                            "anchor_tenant": _extract_anchor(_find_keywords(text)),
                            "giy_stated":   _parse_giy(text),
                            "wault_stated": _parse_wault(text),
                            "broker":       "",
                            "url":          item.get("url", ""),
                            "source":       "Immowelt",
                            "raw_text":     text[:500],
                            "keywords":     _find_keywords(text),
                        })
        except (json.JSONDecodeError, AttributeError, TypeError):
            pass
    if listings:
        return listings

    # ── Attempt 3: inline script blobs with estate arrays ────────────────────
    _blob_re = re.compile(
        r'"(?:estates|estateList|items|resultList)"\s*:\s*(\[.+?\])',
        re.DOTALL,
    )
    for tag in soup.find_all("script"):
        text_content = tag.string or ""
        m = _blob_re.search(text_content)
        if m:
            try:
                items = json.loads(m.group(1))
                for item in items[:50]:
                    t = json.dumps(item)
                    addr = item.get("address", {})
                    city = addr.get("city") or addr.get("addressLocality") or ""
                    listings.append({
                        "title":        item.get("title") or item.get("name") or "",
                        "address":      str(addr.get("street", "")),
                        "city":         city,
                        "bundesland":   _guess_bundesland(city),
                        "asking_price": _parse_price(str(
                            item.get("price") or
                            item.get("prices", {}).get("buy", {}).get("amount", "")
                        )),
                        "gla_sqm":      _parse_gla(str(
                            item.get("livingSpace") or item.get("area") or ""
                        )),
                        "year_built":   item.get("yearOfConstruction"),
                        "anchor_tenant": _extract_anchor(_find_keywords(t)),
                        "giy_stated":   _parse_giy(t),
                        "wault_stated": _parse_wault(t),
                        "broker":       "",
                        "url":          urljoin(base_url, str(item.get("url") or "")),
                        "source":       "Immowelt",
                        "raw_text":     t[:500],
                        "keywords":     _find_keywords(t),
                    })
            except (json.JSONDecodeError, TypeError):
                pass
    if listings:
        return listings

    # ── Attempt 4: static HTML cards ─────────────────────────────────────────
    _selectors = [
        "article",
        "div[class*='EstateItem']", "div[class*='estate-item']",
        "div[class*='EstateCard']", "div[class*='estate-card']",
        "li[class*='search-result']", "li[class*='SearchResult']",
        "div[data-testid*='listing']", "div[data-testid*='estate']",
        "div[class*='result-list-entry']",
    ]
    for sel in _selectors:
        for card in soup.select(sel):
            text = card.get_text(" ", strip=True)
            if len(text) < 20:
                continue
            href = card.find("a")
            url  = urljoin(base_url, href["href"]) if href and href.get("href") else ""
            listings.append({
                "title":        text[:80],
                "address":      "",
                "city":         "",
                "bundesland":   _guess_bundesland(text),
                "asking_price": _parse_price(text),
                "gla_sqm":      _parse_gla(text),
                "year_built":   _parse_year_built(text),
                "anchor_tenant": _extract_anchor(_find_keywords(text)),
                "giy_stated":   _parse_giy(text),
                "wault_stated": _parse_wault(text),
                "broker":       "",
                "url":          url,
                "source":       "Immowelt",
                "raw_text":     text[:500],
                "keywords":     _find_keywords(text),
            })
        if listings:
            return listings

    # ── Diagnostic: log what scripts were present ────────────────────────────
    script_tags = soup.find_all("script")
    _has_next   = any(t.get("id") == "__NEXT_DATA__" for t in script_tags)
    _has_ld     = any(t.get("type") == "application/ld+json" for t in script_tags)
    log.info(
        f"  Immowelt: 0 listings parsed. HTML={len(html):,} chars, "
        f"scripts={len(script_tags)}, __NEXT_DATA__={_has_next}, ld+json={_has_ld}. "
        f"Page likely requires JavaScript rendering."
    )
    return []


def _parse_kleinanzeigen(html: str, base_url: str) -> list:
    """
    Parse Kleinanzeigen commercial property search results.

    Each result is an <article class="aditem"> with predictable CSS classes:
      - a.ellipsis               → title + href
      - div.aditem-main--top--left → location (postal code + city)
      - p.aditem-main--middle--price-shipping--price → price
      - p.aditem-main--middle--description → description snippet
      - span.simpletag           → tags (area, rooms, etc.)

    Prices are German-formatted: "449.000 €" = 449,000 EUR. Values below
    50,000 after parsing are skipped (likely monthly rents not sale prices).
    """
    listings = []
    soup = BeautifulSoup(html, "html.parser")
    articles = soup.find_all("article", {"class": re.compile(r"aditem")})

    for art in articles:
        title_el = art.find("a", {"class": "ellipsis"})
        price_el = art.find("p", {"class": re.compile(r"price-shipping--price")})
        loc_el   = art.find("div", {"class": "aditem-main--top--left"})
        desc_el  = art.find("p", {"class": "aditem-main--middle--description"})
        tags     = [s.get_text(" ", strip=True)
                    for s in art.find_all("span", {"class": "simpletag"})]

        if not title_el:
            continue

        title = title_el.get_text(" ", strip=True)
        href  = title_el.get("href", "")
        url   = urljoin(base_url, href) if href else ""
        price_raw = price_el.get_text(" ", strip=True) if price_el else ""
        loc_raw   = " ".join((loc_el.get_text(" ", strip=True)).split()) if loc_el else ""
        desc_raw  = desc_el.get_text(" ", strip=True) if desc_el else ""
        tag_text  = " ".join(tags)

        # Skip monthly rent listings (price text contains /Monat or /m²)
        if re.search(r"/\s*(?:Monat|m[²2]|Jahr)", price_raw, re.IGNORECASE):
            continue

        # Parse price: "449.000 €" → remove dot (thousands sep) → 449000
        price = None
        _pm = re.search(r"([\d]{1,3}(?:[.\s][\d]{3})*(?:[,][\d]{1,2})?)\s*€", price_raw)
        if _pm:
            _pv = _pm.group(1).replace(".", "").replace(",", ".").replace(" ", "")
            try:
                pv = float(_pv)
                # Only treat as sale price if >= 50,000 EUR
                if pv >= 50_000:
                    price = pv
            except ValueError:
                pass

        # City: strip leading postal code (5 digits)
        city = re.sub(r"^\d{5}\s*", "", loc_raw).strip()

        full_text = f"{title} {price_raw} {loc_raw} {desc_raw} {tag_text}"
        kws = _find_keywords(full_text)

        listings.append({
            "title":        title,
            "address":      "",
            "city":         city,
            "bundesland":   _guess_bundesland(loc_raw),
            "asking_price": price,
            "gla_sqm":      _parse_gla(tag_text + " " + desc_raw),
            "year_built":   _parse_year_built(desc_raw),
            "anchor_tenant": _extract_anchor(kws),
            "giy_stated":   _parse_giy(desc_raw),
            "wault_stated": _parse_wault(desc_raw),
            "broker":       "",
            "url":          url,
            "source":       "Kleinanzeigen",
            "raw_text":     full_text[:500],
            "keywords":     kws,
        })

    if not listings:
        script_tags = soup.find_all("script")
        log.info(
            f"  Kleinanzeigen: 0 listings parsed. HTML={len(html):,} chars, "
            f"articles={len(articles)}, scripts={len(script_tags)}."
        )
    return listings


_PARSERS = {
    "immoscout24":   _parse_immoscout24,
    "immowelt":      _parse_immowelt,
    "kleinanzeigen": _parse_kleinanzeigen,
}


# ===========================================================================
# Scoring & filtering
# ===========================================================================
def score_listing(listing: dict) -> int:
    s = 0

    # Price in target range: up to 20 pts
    price = listing.get("asking_price")
    if price:
        if FILTER["min_price"] <= price <= FILTER["max_price"]:
            s += 20
        elif price <= FILTER["max_price"] * 1.15:
            s += 10  # slightly over but still interesting

    # GLA in range: up to 10 pts
    gla = listing.get("gla_sqm")
    if gla:
        if FILTER["min_gla"] <= gla <= FILTER["max_gla"]:
            s += 10
        elif gla <= FILTER["max_gla"] * 1.2:
            s += 5

    # Preferred Bundesland: 10 pts
    bl = listing.get("bundesland", "")
    if any(pref.lower() in bl.lower() for pref in FILTER["preferred_states"]):
        s += 10

    # Keywords found: 5 pts each, max 25
    kws = listing.get("keywords", [])
    s += min(len(kws) * 5, 25)

    # GIY stated and above threshold: 15 pts; bonus 5 if exceptionally high
    giy = listing.get("giy_stated")
    if giy:
        if giy >= FILTER["min_giy_if_stated"]:
            s += 15
        if giy >= 0.07:
            s += 5

    # anchor_quality_override: WAULT > 8yr AND GIY > 6% → score at least 75
    wault = listing.get("wault_stated")
    if (
        FILTER.get("anchor_quality_override")
        and wault and wault > FILTER["anchor_override_min_wault"]
        and giy  and giy  > FILTER["anchor_override_min_giy"]
    ):
        s = max(s, 75)

    return min(s, 100)


def _passes_filter(listing: dict) -> bool:
    """Return True if listing meets minimum investment criteria."""
    price = listing.get("asking_price")
    gla   = listing.get("gla_sqm")
    giy   = listing.get("giy_stated")
    wault = listing.get("wault_stated")
    kws   = listing.get("keywords", [])

    # Hard price filter (both bounds must be checkable)
    if price is not None:
        if price < FILTER["min_price"] or price > FILTER["max_price"] * 2:
            return False

    # GLA filter if known
    if gla is not None:
        if gla < FILTER["min_gla"] or gla > FILTER["max_gla"] * 2:
            return False

    # GIY hard floor if stated
    if giy is not None and giy < FILTER["min_giy_if_stated"] * 0.8:
        return False

    # anchor_quality_override bypasses keyword requirement
    if (
        FILTER.get("anchor_quality_override")
        and wault and wault > FILTER["anchor_override_min_wault"]
        and giy   and giy   > FILTER["anchor_override_min_giy"]
    ):
        return True

    # Keyword check: required only when both price AND GLA are unknown
    # (i.e. we have nothing else to go on). If price is in range and GLA
    # is plausible we allow the listing through — keywords boost score only.
    price_in_range = (
        price is not None
        and FILTER["min_price"] <= price <= FILTER["max_price"]
    )
    gla_ok = gla is None or FILTER["min_gla"] <= gla <= FILTER["max_gla"]
    if price_in_range and gla_ok:
        return True  # price in range is sufficient without keywords

    if not kws:
        return False

    return True


# ===========================================================================
# Excel pre-population
# ===========================================================================
def pre_populate_excel(listing: dict) -> Optional[pathlib.Path]:
    """Copy the UW template and fill available fields. Returns output path."""
    if not TEMPLATE.exists():
        log.warning("Template not found — skipping pre-population")
        return None
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not available — skipping pre-population")
        return None

    PRE_UW_DIR.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r'[^\w\-_. ]', '_', listing.get("title", "deal"))[:40].strip()
    date_str = datetime.date.today().strftime("%Y%m%d")
    out = PRE_UW_DIR / f"UW_{safe}_{date_str}.xlsx"

    shutil.copy2(str(TEMPLATE), str(out))

    wb = openpyxl.load_workbook(str(out))
    ws = wb["Deal Overview"]

    # Map to DEAL_CELLS (column B, Excel row numbers per CLAUDE.md)
    if listing.get("title"):
        ws["B5"] = listing["title"][:60]
    if listing.get("address"):
        ws["B6"] = listing["address"]
    if listing.get("city"):
        bl    = listing.get("bundesland", "")
        ws["B7"] = f"{listing['city']} / {bl}" if bl else listing["city"]
    if listing.get("gla_sqm"):
        ws["B9"] = listing["gla_sqm"]
    if listing.get("year_built"):
        ws["B10"] = listing["year_built"]
    if listing.get("asking_price"):
        ws["B16"] = listing["asking_price"]

    # GrEst rate from Bundesland
    bl     = listing.get("bundesland", "")
    grest  = GREST_BY_STATE.get(bl, GREST_DEFAULT)
    ws["B20"] = grest

    # Standard acquisition cost rates
    ws["B21"] = 0.015   # notary 1.5%
    ws["B22"] = 0.010   # broker 1.0%

    # Analysis date
    ws["B24"] = datetime.date.today().isoformat()

    # Pre-fill anchor tenant note in a notes cell if present
    anchor = listing.get("anchor_tenant")
    if anchor:
        try:
            ws["B8"] = f"Retail — anchor: {anchor}"
        except Exception:
            pass

    wb.save(str(out))
    log.info(f"Pre-populated: {out.name}")
    return out


# ===========================================================================
# Main scrape loop
# ===========================================================================
def _looks_like_listing(obj: dict) -> bool:
    """True if a dict has enough fields to be a property listing."""
    loc_keys  = {"city", "address", "addressLocality", "region", "street",
                 "addressRegion", "postalCode"}
    fin_keys  = {"price", "prices", "offers", "rent", "kaufpreis",
                 "miete", "amount"}
    prop_keys = {"area", "livingSpace", "plotArea", "rooms", "sqm", "gla",
                 "wohnflaeche", "nutzflaeche", "flaeche"}
    keys = set(obj.keys())
    hits = (bool(keys & loc_keys) + bool(keys & fin_keys) + bool(keys & prop_keys))
    return hits >= 2


def _extract_from_json_blob(data: object, source_url: str = "", _depth: int = 0) -> list:
    """
    Recursively hunt for listing arrays in an arbitrary JSON blob (max depth 6).
    Returns the first promising list of listing-like dicts found, or [].
    """
    if _depth > 6:
        return []
    if isinstance(data, list) and data:
        if isinstance(data[0], dict) and _looks_like_listing(data[0]):
            return data
    if isinstance(data, dict):
        # Prefer well-named keys for listing arrays
        _priority = ["estates", "items", "results", "resultList", "listings",
                     "properties", "objects", "entries", "hits", "data"]
        for key in _priority:
            val = data.get(key)
            if isinstance(val, list) and val and isinstance(val[0], dict):
                if _looks_like_listing(val[0]):
                    log.info(f"  API blob key='{key}' ({len(val)} items) <- {source_url[:60]}")
                    return val
        # Fall back to any array value that looks like listings
        for key, val in data.items():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                if _looks_like_listing(val[0]):
                    log.info(f"  API blob key='{key}' ({len(val)} items) <- {source_url[:60]}")
                    return val
        # Recurse into nested dicts/lists
        for val in data.values():
            if isinstance(val, (dict, list)):
                found = _extract_from_json_blob(val, source_url, _depth + 1)
                if found:
                    return found
    return []


def _norm_api_listing(raw: dict, source: str, base_url: str) -> dict:
    """Convert a raw API listing dict to the standard internal format."""
    text   = json.dumps(raw)
    addr   = raw.get("address") or {}
    if isinstance(addr, str):
        addr = {}
    city   = (addr.get("city") or addr.get("addressLocality") or
              raw.get("city") or "")
    price_raw = (raw.get("price") or
                 (raw.get("prices") or {}).get("buy", {}).get("amount") or
                 (raw.get("offers") or {}).get("price") or "")
    area_raw  = (raw.get("livingSpace") or raw.get("area") or
                 raw.get("plotArea") or
                 (raw.get("areas") or {}).get("living") or
                 (raw.get("areas") or {}).get("total") or "")
    expose_id = raw.get("id") or raw.get("exposeId") or raw.get("estateId") or ""
    listing_url = (raw.get("url") or
                   (urljoin(base_url, f"/expose/{expose_id}") if expose_id else ""))
    kws = _find_keywords(text)
    return {
        "title":        raw.get("title") or raw.get("name") or str(addr.get("street", ""))[:80],
        "address":      f"{addr.get('street','')} {addr.get('houseNumber','')}".strip(),
        "city":         city,
        "bundesland":   _guess_bundesland(city),
        "asking_price": _parse_price(str(price_raw)),
        "gla_sqm":      _parse_gla(str(area_raw)),
        "year_built":   raw.get("constructionYear") or raw.get("yearOfConstruction"),
        "anchor_tenant": _extract_anchor(kws),
        "giy_stated":   _parse_giy(text),
        "wault_stated": _parse_wault(text),
        "broker":       (raw.get("broker") or {}).get("companyName", "") if isinstance(raw.get("broker"), dict) else "",
        "url":          listing_url,
        "source":       source,
        "raw_text":     text[:500],
        "keywords":     kws,
    }


def _pw_fetch(browser, url: str, wait_selectors: list) -> tuple:
    """
    Fetch *url* with headless Chromium. Returns (html, status, api_listings).

    Three-layer extraction strategy
    --------------------------------
    1. XHR/fetch interception  — captures JSON API responses fired during
       page load that contain listing arrays (the SPA data source).
    2. window JS state         — page.evaluate() extracts window.__INITIAL_STATE__
       and similar globals set by the SPA framework.
    3. Rendered DOM            — page.content() after networkidle + selector wait,
       passed to BeautifulSoup parsers as before.

    api_listings is non-empty when layers 1 or 2 yield data; the caller skips
    the BS parser and uses api_listings directly.
    """
    from playwright.sync_api import TimeoutError as _PWTimeout

    ctx  = browser.new_context(
        user_agent=BOT_UA,
        locale="de-DE",
        extra_http_headers={"Accept-Language": "de-DE,de;q=0.9,en;q=0.8"},
        viewport={"width": 1280, "height": 900},
    )
    page    = ctx.new_page()
    # Apply stealth patches: hides navigator.webdriver, JS fingerprints, etc.
    try:
        from playwright_stealth import Stealth  # playwright-stealth 2.x API
        Stealth().apply_stealth_sync(page)
        log.info("  playwright-stealth applied")
    except (ImportError, Exception) as _ste:
        log.info(f"  stealth skipped: {_ste}")
    status  = 0
    _api_raw: list = []   # raw JSON blobs from intercepted responses

    # ── Layer 1: intercept JSON API responses (not JS/CSS/image assets) ─────
    _resp_log: list  = []   # (url_short, size, items_found) for diagnostics
    _json_urls: list = []   # every application/json response URL seen

    _ASSET_EXTS = ('.js', '.css', '.png', '.jpg', '.jpeg', '.gif', '.svg',
                   '.woff', '.woff2', '.ttf', '.ico', '.map', '.webp', '.mp4')

    def _on_response(resp):
        if resp.status != 200:
            return
        resp_url = resp.url
        if any(resp_url.lower().split('?')[0].endswith(ext) for ext in _ASSET_EXTS):
            return  # static asset — not API data
        ct = resp.headers.get("content-type", "").lower()
        if "javascript" in ct:
            return  # JS bundle — not API data
        # Accept application/json, text/plain, empty/unknown content-type
        if ct and "json" not in ct and "text/plain" not in ct:
            return
        # Log every JSON response URL for diagnostics
        if "json" in ct:
            _json_urls.append(resp_url[:80])
        try:
            raw_text = resp.text()
        except Exception as _re:
            _resp_log.append((resp_url[:70], f"body_err", 0))
            return
        size = len(raw_text)
        if size < 500 or size > 10_000_000:
            return
        try:
            body  = json.loads(raw_text)
            found = _extract_from_json_blob(body, resp_url)
            _resp_log.append((resp_url[:70], size, len(found)))
            if found:
                _api_raw.extend(found)
        except json.JSONDecodeError:
            _resp_log.append((resp_url[:70], size, "json_err"))

    page.on("response", _on_response)

    try:
        nav    = page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        status = nav.status if nav else 0
        log.info(f"  PW  {url[:80]} -> {status}")

        if status in (401, 403, 404, 410):
            return "", status, []

        # Phase 1a — dismiss cookie/GDPR consent dialog (common on German sites)
        _consent_selectors = [
            # Usercentrics (used by many German portals)
            "[data-testid='uc-accept-all-button']",
            "button[data-testid='uc-accept-all-button']",
            # OneTrust
            "#onetrust-accept-btn-handler",
            # Sourcepoint / Didomi
            "[data-qa='consent-accept']",
            "button[aria-label*='akzeptieren' i]",
            "button[aria-label*='accept' i]",
            # Generic German text patterns
            "button:text('Alle akzeptieren')",
            "button:text('Akzeptieren')",
            "button:text('Zustimmen')",
            "button:text('Einverstanden')",
            "button:text('OK')",
            # Immowelt-specific
            "[class*='consent'] button",
            "[class*='cookie'] button[class*='accept']",
            "[class*='CookieBanner'] button",
            "[id*='consent'] button",
        ]
        _dismissed = False
        for _cs in _consent_selectors:
            try:
                page.wait_for_selector(_cs, timeout=3_000)
                page.click(_cs)
                log.info(f"  Cookie consent dismissed: {_cs}")
                _dismissed = True
                # After consent, wait for the page to begin loading listing data
                page.wait_for_timeout(2_500)
                break
            except _PWTimeout:
                continue
        if not _dismissed:
            log.info("  No cookie consent dialog detected")

        # Phase 1b — wait for network to settle (XHR for listings fires here)
        try:
            page.wait_for_load_state("networkidle", timeout=20_000)
        except _PWTimeout:
            log.info("  networkidle timeout — adding 4s buffer for in-flight XHR")
            page.wait_for_timeout(4_000)

        # Phase 1c — scroll to trigger lazy-loaded listing cards
        try:
            page.evaluate("window.scrollTo(0, Math.round(document.body.scrollHeight * 0.4))")
            page.wait_for_timeout(1_200)
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1_200)
            page.evaluate("window.scrollTo(0, 0)")
        except Exception:
            pass

        # Phase 2 — wait for first matching listing-card selector
        matched_sel = None
        for sel in wait_selectors:
            try:
                page.wait_for_selector(sel, timeout=5_000)
                matched_sel = sel
                log.info(f"  Selector matched: {sel}")
                break
            except _PWTimeout:
                continue

        # ── Layer 2: window JS state via evaluate ────────────────────────────
        try:
            win_data = page.evaluate(
                """() => {
                    const keys = [
                        '__INITIAL_STATE__', '__APP_STATE__', '__NEXT_DATA__',
                        '__IMMOWELT_STATE__', '__IS24_STATE__',
                        'IS24_APP_DATA', 'immoweltAppState',
                    ];
                    for (const k of keys) {
                        if (window[k] && typeof window[k] === 'object') {
                            return {key: k, data: window[k]};
                        }
                    }
                    return null;
                }"""
            )
            if win_data and win_data.get("data"):
                found = _extract_from_json_blob(
                    win_data["data"], f"window.{win_data['key']}"
                )
                if found:
                    log.info(f"  window.{win_data['key']}: {len(found)} potential listings")
                    _api_raw.extend(found)
        except Exception as _eval_exc:
            log.info(f"  evaluate() skipped: {_eval_exc}")

        html = page.content()
        log.info(
            f"  DOM: {len(html):,} chars | "
            f"selector: {matched_sel or 'none'} | "
            f"API items: {len(_api_raw)}"
        )

        # Diagnostic: ALL application/json response URLs (no truncation)
        if _json_urls:
            log.info(f"  application/json responses ({len(_json_urls)}):")
            for _ju in _json_urls:
                log.info(f"    {_ju}")
        else:
            log.info("  No application/json responses intercepted")
        # Parsed responses that passed size/parse filter
        if _resp_log:
            log.info(f"  Parsed responses ({len(_resp_log)}):")
            for _ru, _rs, _ri in _resp_log:
                log.info(f"    size={_rs} items={_ri} | {_ru}")

        # ── Layer 3: probe serp-bff estate API directly with session cookies ──
        # Immowelt's search page calls serp-bff/estates (seen in network trace).
        # We replicate that call using the browser's session cookies.
        if not _api_raw and "immowelt" in url.lower():
            _bff_candidates = [
                "https://www.immowelt.de/serp-bff/estates?placesIds=AD02DE1&usageTypes=commercial&marketingTypes=KAUFEN&pageSize=50",
                "https://www.immowelt.de/serp-bff/search?placesIds=AD02DE1&usageTypes=commercial&marketingTypes=KAUFEN&pageSize=50",
                "https://www.immowelt.de/search-mfe-bff/estates?placesIds=AD02DE1&usageTypes=commercial&marketingTypes=KAUFEN&pageSize=50",
            ]
            for _bff_url in _bff_candidates:
                try:
                    _bff_result = page.evaluate(
                        f"""async () => {{
                            try {{
                                const r = await fetch("{_bff_url}", {{
                                    headers: {{
                                        "Accept": "application/json, */*",
                                        "Accept-Language": "de-DE,de;q=0.9",
                                        "X-Requested-With": "XMLHttpRequest",
                                        "Referer": "https://www.immowelt.de/gewerbeimmobilien",
                                    }},
                                    credentials: "include",
                                }});
                                const txt = await r.text();
                                let data = null;
                                try {{ data = JSON.parse(txt); }} catch(e) {{}}
                                return {{status: r.status, data: data, size: txt.length}};
                            }} catch(err) {{
                                return {{status: 0, error: err.message}};
                            }}
                        }}"""
                    )
                    _s = (_bff_result or {}).get("status", 0)
                    _sz = (_bff_result or {}).get("size", 0)
                    _bff_data = (_bff_result or {}).get("data")
                    log.info(f"  serp-bff {_s} size={_sz} | {_bff_url[:60]}")
                    if _bff_data and isinstance(_bff_data, (dict, list)):
                        _bff_found = _extract_from_json_blob(_bff_data, _bff_url)
                        if _bff_found:
                            log.info(f"  serp-bff HIT: {len(_bff_found)} listings")
                            _api_raw.extend(_bff_found)
                            break
                except Exception as _bff_e:
                    log.info(f"  serp-bff exception: {_bff_e}")

        # Also scan DOM for <script type="application/json"> data islands
        _soup_diag = BeautifulSoup(html, "html.parser")
        _json_tags = _soup_diag.find_all("script", {"type": "application/json"})
        if _json_tags:
            log.info(f"  DOM has {len(_json_tags)} <script type=application/json> tag(s)")
            for _jt in _json_tags:
                try:
                    _jd = json.loads(_jt.string or "")
                    _jf = _extract_from_json_blob(_jd, "dom_script_json")
                    if _jf:
                        log.info(f"  DOM JSON tag: {len(_jf)} listings found!")
                        _api_raw.extend(_jf)
                except Exception:
                    pass

        return html, status, _api_raw

    except Exception as exc:
        log.warning(f"  Playwright navigation error: {exc}")
        return "", status, []
    finally:
        ctx.close()


def scrape(quiet: bool = False) -> dict:
    """
    Run the full scan. Returns a results dict with:
        total_found, total_passed, listings, top5, errors, timestamp
    """
    if quiet:
        log.setLevel(logging.WARNING)

    # requests session still used only for robots.txt lookups (lightweight)
    sess   = _session()
    robots = _RobotsCache()
    today  = datetime.date.today().isoformat()

    all_raw: list  = []
    errors:  list  = []

    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
      browser = pw.chromium.launch(headless=True)
      try:
        for site in SITES:
            log.info(f"[{site['name']}] Starting scan (Playwright/Chromium)")
            parser_fn     = _PARSERS[site["parser"]]
            wait_sels     = site.get("wait_selectors", [])

            for search_url in site["search_urls"]:
                # robots.txt check (still via requests — fast and appropriate)
                if not robots.allowed(search_url, sess):
                    msg = f"{site['name']}: robots.txt disallows {search_url}"
                    log.warning(msg)
                    errors.append(msg)
                    continue

                time.sleep(REQUEST_DELAY)
                try:
                    html, status, api_raw = _pw_fetch(browser, search_url, wait_sels)
                except Exception as exc:
                    msg = f"{site['name']}: Playwright failed — {exc}"
                    log.warning(msg)
                    errors.append(msg)
                    continue

                if status in (401, 403, 404, 410):
                    msg = f"{site['name']}: HTTP {status} — {search_url[:60]}"
                    log.warning(msg)
                    errors.append(msg)
                    continue

                if not html and not api_raw:
                    msg = f"{site['name']}: empty response — {search_url[:60]}"
                    log.warning(msg)
                    errors.append(msg)
                    continue

                # Prefer API-intercepted listings; fall back to BS parser
                if api_raw:
                    log.info(f"  Using {len(api_raw)} intercepted API listings")
                    raw = [
                        _norm_api_listing(r, site["name"], site["base_url"])
                        for r in api_raw
                    ]
                else:
                    raw = parser_fn(html, site["base_url"])
                log.info(f"  Extracted {len(raw)} listing(s)")

                # Deduplicate by URL within this run
                seen_urls = {l["url"] for l in all_raw}
                for r in raw:
                    if r.get("url") and r["url"] not in seen_urls:
                        r["date_found"] = today
                        all_raw.append(r)
                        seen_urls.add(r["url"])

                time.sleep(REQUEST_DELAY)

      finally:
          browser.close()

    # Score and filter
    for r in all_raw:
        r["score"]        = score_listing(r)
        r["passes_filter"] = _passes_filter(r)

    passed = [r for r in all_raw if r["passes_filter"]]
    top5   = sorted(passed, key=lambda x: x["score"], reverse=True)[:5]

    # Pre-populate Excel for high-scorers
    for r in passed:
        if r["score"] >= FILTER["pre_populate_threshold"]:
            xls = pre_populate_excel(r)
            r["excel_path"] = str(xls) if xls else None
        else:
            r["excel_path"] = None

    return {
        "timestamp":    today,
        "total_found":  len(all_raw),
        "total_passed": len(passed),
        "listings":     all_raw,
        "top5":         top5,
        "errors":       errors,
    }


# ===========================================================================
# Persistence
# ===========================================================================
def _save(results: dict) -> pathlib.Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.date.today().strftime("%Y%m%d")
    path = OUTPUT_DIR / f"deals_found_{date_str}.json"
    path.write_text(
        json.dumps(results, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    # Also write a latest symlink / copy
    latest = OUTPUT_DIR / "deals_latest.json"
    latest.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
    return path


def read_latest() -> Optional[dict]:
    latest = OUTPUT_DIR / "deals_latest.json"
    if not latest.exists():
        return None
    try:
        return json.loads(latest.read_text(encoding="utf-8"))
    except Exception:
        return None


# ===========================================================================
# Report helper
# ===========================================================================
def _report(results: dict) -> str:
    lines = [
        "=" * 68,
        "  DEAL PIPELINE SCAN RESULTS",
        f"  {results['timestamp']}",
        "=" * 68,
        f"  Total listings found : {results['total_found']}",
        f"  Passing filter       : {results['total_passed']}",
    ]
    if results["errors"]:
        lines.append(f"\n  Errors / warnings    : {len(results['errors'])}")
        for e in results["errors"]:
            lines.append(f"    • {e}")

    if not results["top5"]:
        lines.append("\n  No qualifying listings found.")
        lines.append("  Likely causes: JS-rendered pages returned no listing data,")
        lines.append("  or robots.txt disallows scraping. See errors above.")
        lines.append("  Consider the sites' official data API programmes.")
    else:
        lines.append(f"\n  Top {len(results['top5'])} scored deals:")
        lines.append("  " + "-" * 64)
        for i, deal in enumerate(results["top5"], 1):
            price = f"EUR {deal['asking_price']:,.0f}" if deal.get("asking_price") else "n/a"
            gla   = f"{deal['gla_sqm']:,.0f} m²"     if deal.get("gla_sqm")      else "n/a"
            giy   = f"{deal['giy_stated']:.1%}"       if deal.get("giy_stated")   else "n/a"
            kws   = ", ".join(deal.get("keywords", [])[:3]) or "—"
            xls   = f"\n      Excel: {deal.get('excel_path','—')}" if deal.get("excel_path") else ""
            lines += [
                f"\n  #{i}  [{deal['score']}/100]  {deal.get('title','(no title)')[:55]}",
                f"       {deal.get('city','?')} | {deal.get('bundesland','?')}",
                f"       Price: {price}  |  GLA: {gla}  |  GIY: {giy}",
                f"       Keywords: {kws}",
                f"       URL: {deal.get('url','—')[:70]}",
            ]
            if xls:
                lines.append(xls)

    lines.append("\n" + "=" * 68)
    return "\n".join(lines)


# ===========================================================================
# Public entry point
# ===========================================================================
def run(force: bool = False, quiet: bool = False) -> pathlib.Path:
    """
    Orchestrate a full scan, save, and return the output path.
    Skips if today's file already exists and force=False.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_str  = datetime.date.today().strftime("%Y%m%d")
    out_path  = OUTPUT_DIR / f"deals_found_{date_str}.json"

    if out_path.exists() and not force:
        if not quiet:
            log.info(f"Today's scan already exists: {out_path} (use --force to re-run)")
        return out_path

    if not quiet:
        log.info(f"[deal_scanner] Starting property scan — {datetime.date.today()}")

    results = scrape(quiet=quiet)
    path    = _save(results)

    if not quiet:
        report = _report(results)
        sys.stdout.buffer.write((report + "\n").encode("utf-8", errors="replace"))
        sys.stdout.buffer.flush()
        log.info(f"[deal_scanner] Saved: {path}")

    return path


# ===========================================================================
# CLI
# ===========================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GRU Deal Pipeline Scanner")
    parser.add_argument("--force", action="store_true", help="Re-run even if today's file exists")
    parser.add_argument("--quiet", action="store_true", help="Suppress progress output")
    parser.add_argument("--show",  action="store_true", help="Print latest results and exit")
    args = parser.parse_args()

    if args.show:
        data = read_latest()
        if data:
            sys.stdout.buffer.write((_report(data) + "\n").encode("utf-8", errors="replace"))
        else:
            print("No scan results found. Run without --show first.")
    else:
        run(force=args.force, quiet=args.quiet)
