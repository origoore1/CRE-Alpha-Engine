"""
deal_simulator.py  v1.0
========================
Generates three realistic synthetic German Retail CRE scenarios and writes
them into the DE_Retail_Underwriter_Template.xlsx structure using openpyxl.

KEY DESIGN DECISIONS (vs. Gemini's broken approach)
----------------------------------------------------
• Uses openpyxl load_workbook to POPULATE the existing template —
  never overwrites it with pd.ExcelWriter (which destroys all formulas).
• Writes only to known blue-input cells; all formula cells are left intact.
• Each scenario outputs a SEPARATE named file — master template is untouched.
• Cost Sheet rates written to C14:C23 matching the parser's exact cell map.
• All rent roll dates written as datetime objects matching the template format.
• Void cost rates kept consistent with template defaults (€18/m²/yr).
• Documents expected QA verdict for each scenario so you can verify
  the engine is working correctly.

SCENARIOS
---------
  1. PASS        — Scharnitzpark, Regensburg (Core+ Fachmarktzentrum, Bavaria)
                   Strong Edeka anchor, low vacancy, 7.4yr WAULT, clean metrics.
  2. INVESTIGATE — Nordpark, Kassel (Value-Add Retail Park, Hesse)
                   LTV > 65%, vacancy 11%, 3yr expiry risk, borderline WAULT.
  3. DECLINE     — Stadtgalerie Süd, Duisburg (Distressed Shopping Centre, NRW)
                   71% LTV, DSCR 0.80x, WAULT 1.7yr — three simultaneous RED fails.

USAGE
-----
  python deal_simulator.py               → generates all three scenarios
  python deal_simulator.py 1             → generates only scenario 1
  python deal_simulator.py 2             → generates only scenario 2
  python deal_simulator.py 3             → generates only scenario 3
  python deal_simulator.py --verify      → runs parser on each file and prints verdict
"""

import shutil
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from constants import GREST_BY_STATE

# ---------------------------------------------------------------------------
# CELL MAP  (do not change — must match template structure exactly)
# ---------------------------------------------------------------------------

# Deal Overview: all in column B (col 2)
DEAL_CELLS = {
    "property_name":  (5,  2),
    "street_address": (6,  2),
    "city":           (7,  2),
    "asset_type":     (8,  2),
    "gla_sqm":        (9,  2),
    "year_built":     (10, 2),
    "parking":        (13, 2),
    "asking_price":   (16, 2),
    "equity":         (17, 2),
    "grest_rate":     (20, 2),
    "notar_rate":     (21, 2),
    "broker_rate":    (22, 2),
    "analysis_date":  (24, 2),
    "hold_period":    (25, 2),
    "interest_rate":  (28, 2),
    "loan_term_yrs":  (29, 2),
    "amort_type":     (30, 2),
    "amort_rate":     (31, 2),
    "io_period":      (33, 2),
}

# Rent Roll: data rows start at 4; columns A–R (1–18), skip formula cols G(7) and K(11)
RR_FIRST_DATA_ROW = 4
RR_LAST_DATA_ROW  = 103
RR_COLS = {
    "unit_ref":           1,   # A
    "tenant":             2,   # B
    "sector":             3,   # C
    "anchor":             4,   # D
    "area_sqm":           5,   # E
    "rent_sqm_mo":        6,   # F
    # col 7 (G) = Annual Rent formula — DO NOT overwrite
    "lease_start":        8,   # H
    "lease_expiry":       9,   # I
    "break_option":       10,  # J
    # col 11 (K) = Unexpired Term formula — DO NOT overwrite
    "cpi_clause":         12,  # L
    "cpi_passthrough":    13,  # M
    "cpi_trigger":        14,  # N
    "security_deposit":   15,  # O
    "svc_chg_recov":      16,  # P
    "rent_free_months":   17,  # Q
    "notes":              18,  # R
}

# Cost Sheet: rate inputs in column C (col 3), rows 14–23
COST_ROWS = {
    "property_mgmt_pct":  14,  # % of gross rent
    "asset_mgmt_pct":     15,  # % of gross rent
    "insurance_sqm":      16,  # EUR/m² GLA/yr
    "capex_sqm":          17,  # EUR/m² GLA/yr
    "ground_rent_abs":    18,  # EUR/yr absolute
    "grundsteuer_abs":    19,  # EUR/yr absolute (vacant portion)
    "void_svc_sqm":       20,  # EUR/m² vacant/yr
    "void_ins_sqm":       21,  # EUR/m² vacant/yr
    "letting_pct":        22,  # % of gross rent (amortised)
    "other_abs":          23,  # EUR/yr absolute
}

MASTER_TEMPLATE = "DE_Retail_Underwriter_Template.xlsx"


# ---------------------------------------------------------------------------
# SCENARIO DEFINITIONS
# ---------------------------------------------------------------------------

def _d(year, month, day):
    """Helper: return datetime for openpyxl date cells."""
    return datetime(year, month, day)


SCENARIOS = {

    # ════════════════════════════════════════════════════════════════════════
    # SCENARIO 1  —  PASS
    # Scharnitzpark Fachmarktzentrum, Regensburg, Bavaria
    #
    # Design logic:
    #   • Edeka anchor at realistic €11.00/m²/mo (not inflated €15)
    #   • Anchor concentration 54.3% — just inside the 55% threshold
    #   • Vacancy 8.8% — inside 10% threshold
    #   • WAULT ~7.4yr — well above 4yr minimum
    #   • Zero leases expiring within 3yr
    #   • LTV 60.8%, DSCR ~1.59x, ICR ~2.36x — all clean
    #
    # Expected verdict:  PASS
    # Expected QA:       All PASS or PASS with NOI-source WARN if Cost Sheet
    #                    values not loaded from LibreOffice recalc
    # ════════════════════════════════════════════════════════════════════════
    1: {
        "filename": "Scenario_1_PASS_Regensburg.xlsx",
        "expected_verdict": "PASS",
        "expected_issues": "None — clean institutional metrics throughout",
        "deal": {
            "property_name":  "Scharnitzpark Fachmarktzentrum",
            "street_address": "Scharnitzstrasse 4",
            "city":           "Regensburg / Bayern",
            "asset_type":     "Fachmarktzentrum (Retail Park)",
            "gla_sqm":        4550,
            "year_built":     2008,
            "parking":        210,
            "asking_price":   8_550_000,
            "equity":         3_350_000,
            "grest_rate":     GREST_BY_STATE["Bayern"],    # Bavaria: lowest in Germany
            "notar_rate":     0.015,
            "broker_rate":    0.010,
            "analysis_date":  date.today(),
            "hold_period":    10,
            "interest_rate":  0.041,
            "loan_term_yrs":  10,
            "amort_type":     "Annuitaetendarlehen",
            "amort_rate":     0.020,
            "io_period":      2,
        },
        "rent_roll": [
            # unit, tenant, sector, anchor, area_m2, rent_m2_mo,
            # lease_start, lease_expiry, break_option,
            # cpi, cpi_pt, cpi_trig, sec_dep, svc_rec, rf_mo, notes
            ("U-01", "Edeka Markt GmbH",   "Food & Grocery",  "Y", 2400, 11.00,
             _d(2018,1,1), _d(2034,6,30), None,
             "Y", 0.75, 0.10, 79_200, "Y", 0,
             "Anchor — 10yr lease, CPI 75% pass-through"),

            ("U-02", "Rossmann GmbH",      "Health & Beauty",  "N", 550, 15.50,
             _d(2021,3,1), _d(2032,12,31), None,
             "Y", 1.00, 0.10, 25_600, "Y", 0,
             "Strong national covenant"),

            ("U-03", "Deichmann SE",       "Footwear",         "N", 420, 13.00,
             _d(2022,6,1), _d(2031,6,30), None,
             "Y", 1.00, 0.10, 16_400, "Y", 0,
             ""),

            ("U-04", "Fressnapf GmbH",     "Pet Supplies",     "N", 480, 11.50,
             _d(2019,4,1), _d(2033,3,31), None,
             "Y", 1.00, 0.10, 16_600, "Y", 0,
             "Long CPI-indexed lease"),

            ("U-05", "KiK Textilien",      "Discount Fashion", "N", 320,  9.00,
             _d(2023,1,1), _d(2030,12,31), None,
             "Y", 0.75, 0.10, 8_600, "Y", 0,
             ""),

            ("U-06", "VACANT",             "—",                "N", 380,  0.00,
             None, None, None,
             "N", 0.00, 0.00, 0, "N", 0,
             "Void since Q1 2025 — under active letting"),
        ],
        "costs": {
            "property_mgmt_pct": 0.015,
            "asset_mgmt_pct":    0.005,
            "insurance_sqm":     2.50,
            "capex_sqm":         5.50,
            "ground_rent_abs":   0,
            "grundsteuer_abs":   0,
            "void_svc_sqm":      18.0,
            "void_ins_sqm":      1.5,
            "letting_pct":       0.040,
            "other_abs":         1_500,
        },
    },

    # ════════════════════════════════════════════════════════════════════════
    # SCENARIO 2  —  INVESTIGATE
    # Nordpark Fachmarkt, Kassel, Hesse
    #
    # Design logic:
    #   • LTV 65.8% — just above 65% threshold → WARN
    #   • Vacancy 11.1% — just above 10% threshold → WARN
    #   • 3yr expiry risk 17.1% (Woolworth + Takko) → WARN
    #   • WAULT 4.79yr — above 4yr, avoids RED; deliberate (RED = DECLINE not INVESTIGATE)
    #   • DSCR 1.55x — passes, but tighter than Scenario 1
    #   • GrESt 6.0% (Hesse) — increases acquisition cost vs. Bavaria
    #   • Three yellow warns → algorithm returns INVESTIGATE
    #
    # Expected verdict:  INVESTIGATE
    # Expected QA:       LTV WARN, Vacancy WARN, 3yr Expiry WARN
    #                    DSCR PASS, WAULT PASS, ICR PASS
    # ════════════════════════════════════════════════════════════════════════
    2: {
        "filename": "Scenario_2_INVESTIGATE_Kassel.xlsx",
        "expected_verdict": "INVESTIGATE",
        "expected_issues": (
            "LTV 65.1% (just above 65%), Vacancy 11.1%, "
            "3yr expiry risk 33.8% (Woolworth + Takko + Netto + small unit)"
        ),
        "deal": {
            "property_name":  "Nordpark Fachmarkt",
            "street_address": "Wolfhager Strasse 120",
            "city":           "Kassel / Hessen",
            "asset_type":     "Fachmarktzentrum (Retail Park)",
            "gla_sqm":        5400,
            "year_built":     2005,
            "parking":        240,
            "asking_price":   7_300_000,
            "equity":         2_550_000,
            "grest_rate":     GREST_BY_STATE["Hessen"],    # Hesse: 6.0%
            "notar_rate":     0.015,
            "broker_rate":    0.010,
            "analysis_date":  date.today(),
            "hold_period":    10,
            "interest_rate":  0.041,
            "loan_term_yrs":  10,
            "amort_type":     "Annuitaetendarlehen",
            "amort_rate":     0.020,
            "io_period":      2,
        },
        "rent_roll": [
            ("U-01", "Rewe Markt GmbH",    "Food & Grocery",     "Y", 2600,  9.80,
             _d(2021,10,1), _d(2031,9,30), None,
             "Y", 0.75, 0.10, 22_050, "Y", 0,
             "Anchor — CPI-indexed, sound covenant"),

            ("U-02", "Woolworth GmbH",     "General Merchandise","N",  600,  8.50,
             _d(2020,10,1), _d(2028,9,30), None,
             "N", 0.00, 0.00, 4_250, "Y", 0,
             "Expiry risk: 2.5yr remaining — no CPI clause"),

            ("U-03", "Takko Fashion",      "Value Fashion",      "N",  380,  8.00,
             _d(2020,4,1),  _d(2028,3,31), None,
             "N", 0.00, 0.00, 2_560, "Y", 3,
             "Expiry risk: 2.0yr remaining — rent-free at renewal"),

            ("U-04", "dm-drogerie markt",  "Health & Beauty",    "N",  380, 16.00,
             _d(2022,10,1), _d(2032,9,30), None,
             "Y", 1.00, 0.10, 7_296, "Y", 0,
             "Strong covenant, long term"),

            ("U-05", "Netto Marken-Discount","Discount Food",    "N",  500,  8.50,
             _d(2022,4,1),  _d(2027,9,30), None,   # 1.5yr — expiry risk
             "Y", 0.75, 0.10, 3_825, "Y", 0,
             "Expiry risk: 1.5yr — short remaining term"),

            ("U-06", "Small Unit (Dienstleistung)","Services",   "N",  340, 11.00,
             _d(2023,4,1),  _d(2028,9,30), None,   # 2.5yr — expiry risk
             "N", 0.00, 0.00, 3_080, "Y", 0,
             "Hairdresser / services — 2.5yr remaining, no CPI protection"),

            ("U-07", "VACANT",             "—",                  "N",  600,  0.00,
             None, None, None,
             "N", 0.00, 0.00, 0, "N", 0,
             "Void since Q3 2024 — former fashion unit"),
        ],
        "costs": {
            "property_mgmt_pct": 0.015,
            "asset_mgmt_pct":    0.005,
            "insurance_sqm":     2.50,
            "capex_sqm":         5.50,
            "ground_rent_abs":   0,
            "grundsteuer_abs":   1_200,  # Grundsteuer on vacant unit
            "void_svc_sqm":      18.0,
            "void_ins_sqm":      1.5,
            "letting_pct":       0.050,  # Slightly higher — harder market
            "other_abs":         2_500,
        },
    },

    # ════════════════════════════════════════════════════════════════════════
    # SCENARIO 3  —  DECLINE
    # Stadtgalerie Süd, Duisburg, NRW
    #
    # Design logic:
    #   • THREE simultaneous RED fails — DECLINE is triggered at ≥2 red fails
    #       1. WAULT 1.70yr     → RED (< 4yr minimum)
    #       2. DSCR  0.80x      → RED (< 1.20x hard floor)
    #       3. LTV   71.4%      → RED (> 70% hard ceiling)
    #   • 53.8% vacancy — extreme (former anchor space + two dark units)
    #   • All active leases expire within 2.5yr — 100% expiry risk
    #   • NOI margin ~49% — void costs dominate
    #   • GrESt 6.5% (NRW — highest in Germany)
    #   • This scenario is deliberately severe — tests that all three RED
    #     channels fire simultaneously and the engine does not give a false PASS
    #
    # Expected verdict:  DECLINE
    # Expected QA:       WAULT RED FAIL, DSCR RED FAIL, LTV RED FAIL,
    #                    Vacancy FAIL (yellow), 3yr expiry risk FAIL (yellow)
    # ════════════════════════════════════════════════════════════════════════
    3: {
        "filename": "Scenario_3_DECLINE_Duisburg.xlsx",
        "expected_verdict": "DECLINE",
        "expected_issues": (
            "WAULT 1.70yr (< 4yr), DSCR 0.80x (< 1.20x), LTV 71.4% (> 70%), "
            "Vacancy 53.8%, all leases expire < 3yr"
        ),
        "deal": {
            "property_name":  "Stadtgalerie Sued",
            "street_address": "Koenigsborner Strasse 42",
            "city":           "Duisburg / Nordrhein-Westfalen",
            "asset_type":     "Shopping Centre (Distressed)",
            "gla_sqm":        6500,
            "year_built":     1998,
            "parking":        180,
            "asking_price":   4_200_000,
            "equity":         1_200_000,
            "grest_rate":     GREST_BY_STATE["Nordrhein-Westfalen"],    # NRW: highest GrESt in Germany
            "notar_rate":     0.015,
            "broker_rate":    0.010,
            "analysis_date":  date.today(),
            "hold_period":    10,
            "interest_rate":  0.041,
            "loan_term_yrs":  10,
            "amort_type":     "Annuitaetendarlehen",
            "amort_rate":     0.020,
            "io_period":      0,        # No IO — lender demands immediate amortisation
        },
        "rent_roll": [
            ("U-01", "Penny Markt GmbH",   "Discount Food",      "Y", 1800,  7.00,
             _d(2022,4,1),  _d(2028,3,31), _d(2026,9,30),
             "N", 0.00, 0.00, 10_800, "Y", 0,
             "Acting anchor — break option in 6 months. HIGH RISK."),

            ("U-02", "Woolworth GmbH",     "General Merchandise","N",  600,  7.50,
             _d(2020,10,1), _d(2027,9,30), None,
             "N", 0.00, 0.00, 3_375, "Y", 0,
             "1.5yr remaining — no renewal interest signalled"),

            ("U-03", "Nails & Beauty",     "Personal Services",  "N",  250, 12.00,
             _d(2023,4,1),  _d(2027,3,31), None,
             "N", 0.00, 0.00, 3_000, "N", 0,
             "Independent operator — covenant risk"),

            ("U-04", "Phone Repair Shop",  "Electronics Services","N", 200, 11.00,
             _d(2024,10,1), _d(2026,9,30), None,
             "N", 0.00, 0.00, 1_320, "N", 0,
             "URGENT: expires in 6 months — no renewal signed"),

            ("U-05", "Lottery / Lotto",    "Personal Services",  "N",  150, 15.00,
             _d(2023,10,1), _d(2028,9,30), None,
             "N", 0.00, 0.00, 1_800, "N", 0,
             "Small stable unit — insufficient to offset vacancy"),

            ("U-06", "VACANT",              "—",                  "N", 1500,  0.00,
             None, None, None,
             "N", 0.00, 0.00, 0, "N", 0,
             "Former Saturn anchor — departed Q2 2024, 1,500m2 dark"),

            ("U-07", "VACANT",              "—",                  "N", 1200,  0.00,
             None, None, None,
             "N", 0.00, 0.00, 0, "N", 0,
             "Former Karstadt Sports — dark since 2023"),

            ("U-08", "VACANT",              "—",                  "N",  800,  0.00,
             None, None, None,
             "N", 0.00, 0.00, 0, "N", 0,
             "Multiple corridor units — never re-let after 2022"),
        ],
        "costs": {
            "property_mgmt_pct": 0.015,
            "asset_mgmt_pct":    0.005,
            "insurance_sqm":     3.00,   # Higher — older, larger building
            "capex_sqm":         7.00,   # Higher — 1998 build, ageing systems
            "ground_rent_abs":   0,
            "grundsteuer_abs":   4_500,  # Grundsteuer on 3,500m2 vacant
            "void_svc_sqm":      20.0,   # Higher rate — complex building
            "void_ins_sqm":      2.0,
            "letting_pct":       0.080,  # Very high — specialist retail leasing agent needed
            "other_abs":         5_000,
        },
    },
}


# ---------------------------------------------------------------------------
# CORE ENGINE  —  write one scenario into a copy of the master template
# ---------------------------------------------------------------------------

def _clear_rent_roll(ws):
    """Erase all tenant data from rows 4–103. Leave formulas in G and K intact."""
    for row in range(RR_FIRST_DATA_ROW, RR_LAST_DATA_ROW + 1):
        for col_name, col_idx in RR_COLS.items():
            ws.cell(row=row, column=col_idx).value = None


def write_scenario(scenario_id: int, master_path: str = MASTER_TEMPLATE) -> str:
    """
    Populate a copy of the master template with a scenario's data.
    Returns the output filepath.
    """
    scen = SCENARIOS[scenario_id]
    out_path = scen["filename"]

    # Never touch the master — always work on a fresh copy
    shutil.copy2(master_path, out_path)
    wb = load_workbook(out_path)

    # ── Deal Overview ────────────────────────────────────────────────────────
    ws_deal = wb["Deal Overview"]
    for field, (row, col) in DEAL_CELLS.items():
        val = scen["deal"].get(field)
        if val is not None:
            ws_deal.cell(row=row, column=col).value = val

    # ── Rent Roll ────────────────────────────────────────────────────────────
    ws_rr = wb["Rent Roll"]
    _clear_rent_roll(ws_rr)

    for i, tenant in enumerate(scen["rent_roll"]):
        row = RR_FIRST_DATA_ROW + i
        (unit_ref, tenant_name, sector, anchor, area_sqm, rent_sqm_mo,
         lease_start, lease_expiry, break_option,
         cpi_clause, cpi_pt, cpi_trig, sec_dep, svc_rec, rf_mo, notes) = tenant

        ws_rr.cell(row=row, column=RR_COLS["unit_ref"])         .value = unit_ref
        ws_rr.cell(row=row, column=RR_COLS["tenant"])           .value = tenant_name
        ws_rr.cell(row=row, column=RR_COLS["sector"])           .value = sector
        ws_rr.cell(row=row, column=RR_COLS["anchor"])           .value = anchor
        ws_rr.cell(row=row, column=RR_COLS["area_sqm"])         .value = area_sqm
        ws_rr.cell(row=row, column=RR_COLS["rent_sqm_mo"])      .value = rent_sqm_mo
        ws_rr.cell(row=row, column=RR_COLS["lease_start"])      .value = lease_start
        ws_rr.cell(row=row, column=RR_COLS["lease_expiry"])     .value = lease_expiry
        ws_rr.cell(row=row, column=RR_COLS["break_option"])     .value = break_option
        ws_rr.cell(row=row, column=RR_COLS["cpi_clause"])       .value = cpi_clause
        ws_rr.cell(row=row, column=RR_COLS["cpi_passthrough"])  .value = cpi_pt
        ws_rr.cell(row=row, column=RR_COLS["cpi_trigger"])      .value = cpi_trig
        ws_rr.cell(row=row, column=RR_COLS["security_deposit"]) .value = sec_dep
        ws_rr.cell(row=row, column=RR_COLS["svc_chg_recov"])    .value = svc_rec
        ws_rr.cell(row=row, column=RR_COLS["rent_free_months"]) .value = rf_mo
        ws_rr.cell(row=row, column=RR_COLS["notes"])            .value = notes

    # ── Cost Sheet ────────────────────────────────────────────────────────────
    ws_cost = wb["Cost Sheet"]
    for cost_key, row in COST_ROWS.items():
        val = scen["costs"].get(cost_key, 0)
        ws_cost.cell(row=row, column=3).value = val   # column C

    wb.save(out_path)
    # NOTE: No LibreOffice recalc needed. The 3-layer Python fallback in
    # rent_roll_parser.load_cost_sheet() and load_rent_roll() recomputes all
    # formula cells from raw inputs when data_only=True returns None.
    # This makes the system fully OS-agnostic and cloud-ready.
    return out_path


# ---------------------------------------------------------------------------
# VERIFICATION  —  run the parser on each generated file
# ---------------------------------------------------------------------------

def verify_scenarios(scenario_ids):
    """Run rent_roll_parser on each file and compare verdict to expected."""
    try:
        import rent_roll_parser as rrp
    except ImportError:
        print("  [!] rent_roll_parser.py not found in this folder. Skipping verify.")
        return

    print("\n" + "=" * 72)
    print("  VERIFICATION  —  Parser Output vs. Expected Verdicts")
    print("=" * 72)

    for sid in scenario_ids:
        scen = SCENARIOS[sid]
        path = scen["filename"]
        expected = scen["expected_verdict"]

        if not Path(path).exists():
            print(f"\n  Scenario {sid}: file not found — run without --verify first")
            continue

        print(f"\n  Scenario {sid}: {path}")
        print(f"  Expected verdict : {expected}")
        print(f"  Expected issues  : {scen['expected_issues']}")
        print()

        try:
            deal    = rrp.load_deal_overview(path)
            rr      = rrp.load_rent_roll(path)
            _rr_active  = rr[~rr["is_vacant"]]
            _rr_vacant  = rr[rr["is_vacant"]]
            costs   = rrp.load_cost_sheet(
                path,
                rr_gross_rent  = float(_rr_active["annual_rent"].sum()),
                rr_gla_total   = float(rr["area_sqm"].sum()),
                rr_vacant_area = float(_rr_vacant["area_sqm"].sum()),
            )
            metrics = rrp.compute_metrics(deal, rr, costs)
            qa      = rrp.run_qa_audit(metrics, deal)
            verdict = rrp.generate_verdict(metrics, qa)

            actual = verdict["recommendation"]
            match  = "MATCH" if actual == expected else "MISMATCH"

            print(f"  Gross Rent   : EUR {metrics['gross_passing_rent']:>12,.0f}")
            print(f"  NOI          : EUR {metrics['noi_proxy']:>12,.0f}  "
                  f"({metrics['noi_margin']:.1%} margin, source: {metrics['noi_source']})")
            print(f"  WAULT        : {metrics['wault']:.2f} yrs")
            print(f"  Vacancy      : {metrics['vacancy_rate']:.1%}")
            print(f"  LTV          : {metrics['ltv']:.1%}")
            dscr_s = f"{metrics['dscr']:.2f}x" if metrics.get("dscr") else "--"
            icr_s  = f"{metrics['icr']:.2f}x"  if metrics.get("icr")  else "--"
            print(f"  DSCR / ICR   : {dscr_s} / {icr_s}")
            print(f"  Anchor Conc. : {metrics['anchor_pct']:.1%}")
            print()
            for r in qa:
                icon = "OK" if r.status == "PASS" else ("!!" if r.status == "WARN" else "XX")
                print(f"    [{icon}]  {r.check:<40}  {r.actual}")
            print()
            print(f"  ACTUAL VERDICT  : {actual}")
            print(f"  ── {match} {'✅' if match == 'MATCH' else '❌'} ──")

        except Exception as e:
            import traceback
            print(f"  ERROR: {e}")
            traceback.print_exc()

    print("\n" + "=" * 72 + "\n")


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]

    if not Path(MASTER_TEMPLATE).exists():
        print(f"ERROR: '{MASTER_TEMPLATE}' not found in this folder.")
        print("Place this script in the same folder as the master template.")
        sys.exit(1)

    verify_only = "--verify" in args
    args = [a for a in args if a != "--verify"]

    # Determine which scenarios to run
    if args:
        try:
            ids = [int(a) for a in args if a.isdigit()]
            if not ids:
                raise ValueError
        except ValueError:
            print(f"Usage: python deal_simulator.py [1|2|3] [--verify]")
            sys.exit(1)
    else:
        ids = [1, 2, 3]

    if not verify_only:
        print("\n" + "=" * 72)
        print("  DEAL SIMULATOR  v1.0  —  German Retail CRE Stress Testing")
        print("=" * 72)

        for sid in ids:
            scen = SCENARIOS[sid]
            print(f"\n  Generating Scenario {sid}: {scen['deal']['property_name']} ...")
            out = write_scenario(sid)
            print(f"  Output file     : {out}")
            print(f"  Expected verdict: {scen['expected_verdict']}")
            print(f"  Known issues    : {scen['expected_issues']}")

        print("\n" + "─" * 72)
        print("  NEXT STEPS")
        print("─" * 72)
        print("  1. Run parser verification:")
        print("       python deal_simulator.py --verify")
        print("  2. Launch Streamlit and upload each scenario file:")
        print("       python -m streamlit run app.py")
        print("  3. Compare Banker Lens / Investor Lens verdicts across the 3 files.")
        print("=" * 72 + "\n")

    verify_scenarios(ids)


if __name__ == "__main__":
    main()
