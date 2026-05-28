"""
German Retail Underwriter — Deal Intake System
===============================================
When you receive a deal (broker email, PDF, exposé, notes in any format),
share the details with Claude and Claude fills this file, then runs it
to generate a ready-to-upload Excel file.

HOW TO USE:
  1. Share deal details with Claude in any format
  2. Claude fills in the DEAL dict below
  3. Claude runs: python intake_deal.py
  4. Download the generated .xlsx file and upload to the app

FIELD GUIDE:
  - All money in EUR (integers)
  - Rates as decimals: 0.041 = 4.1%
  - Dates as datetime objects or None
  - Tenant rent: EUR/m²/month (NOT annual)
  - VACANT rows: rent_sqm_mo=0, lease dates=None
  - Anchor tenant: anchor="Y", others "N"
  - CPI clause: "Y"/"N"
  - Security deposit: EUR absolute (standard = 3 months rent)
"""

from datetime import datetime
from pathlib import Path
import shutil
from openpyxl import load_workbook
from constants import GREST_DEFAULT
from deal_simulator import (
    DEAL_CELLS, RR_COLS, COST_ROWS,
    RR_FIRST_DATA_ROW, _clear_rent_roll, MASTER_TEMPLATE
)

# =============================================================================
# FILL THIS SECTION WITH DEAL DATA
# Claude populates everything below from the deal materials you share
# =============================================================================

DEAL = {

    # ── Property Identity ────────────────────────────────────────────────────
    "property_name":  "PROPERTY NAME HERE",
    "street_address": "STREET ADDRESS",
    "city":           "CITY / BUNDESLAND",        # e.g. "Munich / Bayern"
    "asset_type":     "Fachmarktzentrum (Retail Park)",  # or Nahversorgungszentrum etc
    "gla_sqm":        0,                           # total GLA in m²
    "year_built":     None,                        # or integer year
    "parking":        0,                           # number of parking spaces

    # ── Transaction ──────────────────────────────────────────────────────────
    "asking_price":   0,                           # EUR
    "equity":         0,                           # EUR equity contribution
    "grest_rate":     GREST_DEFAULT,               # Auto-detected from city; override per GREST_BY_STATE
    "notar_rate":     0.015,                       # standard 1.5%
    "broker_rate":    0.010,                       # standard 1.0% (if applicable)

    # ── Financing ────────────────────────────────────────────────────────────
    "interest_rate":  0.041,                       # 4.1% typical 5yr fixed 2026
    "loan_term_yrs":  10,                          # years
    "amort_type":     "Annuitaetendarlehen",        # or "Lineare Tilgung"
    "amort_rate":     0.020,                       # 2.0% standard
    "io_period":      2,                           # interest-only years

    # ── Analysis Parameters ──────────────────────────────────────────────────
    "analysis_date":  datetime.today().date(),
    "hold_period":    10,
}

# =============================================================================
# RENT ROLL
# One tuple per tenant/unit. VACANT units must have rent_sqm_mo=0.
# Format: (unit_ref, tenant_name, sector, anchor, area_sqm, rent_sqm_mo,
#           lease_start, lease_expiry, break_option,
#           cpi_clause, cpi_passthrough, cpi_trigger,
#           security_deposit, svc_chg_recoverable, rent_free_months, notes)
# =============================================================================

# Sector options: Food & Grocery / Health & Beauty / Discount Fashion /
# Discount Food / Pet Supplies / DIY & Home / Electronics / Services /
# General Merchandise / Personal Services / Other

RENT_ROLL = [

    # Example structure — Claude replaces with actual deal data:
    ("U-01", "TENANT NAME",   "Food & Grocery",  "Y",  0, 0.00,
     None, None, None,
     "Y", 0.75, 0.10, 0, "Y", 0,
     "Notes about this tenant"),

    ("U-02", "VACANT",        "—",               "N",  0, 0.00,
     None, None, None,
     "N", 0.00, 0.00, 0, "N", 0,
     "Vacant unit description"),

]

# =============================================================================
# COST SHEET
# All rates as decimals or EUR absolute values
# Benchmarks: property_mgmt 1.5-2%, asset_mgmt 0.5%, insurance EUR 2-3/m²,
# capex EUR 4-8/m² (IREBS), void_svc EUR 18-22/m², void_ins EUR 1.5-2/m²
# =============================================================================

COSTS = {
    "property_mgmt_pct": 0.015,   # % of gross rent
    "asset_mgmt_pct":    0.005,   # % of gross rent
    "insurance_sqm":     2.50,    # EUR/m² GLA/yr
    "capex_sqm":         5.00,    # EUR/m² GLA/yr (CapEx reserve)
    "ground_rent_abs":   0,       # EUR/yr (Erbpacht — only if leasehold)
    "grundsteuer_abs":   0,       # EUR/yr (property tax — fill if known)
    "void_svc_sqm":      20.0,    # EUR/m² vacant/yr (service charge on voids)
    "void_ins_sqm":      1.75,    # EUR/m² vacant/yr (insurance on voids)
    "letting_pct":       0.050,   # % of gross rent (leasing commissions)
    "other_abs":         2_000,   # EUR/yr (other non-recoverables)
}

# =============================================================================
# INTAKE NOTES
# Claude fills this with a summary of what was confirmed vs assumed
# =============================================================================

INTAKE_NOTES = """
CONFIRMED FROM DEAL MATERIALS:
- (Claude lists what was in the source documents)

ASSUMED / ESTIMATED:
- (Claude lists what was not in the source and was assumed using market benchmarks)

MISSING — NEED TO VERIFY:
- (Claude lists what is unknown and could materially change the analysis)

SOURCE DOCUMENTS:
- (Claude lists what was provided: email, PDF, etc.)
"""

# =============================================================================
# ENGINE — do not modify below this line
# =============================================================================

from constants import GREST_BY_STATE, GREST_DEFAULT


def _auto_grest(city_string: str) -> float:
    """Auto-detect GrESt rate from city/state string."""
    for state, rate in GREST_BY_STATE.items():
        if state.lower() in city_string.lower():
            return rate
    return GREST_DEFAULT


def build_excel(deal: dict, rent_roll: list, costs: dict,
                output_filename: str = None) -> str:
    """
    Generate a populated Excel file from intake data.
    Returns the output path.
    """
    if output_filename is None:
        clean = deal.get('property_name', 'Deal').replace(' ', '_')[:40]
        date_str = datetime.today().strftime('%Y%m%d')
        output_filename = f"INTAKE_{date_str}_{clean}.xlsx"

    out_path = Path(__file__).parent / output_filename
    shutil.copy2(MASTER_TEMPLATE, out_path)
    wb = load_workbook(out_path)

    # ── Deal Overview ─────────────────────────────────────────────────────────
    ws_deal = wb['Deal Overview']

    # Auto-detect GrESt if not explicitly set
    if deal.get('grest_rate', 0) == 0.055:
        detected = _auto_grest(deal.get('city', ''))
        if detected != 0.055:
            deal['grest_rate'] = detected
            print(f"  Auto-detected GrESt: {detected:.1%} from city: {deal['city']}")

    for field, (row, col) in DEAL_CELLS.items():
        val = deal.get(field)
        if val is not None:
            ws_deal.cell(row=row, column=col).value = val

    # ── Rent Roll ─────────────────────────────────────────────────────────────
    ws_rr = wb['Rent Roll']
    _clear_rent_roll(ws_rr)

    total_area = 0
    vacant_area = 0
    tenant_count = 0
    vacant_count = 0

    for i, tenant in enumerate(rent_roll):
        row = RR_FIRST_DATA_ROW + i
        (unit_ref, tenant_name, sector, anchor, area_sqm, rent_sqm_mo,
         lease_start, lease_expiry, break_option, cpi_clause, cpi_pt, cpi_trig,
         sec_dep, svc_rec, rf_mo, notes) = tenant

        ws_rr.cell(row=row, column=RR_COLS['unit_ref']).value          = unit_ref
        ws_rr.cell(row=row, column=RR_COLS['tenant']).value            = tenant_name
        ws_rr.cell(row=row, column=RR_COLS['sector']).value            = sector
        ws_rr.cell(row=row, column=RR_COLS['anchor']).value            = anchor
        ws_rr.cell(row=row, column=RR_COLS['area_sqm']).value          = area_sqm
        ws_rr.cell(row=row, column=RR_COLS['rent_sqm_mo']).value       = rent_sqm_mo
        ws_rr.cell(row=row, column=RR_COLS['lease_start']).value       = lease_start
        ws_rr.cell(row=row, column=RR_COLS['lease_expiry']).value      = lease_expiry
        ws_rr.cell(row=row, column=RR_COLS['break_option']).value      = break_option
        ws_rr.cell(row=row, column=RR_COLS['cpi_clause']).value        = cpi_clause
        ws_rr.cell(row=row, column=RR_COLS['cpi_passthrough']).value   = cpi_pt
        ws_rr.cell(row=row, column=RR_COLS['cpi_trigger']).value       = cpi_trig
        ws_rr.cell(row=row, column=RR_COLS['security_deposit']).value  = sec_dep
        ws_rr.cell(row=row, column=RR_COLS['svc_chg_recov']).value     = svc_rec
        ws_rr.cell(row=row, column=RR_COLS['rent_free_months']).value  = rf_mo
        ws_rr.cell(row=row, column=RR_COLS['notes']).value             = notes

        total_area += area_sqm
        if tenant_name.upper() == 'VACANT' or rent_sqm_mo == 0:
            vacant_area += area_sqm
            vacant_count += 1
        else:
            tenant_count += 1

    # ── Cost Sheet ────────────────────────────────────────────────────────────
    ws_cost = wb['Cost Sheet']
    for cost_key, row_num in COST_ROWS.items():
        ws_cost.cell(row=row_num, column=3).value = costs.get(cost_key, 0)

    wb.save(out_path)
    return str(out_path), total_area, vacant_area, tenant_count, vacant_count


def _quick_check(deal, rent_roll, costs):
    """
    Pre-flight sanity check before generating Excel.
    Returns list of issues.
    """
    issues = []

    if not deal.get('property_name') or deal['property_name'] == 'PROPERTY NAME HERE':
        issues.append("MISSING: property_name not filled in")
    if not deal.get('asking_price') or deal['asking_price'] == 0:
        issues.append("MISSING: asking_price is 0")
    if not deal.get('equity') or deal['equity'] == 0:
        issues.append("MISSING: equity is 0")
    if not deal.get('gla_sqm') or deal['gla_sqm'] == 0:
        issues.append("MISSING: gla_sqm is 0")

    if deal.get('equity', 0) >= deal.get('asking_price', 1):
        issues.append("WARNING: equity >= asking_price (all-equity deal?)")

    ltv = (deal.get('asking_price', 0) - deal.get('equity', 0)) / deal.get('asking_price', 1)
    if ltv > 0.80:
        issues.append(f"WARNING: LTV {ltv:.1%} very high — verify equity/price")

    tenant_rows = [t for t in rent_roll if t[1].upper() != 'VACANT' and t[5] > 0]
    if len(tenant_rows) == 0:
        issues.append("MISSING: no active tenants in rent roll")

    for i, t in enumerate(rent_roll):
        if t[5] > 50:
            issues.append(f"WARNING: row {i+1} ({t[1]}) rent EUR {t[5]}/m²/mo — "
                          f"very high, verify units (should be monthly per m²)")
        if t[4] > 10000:
            issues.append(f"WARNING: row {i+1} ({t[1]}) area {t[4]}m² — verify units")

    return issues


def print_intake_summary(deal, rent_roll, costs, path,
                          total_area, vacant_area, tenant_count, vacant_count):
    """Print a clean summary of what was built."""
    loan = deal.get('asking_price', 0) - deal.get('equity', 0)
    gross_rent = sum(t[4] * t[5] * 12 for t in rent_roll
                     if t[1].upper() != 'VACANT' and t[5] > 0)
    giy = gross_rent / deal.get('asking_price', 1)
    ltv = loan / deal.get('asking_price', 1)
    vacancy = vacant_area / total_area if total_area > 0 else 0
    acq_cost = deal.get('asking_price', 0) * (
        1 + deal.get('grest_rate', 0) +
        deal.get('notar_rate', 0.015) +
        deal.get('broker_rate', 0.010)
    )

    print()
    print("=" * 60)
    print(f"  INTAKE COMPLETE: {deal.get('property_name', '?')}")
    print("=" * 60)
    print(f"  City:          {deal.get('city', '?')}")
    print(f"  Asset type:    {deal.get('asset_type', '?')}")
    print(f"  GLA:           {total_area:,} m²  ({tenant_count} tenants, {vacant_count} vacant)")
    print(f"  Vacancy:       {vacancy:.1%}")
    print()
    print(f"  Asking price:  EUR {deal.get('asking_price', 0):,}")
    print(f"  Total acq cost EUR {acq_cost:,.0f}  (incl GrESt {deal.get('grest_rate',0):.1%})")
    print(f"  Equity:        EUR {deal.get('equity', 0):,}")
    print(f"  Loan:          EUR {loan:,}  (LTV {ltv:.1%})")
    print()
    print(f"  Gross rent:    EUR {gross_rent:,}/yr")
    print(f"  GIY:           {giy:.2%}")
    print()
    print(f"  Output file:   {Path(path).name}")
    print("=" * 60)
    print()
    print("  NEXT STEP: Upload this file to the German Retail Underwriter app")
    print()


# =============================================================================
# RUN
# =============================================================================

if __name__ == '__main__':
    print()
    print("  German Retail Underwriter — Deal Intake")
    print()

    # Pre-flight check
    issues = _quick_check(DEAL, RENT_ROLL, COSTS)
    if issues:
        print("  PRE-FLIGHT ISSUES:")
        for issue in issues:
            print(f"    ⚠️  {issue}")
        print()
        if any(i.startswith("MISSING") for i in issues):
            print("  Stopping — fill in missing fields before generating Excel.")
            exit(1)
        print("  Warnings noted — continuing...")
        print()

    # Print intake notes
    if INTAKE_NOTES.strip():
        print("  INTAKE NOTES:")
        for line in INTAKE_NOTES.strip().split('\n'):
            print(f"  {line}")
        print()

    # Build Excel
    print("  Building Excel file...")
    path, total_area, vacant_area, tenant_count, vacant_count = build_excel(
        DEAL, RENT_ROLL, COSTS
    )

    print_intake_summary(
        DEAL, RENT_ROLL, COSTS, path,
        total_area, vacant_area, tenant_count, vacant_count
    )

    # Quick engine verification
    print("  Running engine verification...")
    import io
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    import rent_roll_parser as rrp

    with open(path, 'rb') as f:
        data = io.BytesIO(f.read())

    deal_parsed = rrp.load_deal_overview(data)
    rr_df       = rrp.load_rent_roll(data)
    act         = rr_df[~rr_df['is_vacant']]
    vac         = rr_df[rr_df['is_vacant']]
    costs_data  = rrp.load_cost_sheet(data,
        rr_gross_rent  = float(act['annual_rent'].sum()),
        rr_gla_total   = float(rr_df['area_sqm'].sum()),
        rr_vacant_area = float(vac['area_sqm'].sum())
    )
    metrics  = rrp.compute_metrics(deal_parsed, rr_df, costs_data)
    qa       = rrp.run_qa_audit(metrics, deal_parsed)
    verdict  = rrp.generate_verdict(metrics, qa)

    print()
    print("  ── QUICK METRICS ──────────────────────────────────")
    print(f"  NOI:      EUR {metrics.get('noi_proxy', 0):>9,.0f}  "
          f"({metrics.get('noi_margin', 0):.0%} margin, source: {metrics.get('noi_source','')})")
    print(f"  GIY:      {metrics.get('giy', 0):.2%}   "
          f"NIY: {metrics.get('niy', 0):.2%}   "
          f"Debt Yield: {metrics.get('debt_yield', 0) or 0:.2%}")
    print(f"  LTV:      {metrics.get('ltv', 0):.1%}   "
          f"DSCR: {metrics.get('dscr', 0):.2f}x   "
          f"ICR: {metrics.get('icr', 0):.2f}x")
    print(f"  WAULT:    {metrics.get('wault', 0):.1f} yrs   "
          f"Vacancy: {metrics.get('vacancy_rate', 0):.1%}   "
          f"Anchor: {metrics.get('anchor_pct', 0):.1%}")
    print(f"  AfA:      EUR {metrics.get('afa_annual', 0) or 0:>9,.0f}/yr")

    fails = [r for r in qa if r.status == 'FAIL']
    warns = [r for r in qa if r.status == 'WARN']
    print()
    print(f"  QA:  {len(fails)} FAILs | {len(warns)} WARNs")
    if fails:
        for r in fails:
            print(f"    ❌ {r.check}: {r.actual}")
    if warns:
        for r in warns:
            print(f"    ⚠️  {r.check}: {r.actual}")

    v_color = {'PASS': '✅', 'INVESTIGATE': '⚠️', 'DECLINE': '❌'}
    print()
    print(f"  {v_color.get(verdict['recommendation'], '')} "
          f"VERDICT: {verdict['recommendation']}")
    print(f"  {verdict.get('rationale', '')}")
    print()
    print(f"  File ready: {Path(path).name}")
    print("  Upload this file to the app to see the full IC analysis.")
    print()
