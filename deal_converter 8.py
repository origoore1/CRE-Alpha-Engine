"""
German Retail Underwriter — Universal Deal Converter
=====================================================
Converts any seller document format into our standard Excel template.

Supported inputs:
  1. Broker Excel files (any column naming, German or English)
  2. Free text / email (paste raw text, extract key figures)
  3. PDF with embedded text (pdfplumber extraction)

Output: Populated DE_Retail_Underwriter_Template.xlsx ready to upload.
"""
import pandas as pd
import numpy as np
import re
import io
from datetime import datetime, date
from pathlib import Path

from constants import GREST_BY_STATE, GREST_DEFAULT, ANCHOR_NAMES

# ─── GERMAN FIELD KEYWORD MAPPING ─────────────────────────────────────────────
FIELD_KEYWORDS = {
    'tenant':       ['mieter', 'tenant', 'nutzer', 'name', 'firma', 'gesellschaft'],
    'sector':       ['nutzart', 'branche', 'sector', 'type', 'nutzung', 'kategorie'],
    'area_sqm':     ['fläche', 'nutzfläche', 'area', 'qm', 'm²', 'sqm', 'größe', 'mietfläche'],
    'rent_sqm_mo':  ['€/m²', 'miete/m', 'kaltmiete', 'monatsmiet', 'nettomiete',
                     'rent/m', 'rent per', 'mietzins'],
    'annual_rent':  ['jahresmiete', 'jahres', 'annual', 'p.a.', 'yearly', 'jahresnetto'],
    'lease_start':  ['beginn', 'mietbeginn', 'start', 'von ', 'ab ', 'laufzeitbeginn'],
    'lease_expiry': ['ende', 'mietende', 'expiry', 'bis ', 'ablauf', 'laufzeitende',
                     'vertragsende'],
    'break_option': ['skr', 'sonder', 'break', 'kündigung', 'option', 'kündigungsrecht'],
    'cpi_clause':       ['index', 'indexiert', 'cpi', 'wertsicherung', 'inflationsanpassung'],
    'cpi_passthrough':  ['cpi pass', 'passthrough', 'pass-through', 'weitergabe',
                         'indexanpassung%', 'cpi%', 'anpassungssatz'],
    'cpi_trigger':      ['cpi trigger', 'trigger', 'auslöser', 'aktivierungsschwelle',
                         'schwellenwert', 'indexschwelle'],
    'security_dep': ['kaution', 'sicherheit', 'deposit', 'bürgschaft'],
    'notes':        ['anmerkung', 'bemerkung', 'note', 'kommentar', 'hinweis'],
}

VACANT_KEYWORDS = ['leerstand', 'leer', 'vacant', 'frei', 'unvermietet', 'vacancy']
# ANCHOR_NAMES and GREST_BY_STATE are imported from constants.py


# ─── UTILITY FUNCTIONS ────────────────────────────────────────────────────────

def extract_float(s):
    """Extract numeric value from any string, handling German/English number formats."""
    if s is None: return 0.0
    s = str(s).strip()
    if s in ['', 'nan', 'None', '—', '-', 'n/a', 'N/A']: return 0.0
    s = re.sub(r'[€$£\s%]', '', s)
    # German thousands: 1.234,56 or English thousands: 1,234.56
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', s):
        # German: 1.234.567,89
        s = s.replace('.', '').replace(',', '.')
    elif re.match(r'^\d{1,3}(,\d{3})+(\.(\d+))?$', s):
        # English: 1,234,567.89
        s = s.replace(',', '')
    elif re.match(r'^\d+,\d{3}$', s):
        # 3,450 → thousands separator (3 digits after comma)
        s = s.replace(',', '')
    else:
        # Decimal comma: 9,80 → 9.80
        s = s.replace(',', '.')
    m = re.search(r'[\d]+\.?[\d]*', s)
    try: return float(m.group()) if m else 0.0
    except: return 0.0


def parse_any_date(s):
    """Parse date from any common format."""
    if not s or str(s).strip() in ['', 'nan', 'None', '—', '-']: return None
    s = str(s).strip()
    formats = ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y',
               '%d.%m.%y', '%B %d, %Y', '%d %B %Y']
    for fmt in formats:
        try: return datetime.strptime(s, fmt).date()
        except: continue
    try: return pd.to_datetime(s).date()
    except: return None


def fuzzy_map_columns(df):
    """Map arbitrary DataFrame column names to our standard field names."""
    mapping = {}
    used_cols = set()
    for field, keywords in FIELD_KEYWORDS.items():
        best_col, best_score = None, 0
        for col in df.columns:
            col_lower = col.lower().replace(' ', '').replace('/', '').replace('€', '')
            score = sum(1 for kw in keywords
                        if kw.replace(' ','').replace('€','') in col_lower)
            # Bonus: exact match
            if col.lower().strip() in [k.strip() for k in keywords]:
                score += 3
            if score > best_score and col not in used_cols:
                best_score = score
                best_col   = col
        if best_col and best_score > 0:
            mapping[field] = best_col
            used_cols.add(best_col)
    return mapping


def infer_anchor(tenant_name):
    """Guess if tenant is an anchor based on known German anchor names."""
    name_lower = str(tenant_name).lower()
    return 'Y' if any(a in name_lower for a in ANCHOR_NAMES) else 'N'


def infer_sector(tenant_name, existing_sector=None):
    """Infer trade sector from tenant name if not provided."""
    if existing_sector and str(existing_sector).strip() not in ['', 'nan', '—', '-']:
        return str(existing_sector).strip()
    name_lower = str(tenant_name).lower()
    if any(x in name_lower for x in ['rewe','edeka','kaufland','aldi','lidl','penny','netto','tegut']):
        return 'Food & Grocery'
    if any(x in name_lower for x in ['dm','rossmann','müller','budni']):
        return 'Health & Beauty'
    if any(x in name_lower for x in ['obi','bauhaus','hornbach','toom']):
        return 'DIY & Home Improvement'
    if any(x in name_lower for x in ['kik','takko','primark','h&m','zara','ernsting']):
        return 'Discount Fashion'
    if any(x in name_lower for x in ['mediamarkt','saturn','expert','cyberport']):
        return 'Electronics'
    if any(x in name_lower for x in ['fressnapf','zoo','tierpark']):
        return 'Pet Supplies'
    if any(x in name_lower for x in ['action','tedi','woolworth']):
        return 'General Merchandise'
    if any(x in name_lower for x in ['lotto','tabak','kiosk','leerstand','vacant']):
        return 'Personal Services'
    return 'Other'


# ─── CONVERTER 1: Excel Rent Roll ─────────────────────────────────────────────

def convert_excel_rent_roll(file_bytes):
    """
    Convert any broker Excel rent roll to our standard format.
    Returns (tenant_list, warnings, confidence_score)
    """
    warnings = []
    tenants  = []

    try:
        # Try reading with headers on first row, then row 1, 2
        for header_row in [0, 1, 2]:
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
                # Check if we got meaningful column names
                non_null_cols = [c for c in df.columns
                                 if not str(c).startswith('Unnamed')]
                if len(non_null_cols) >= 3:
                    break
            except: continue

        # Remove completely empty rows
        df = df.dropna(how='all')

        # Map columns
        col_map = fuzzy_map_columns(df)
        if 'tenant' not in col_map:
            warnings.append("Could not identify tenant name column")
            return [], warnings, 0

        # Count how many key fields we found
        key_fields = ['tenant','area_sqm','rent_sqm_mo','lease_expiry']
        found_keys = sum(1 for f in key_fields if f in col_map)
        confidence = found_keys / len(key_fields)

        if found_keys < 2:
            warnings.append(f"Only {found_keys} key columns detected — low confidence")

        # Extract rows
        for i, row in df.iterrows():
            tenant_name = str(row.get(col_map.get('tenant',''), '')).strip()
            if not tenant_name or tenant_name in ['nan', 'None', '']: continue

            is_vacant = (any(kw in tenant_name.lower() for kw in VACANT_KEYWORDS)
                 or tenant_name.strip() in ['—', '-', ''])

            area     = extract_float(row.get(col_map.get('area_sqm',''), 0))
            rent     = extract_float(row.get(col_map.get('rent_sqm_mo',''), 0))

            # If no monthly rent but annual rent exists, back-calculate
            if rent == 0 and area > 0 and 'annual_rent' in col_map:
                annual = extract_float(row.get(col_map['annual_rent'], 0))
                if annual > 0 and area > 0:
                    rent = annual / (area * 12)

            if is_vacant: rent = 0

            start  = parse_any_date(row.get(col_map.get('lease_start',''), ''))
            expiry = parse_any_date(row.get(col_map.get('lease_expiry',''), ''))
            brk    = parse_any_date(row.get(col_map.get('break_option',''), ''))

            cpi_raw = str(row.get(col_map.get('cpi_clause',''), '')).lower()
            cpi = 'Y' if any(x in cpi_raw for x in ['ja','yes','y','1','true','x']) else 'N'

            # CPI passthrough — read from source column when present; normalise 75→0.75;
            # fall back to 0.75 default only when field is genuinely absent or zero.
            def _read_pct_col(col_key, fallback):
                raw = row.get(col_map.get(col_key, ''), None) if col_key in col_map else None
                val = extract_float(raw) if raw is not None else None
                if val is not None and val > 0:
                    return min(1.0, val / 100 if val > 1 else val)
                return fallback

            cpi_passthrough = _read_pct_col('cpi_passthrough', 0.75 if cpi == 'Y' else 0.0)
            cpi_trigger     = _read_pct_col('cpi_trigger',     0.10 if cpi == 'Y' else 0.0)

            sector  = infer_sector(tenant_name, row.get(col_map.get('sector',''), ''))
            anchor  = infer_anchor(tenant_name)
            sec_dep = extract_float(row.get(col_map.get('security_dep',''), 0))
            if sec_dep == 0 and rent > 0 and area > 0:
                sec_dep = round(area * rent * 3)  # Standard: 3 months
            notes = str(row.get(col_map.get('notes',''), '')).strip()
            if notes == 'nan': notes = ''

            if area > 0:
                tenants.append({
                    'tenant_name': tenant_name if not is_vacant else 'VACANT',
                    'sector':      '—' if is_vacant else sector,
                    'anchor':      'N' if is_vacant else anchor,
                    'area_sqm':    area,
                    'rent_sqm_mo': rent,
                    'lease_start': start,
                    'lease_expiry':expiry,
                    'break_option':brk,
                    'cpi_clause':  'N' if is_vacant else cpi,
                    'cpi_passthrough': 0 if is_vacant else cpi_passthrough,
                    'cpi_trigger':     0 if is_vacant else cpi_trigger,
                    'security_dep':sec_dep,
                    'svc_recoverable': 'N' if is_vacant else 'Y',
                    'rent_free':   0,
                    'notes':       notes,
                    'is_vacant':   is_vacant,
                    '_col_map':    col_map,
                })

        warnings.append(f"Detected {found_keys}/4 key columns")
        if 'area_sqm'    not in col_map: warnings.append("⚠️ Area column not detected — check manually")
        if 'rent_sqm_mo' not in col_map: warnings.append("⚠️ Rent/m² column not detected — check manually")
        if 'lease_expiry'not in col_map: warnings.append("⚠️ Lease expiry not detected — check manually")

        return tenants, warnings, confidence

    except Exception as e:
        return [], [f"Excel parse error: {e}"], 0


# ─── CONVERTER 2: Free Text / Email Parser ────────────────────────────────────

def parse_free_text(text):
    """
    Extract deal data from unstructured text (broker email, notes, exposé copy).
    Returns (deal_dict, tenant_list, warnings, confidence)
    """
    deal     = {}
    tenants  = []
    warnings = []
    found    = 0
    total    = 8  # key fields we try to find

    # ── Asking price ──
    m = re.search(r'(?:asking|preis|kaufpreis|price)[^\d€]*(?:EUR|€|eur)?\s*'
                  r'([\d][,\.\d]+)\s*(?:mio|million|mrd)?', text, re.I)
    if m:
        val = extract_float(m.group(1))
        if val < 1000: val *= 1_000_000  # "8.75 Mio"
        deal['asking_price'] = int(val); found += 1

    # ── GLA ──
    m = re.search(r'(?:GLA|mietfläche|nutzfläche|gesamt)[^\d]*([\d][,\.\d]*)\s*m²', text, re.I)
    if m: deal['gla_sqm'] = int(extract_float(m.group(1))); found += 1

    # ── Parking ──
    m = re.search(r'(?:parking|stellplätze|parkplätze|pkw)[^\d]*([\d]+)', text, re.I)
    if m: deal['parking'] = int(m.group(1)); found += 1

    # ── Year built ──
    m = re.search(r'(?:built|baujahr|errichtet|erbaut)[^\d]*((?:19|20)\d{2})', text, re.I)
    if m: deal['year_built'] = int(m.group(1)); found += 1

    # ── City / Bundesland ──
    for state in GREST_BY_STATE.keys():
        if state.lower() in text.lower():
            m_city = re.search(r'([A-ZÄÖÜ][a-zäöüß\-]+(?:\s[A-ZÄÖÜ][a-zäöüß\-]+)?)'
                               r'[,\s]*' + state, text, re.I)
            city_name = m_city.group(1).strip() if m_city else 'Unknown'
            deal['city']      = f"{city_name} / {state}"
            deal['grest_rate']= GREST_BY_STATE[state]
            found += 1
            break

    # ── WAULT stated ──
    m = re.search(r'WAULT[^\d]*([\d\.]+)\s*(?:years?|jahre?|yr)', text, re.I)
    if m: deal['wault_stated'] = float(m.group(1)); found += 1

    # ── Gross rent stated ──
    m = re.search(r'(?:gross rent|jahresmiete|jahresnettomiete)[^\d€]*(?:EUR|€|ca\.?)?\s*'
                  r'([\d][,\.\d]*)', text, re.I)
    if m:
        val = extract_float(m.group(1))
        deal['gross_rent_stated'] = int(val); found += 1

    # ── GIY stated ──
    m = re.search(r'(?:GIY|rendite|yield|anfangsrendite)[^\d]*([\d\.]+)\s*%', text, re.I)
    if m:
        deal['giy_stated'] = float(m.group(1)) / 100
        # Back-calculate asking price if not found
        if 'asking_price' not in deal and 'gross_rent_stated' in deal:
            deal['asking_price'] = int(deal['gross_rent_stated'] / deal['giy_stated'])
        found += 1

    # ── Tenant lines parser ──
    # Looks for patterns like: "Rewe: 2,000m², €11.50/m²/mo, expires 31.03.2034"
    tenant_patterns = [
        # "- Rewe Markt: 2000m², €11.50/m²/mo, expires 2033"
        r'[-•*]\s*([A-Za-zÄÖÜäöüß&\s\.]+?):\s*([\d,\.]+)\s*m²[,\s]*(?:€|EUR|eur)?\s*([\d,\.]+)',
        # "- Kaufland  2,100m²  €9.80"
        r'[-•*]\s*([A-Za-zÄÖÜäöüß&\s\.]+?)\s+([\d][\d,\.]+)\s*m²[,\s]*(?:€|EUR|eur)?([\d,\.]+)',
        # "Kaufland: 2100m², 9.80 EUR"  (no bullet)
        r'^([A-ZÄÖÜ][A-Za-zÄÖÜäöüß&\s\.]{3,}):\s*([\d,\.]+)\s*m²[,\s]*(?:€|EUR)?\s*([\d,\.]+)',
    ]

    for pattern in tenant_patterns:
        for m in re.finditer(pattern, text, re.I):
            name  = m.group(1).strip()
            area  = extract_float(m.group(2))
            rent  = extract_float(m.group(3))
            if area > 0 and rent > 0 and len(name) > 2:
                # Find expiry date near this match
                context = text[m.start():m.start()+200]
                exp_m = re.search(r'(\d{2}[\./]\d{2}[\./]\d{2,4})', context)
                expiry = parse_any_date(exp_m.group(1)) if exp_m else None
                brk_m  = re.search(r'break[^\d]*([\d]{2}[\./]\d{2}[\./]\d{4})', context, re.I)
                brk    = parse_any_date(brk_m.group(1)) if brk_m else None
                cpi    = 'Y' if any(x in context.lower() for x in ['cpi','index','yes','ja']) else 'N'
                # Try to read explicit passthrough % from context (e.g. "75% CPI", "zu 100% angepasst")
                _pt_m = re.search(r'(\d{2,3})\s*%\s*(?:cpi|index|pass|anpass|weitergabe)',
                                  context, re.I)
                if _pt_m:
                    cpi_pt = min(1.0, int(_pt_m.group(1)) / 100)
                else:
                    cpi_pt = 0.75 if cpi == 'Y' else 0.0
                tenants.append({
                    'tenant_name':   name,
                    'sector':        infer_sector(name),
                    'anchor':        infer_anchor(name),
                    'area_sqm':      area,
                    'rent_sqm_mo':   rent,
                    'lease_start':   None,
                    'lease_expiry':  expiry,
                    'break_option':  brk,
                    'cpi_clause':    cpi,
                    'cpi_passthrough': cpi_pt,
                    'cpi_trigger':   0.10 if cpi == 'Y' else 0,
                    'security_dep':  round(area * rent * 3),
                    'svc_recoverable':'Y',
                    'rent_free':     0,
                    'notes':         'Extracted from email/text',
                    'is_vacant':     False,
                })
        if tenants: break

    # Vacant units
    for m in re.finditer(r'(?:vacant|leerstand|leer)[^\d]*([\d,\.]+)\s*m²', text, re.I):
        area = extract_float(m.group(1))
        if area > 0:
            tenants.append({
                'tenant_name':'VACANT','sector':'—','anchor':'N',
                'area_sqm':area,'rent_sqm_mo':0,'lease_start':None,
                'lease_expiry':None,'break_option':None,'cpi_clause':'N',
                'cpi_passthrough':0,'cpi_trigger':0,'security_dep':0,
                'svc_recoverable':'N','rent_free':0,'notes':'Vacant unit from text',
                'is_vacant':True,
            })

    confidence = found / total
    if found < 4:
        warnings.append(f"Low confidence: only {found}/{total} deal fields detected from text")
    if not tenants:
        warnings.append("No tenant data detected — enter rent roll manually or attach Excel")
    else:
        warnings.append(f"Detected {len(tenants)} tenant(s) from text — verify each row")

    return deal, tenants, warnings, confidence


# ─── GENERATE EXCEL FROM CONVERTED DATA ──────────────────────────────────────

def generate_excel_from_conversion(deal_dict, tenant_list, default_costs=None):
    """
    Write converted data into our standard Excel template.
    Returns BytesIO of the populated Excel file.
    """
    import shutil
    from openpyxl import load_workbook
    from deal_simulator import DEAL_CELLS, RR_COLS, COST_ROWS, RR_FIRST_DATA_ROW, _clear_rent_roll, MASTER_TEMPLATE

    out = io.BytesIO()
    # Load template
    with open(MASTER_TEMPLATE, 'rb') as f:
        wb = load_workbook(f)

    ws_deal = wb['Deal Overview']
    for field, (row, col) in DEAL_CELLS.items():
        val = deal_dict.get(field)
        if val is not None:
            ws_deal.cell(row=row, column=col).value = val

    ws_rr = wb['Rent Roll']
    _clear_rent_roll(ws_rr)
    unit_num = 1
    for t in tenant_list:
        row = RR_FIRST_DATA_ROW + (unit_num - 1)
        ws_rr.cell(row=row, column=RR_COLS['unit_ref']).value          = f"U-{unit_num:02d}"
        ws_rr.cell(row=row, column=RR_COLS['tenant']).value            = t.get('tenant_name','')
        ws_rr.cell(row=row, column=RR_COLS['sector']).value            = t.get('sector','')
        ws_rr.cell(row=row, column=RR_COLS['anchor']).value            = t.get('anchor','N')
        ws_rr.cell(row=row, column=RR_COLS['area_sqm']).value          = t.get('area_sqm',0)
        ws_rr.cell(row=row, column=RR_COLS['rent_sqm_mo']).value       = t.get('rent_sqm_mo',0)
        ws_rr.cell(row=row, column=RR_COLS['lease_start']).value       = t.get('lease_start')
        ws_rr.cell(row=row, column=RR_COLS['lease_expiry']).value      = t.get('lease_expiry')
        ws_rr.cell(row=row, column=RR_COLS['break_option']).value      = t.get('break_option')
        ws_rr.cell(row=row, column=RR_COLS['cpi_clause']).value        = t.get('cpi_clause','N')
        ws_rr.cell(row=row, column=RR_COLS['cpi_passthrough']).value   = t.get('cpi_passthrough',0)
        ws_rr.cell(row=row, column=RR_COLS['cpi_trigger']).value       = t.get('cpi_trigger',0)
        ws_rr.cell(row=row, column=RR_COLS['security_deposit']).value  = t.get('security_dep',0)
        ws_rr.cell(row=row, column=RR_COLS['svc_chg_recov']).value     = t.get('svc_recoverable','Y')
        ws_rr.cell(row=row, column=RR_COLS['rent_free_months']).value  = t.get('rent_free',0)
        ws_rr.cell(row=row, column=RR_COLS['notes']).value             = t.get('notes','')
        unit_num += 1

    # Default costs if none provided
    costs = default_costs or {
        'property_mgmt_pct':0.015,'asset_mgmt_pct':0.005,
        'insurance_sqm':2.50,'capex_sqm':5.00,'ground_rent_abs':0,
        'grundsteuer_abs':0,'void_svc_sqm':20.0,'void_ins_sqm':1.75,
        'letting_pct':0.050,'other_abs':2000,
    }
    ws_cost = wb['Cost Sheet']
    for cost_key, row_num in COST_ROWS.items():
        ws_cost.cell(row=row_num, column=3).value = costs.get(cost_key, 0)

    wb.save(out)
    out.seek(0)
    return out


if __name__ == '__main__':
    print("deal_converter.py loaded — conversion functions available")
    print("Use via app.py Import tab or directly:")
    print("  from deal_converter import convert_excel_rent_roll, parse_free_text")
